import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

folder = r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu"
files = [f for f in os.listdir(folder) if f.endswith(".xlsx")]

for f in files:
    path = os.path.join(folder, f)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        # Find candidate sheet
        candidates = [s for s in wb.sheetnames if s not in ['_com.sap.ip.bi.xl.hiddensheet', 'RC']]
        sheet = wb[candidates[0]]
        
        rsms = set()
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i < 2:
                continue
            rsm_val = row[1] if len(row) > 1 else None
            if rsm_val:
                rsms.add(rsm_val)
            if i > 1000: # Just check the first 1000 rows to see if there is any mixture
                break
        print(f"{f}: {rsms}")
    except Exception as e:
        print(f"Error {f}: {e}")
