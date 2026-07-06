# -*- coding: utf-8 -*-
"""
TOOL ĐỔ DỮ LIỆU PHÂN TÍCH TỒN KHO
Phiên bản: 1.1.0
Mô tả: Tự động tính toán các chỉ số tồn kho, tra cứu hạn sử dụng tham khảo
       và phân loại sản phẩm cần kiểm tra theo đúng công thức nghiệp vụ.
       Tối ưu hóa hiệu năng xử lý tốc độ cao (One-pass stream).
"""

import os
import sys
import time
import zipfile
import xml.etree.ElementTree as ET
import ctypes
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, date
from threading import Thread
import tempfile
import shutil

# Force UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

# ===================================================================
# CẤU HÌNH MẶC ĐỊNH
# ===================================================================
# Hỗ trợ chạy cả dạng script thường và dạng đóng gói (.exe)
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DEFAULT_XLSB_PATH = os.path.join(BASE_DIR, "raw mẫu", "updated_data_moi.xlsb")
APP_VERSION = "1.2.0"
APP_TITLE = "Tool Đổ Dữ Liệu Phân Tích Tồn Kho"
SCRATCH_DIR = tempfile.mkdtemp(prefix="tool_do_du_lieu_")


# ===================================================================
# HÀM TRỢ GIÚP
# ===================================================================
def col_to_idx(col_letter: str) -> int:
    """Chuyển đổi tên cột Excel (A, B, ..., AA, AB) sang index 0-based."""
    val = 0
    for char in col_letter.upper():
        val = val * 26 + (ord(char) - 64)
    return val - 1


def get_column_letter(col_idx: int) -> str:
    """Chuyển đổi index 1-based sang chữ cái cột Excel (ví dụ: 1->A, 27->AA)."""
    result = []
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result.append(chr(65 + remainder))
    return ''.join(reversed(result))


def get_actual_col(col_letter: str, offset: int) -> str:
    """Trả về ký tự cột thực tế dựa trên offset của cột RSM.
    
    col_letter: Ký tự cột chuẩn của file gốc (ví dụ: 'K', 'W', 'X')
    offset: 1 nếu RSM ở B (file chuẩn), 0 nếu RSM ở A (file khuyết cột A)
    """
    if offset == 1:
        return col_letter
    idx = col_to_idx(col_letter)
    actual_idx = idx - (1 - offset)
    if actual_idx < 0:
        return ''
    return get_column_letter(actual_idx + 1)


def safe_float(val) -> float | None:
    """Chuyển đổi giá trị sang float an toàn."""
    if val is None or val == '':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def excel_date_to_datetime(val) -> datetime | None:
    """Chuyển đổi số serial Excel hoặc chuỗi ngày sang đối tượng datetime."""
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        # Excel bug: 1900 được coi là năm nhuận, nên mốc bắt đầu thực tế là 1899-12-30
        try:
            return datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(val))
        except (ValueError, OverflowError):
            return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    
    # Thử parse định dạng chuỗi
    val_str = str(val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d.%m.%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(val_str, fmt)
        except ValueError:
            continue
    return None


# ===================================================================
# THAO TÁC FILE XLSB (TRA CỨU HẠN SỬ DỤNG)
# ===================================================================
def load_shelf_life_map(xlsb_path: str, log_func=None) -> dict:
    """
    Đọc bảng tra cứu hạn sử dụng từ file XLSB sử dụng pyxlsb.
    Trả về dict: {ma_san_pham: han_su_dung_tham_khao}
    """
    shelf_life_map = {}
    if not os.path.exists(xlsb_path):
        if log_func:
            log_func(f"  [CẢNH BÁO] Không tìm thấy file tra cứu tại: {xlsb_path}")
        return shelf_life_map

    if log_func:
        log_func(f"  Đang đọc bảng tra cứu từ: {os.path.basename(xlsb_path)}")

    t0 = time.time()
    try:
        from pyxlsb import open_workbook
        with open_workbook(xlsb_path) as wb:
            # Tìm sheet 'master article' hoặc sheet có tên tương tự
            sheet_names = wb.sheets
            target_sheet = None
            for name in sheet_names:
                if 'master' in name.lower() or 'article' in name.lower():
                    target_sheet = name
                    break
            if not target_sheet and sheet_names:
                target_sheet = sheet_names[0]

            if not target_sheet:
                if log_func:
                    log_func("  ✗ File XLSB không chứa sheet nào.")
                return shelf_life_map

            with wb.get_sheet(target_sheet) as sheet:
                # Tìm index của các cột 'Article' và 'Date tham khảo' ở dòng header
                header_row = None
                for row in sheet.rows():
                    header_row = row
                    break

                if not header_row:
                    return shelf_life_map

                col_art_idx = -1
                col_date_idx = -1

                for idx, cell in enumerate(header_row):
                    val = str(cell.v).strip().lower() if cell.v is not None else ''
                    if 'article' in val:
                        col_art_idx = idx
                    elif 'date' in val and 'tham' in val:
                        col_date_idx = idx

                # Nếu không tìm thấy cột chuẩn, sử dụng index mặc định (A=0, Z=25)
                if col_art_idx == -1:
                    col_art_idx = 0
                if col_date_idx == -1:
                    col_date_idx = 25

                # Duyệt qua các hàng dữ liệu
                count_valid = 0
                is_first = True
                for row in sheet.rows():
                    if is_first:
                        is_first = False
                        continue
                    
                    if len(row) <= max(col_art_idx, col_date_idx):
                        continue

                    art_cell = row[col_art_idx].v
                    date_cell = row[col_date_idx].v

                    if art_cell is not None:
                        # Chuẩn hóa mã sản phẩm thành chuỗi số nguyên
                        art_str = str(art_cell).strip()
                        if art_str.endswith('.0'):
                            art_str = art_str[:-2]
                        
                        # Đọc hạn sử dụng
                        shelf_life = None
                        if date_cell is not None:
                            try:
                                shelf_life = int(float(date_cell))
                            except (ValueError, TypeError):
                                pass

                        shelf_life_map[art_str] = shelf_life
                        if shelf_life is not None:
                            count_valid += 1

        t1 = time.time()
        if log_func:
            log_func(f"  ✓ Đọc xong bảng tra cứu: {len(shelf_life_map):,} SP, trong đó {count_valid:,} SP có date tham khảo hợp lệ")
            log_func(f"  → Thời gian nạp bảng tra cứu: {t1 - t0:.1f}s")
    except Exception as e:
        if log_func:
            log_func(f"  ✗ Lỗi nạp XLSB: {e}")
    return shelf_life_map


# ===================================================================
# LOGIC TÍNH TOÁN NGHIỆP VỤ (FORMULAS)
# ===================================================================
def compute_row(row_dict: dict, milestone_date: datetime, ref_amount: float,
                shelf_life_map: dict) -> dict:
    """
    Tính toán các chỉ số cho một dòng dữ liệu (Row 3+).
    Tham số truyền vào row_dict đã được chuẩn hóa về các cột chuẩn (B->X).
    """
    result = {}
    
    art_code_raw = row_dict.get('K')
    if art_code_raw is None:
        art_code_raw = ''
    art_code_str = str(art_code_raw).strip()
    if art_code_str.endswith('.0'):
        art_code_str = art_code_str[:-2]

    # Trích xuất các giá trị số
    ton_dau_ky = safe_float(row_dict.get('N')) or 0
    nhap_trong_ky = safe_float(row_dict.get('O')) or 0
    xuat_trong_ky = safe_float(row_dict.get('P')) or 0
    ton_cuoi_ky = safe_float(row_dict.get('Q')) or 0
    gia_tri_ton = safe_float(row_dict.get('R')) or 0
    doanh_thu = safe_float(row_dict.get('S')) or 0
    gia_von = safe_float(row_dict.get('T')) or 0
    ton_d15 = safe_float(row_dict.get('U')) or 0
    gia_ton_d15 = safe_float(row_dict.get('V')) or 0

    # ================================================================
    # Đọc Last GR (W) và Last Sale (X) - có thể là số serial hoặc ">90"
    # Khi giá trị > 90 ngày trước ngày mốc, SAP xuất ">90" dạng text
    # ================================================================
    last_gr_raw = row_dict.get('W')
    last_sale_raw = row_dict.get('X')

    # Xác định xem Last GR và Last Sale có phải ">90" không
    last_gr_gt90 = (str(last_gr_raw).strip() == '>90') if last_gr_raw is not None else False
    last_sale_gt90 = (str(last_sale_raw).strip() == '>90') if last_sale_raw is not None else False

    last_gr = None if last_gr_gt90 else excel_date_to_datetime(last_gr_raw)
    last_sale = None if last_sale_gt90 else excel_date_to_datetime(last_sale_raw)

    # --- Col Y: Số ngày không nhập ---
    # Chỉnh sửa: số ngày không nhập trừ quá 90 hoặc không có giá trị số thì điền ">90"
    if last_gr_gt90 or last_gr is None or milestone_date is None:
        result['Y'] = '>90'
        y_days = 91               # Dùng 91 để so sánh logic nội bộ (>90)
    else:
        delta = milestone_date - last_gr
        diff_days = delta.days
        if diff_days > 90:
            result['Y'] = '>90'
            y_days = 91
        else:
            result['Y'] = diff_days
            y_days = diff_days

    # --- Col Z: Date tham khảo = tra cứu từ XLSB theo mã SP ---
    z_shelf_life = shelf_life_map.get(art_code_str)
    if z_shelf_life is None:
        z_shelf_life = safe_float(row_dict.get('Z'))
    result['Z'] = z_shelf_life

    # --- Col AA: DIO (D-15) ---
    # Công thức: IF(T3=0,9999,V3/T3*90)
    if gia_von == 0:
        aa_val = 9999  # Giống công thức Excel
    else:
        aa_val = gia_ton_d15 / gia_von * 90
    result['AA'] = aa_val

    # --- Col AB: DIO/Date = AA / Z ---
    if aa_val is not None and z_shelf_life is not None and z_shelf_life > 0:
        result['AB'] = aa_val / z_shelf_life
    else:
        result['AB'] = None
    ab_val = result['AB']

    # --- Col AC: Note ---
    # Công thức: IF(X3=">90", IF(W3=">90","không giao dịch >90 ngày","Không sale 90 ngày"), "")
    # → Last Sale ">90": không có bán trong 90 ngày
    #   - Nếu Last GR cũng ">90": "không giao dịch >90 ngày" (Nghi vấn tồn ảo)
    #   - Nếu Last GR không phải ">90": "Không sale 90 ngày" (Non-moving)
    # → Last Sale không phải ">90": AC = rỗng (có bán gần đây)
    la_nghi_van = False
    la_non_moving = False
    ac_note = ''

    if last_sale_gt90:
        if last_gr_gt90:
            ac_note = 'không giao dịch >90 ngày'
            la_nghi_van = True
        else:
            ac_note = 'Không sale 90 ngày'
            la_non_moving = True
    result['AC'] = ac_note if ac_note else None

    # --- Col AD: Note check slow ---
    # Công thức: IF(AND(AB3>5,R3>$AF$1),"Check","")
    if ab_val is not None and ab_val > 5 and gia_tri_ton > ref_amount:
        result['AD'] = 'Check'
    else:
        result['AD'] = None

    # --- Col AE: note hết hạn ---
    # Công thức: IF(Y3=">90", IF(Z3<90,"Hết hạn",""), IF(Y3>Z3,"Hết hạn",""))
    # → Nếu Y = ">90" (Last GR không rõ/ảo): hết hạn nếu Date tham khảo < 90 ngày
    # → Nếu Y là số ngày: hết hạn nếu Số ngày ko nhập > Date tham khảo
    ae_note = ''
    if z_shelf_life is not None:
        if last_gr_gt90 or y_days == 91:  # Y = ">90"
            if z_shelf_life < 90:
                ae_note = 'Hết hạn'
        elif y_days is not None:
            if y_days > z_shelf_life:
                ae_note = 'Hết hạn'
    result['AE'] = ae_note if ae_note else None

    # --- Col AF: Phân loại kiểm tra ---
    # Công thức: IF(AE3<>"","Hết hạn",
    #               IF(AC3="Không giao dịch >90 ngày","Nghi vấn tồn ảo",
    #                  IF(AC3="Không sale 90 ngày","Non-moving",
    #                     IF(AD3<>"","Slow moving",""))))
    phan_loai = None

    if ae_note:
        phan_loai = 'Hết hạn'
    elif la_nghi_van:
        phan_loai = 'Nghi vấn tồn ảo'
    elif la_non_moving:
        phan_loai = 'Non-moving'
    elif result['AD'] == 'Check':
        phan_loai = 'Slow moving'

    result['AF'] = phan_loai
    return result


# ===================================================================
# XỬ LÝ CHÍNH EXCEL XML STREAM (ONE-PASS OPTIMIZED)
# ===================================================================
def process_excel_file(src_path: str, shelf_life_map: dict, milestone_date: datetime,
                       ref_amount: float, log_func=None) -> str | None:
    """
    Xử lý file dữ liệu Excel bằng iterparse (one-pass) để tối ưu hóa tốc độ.
    Lọc và giữ lại các hàng có Phân loại kiểm tra.
    Trả về đường dẫn file output hoặc None nếu lỗi.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.cell import WriteOnlyCell

    filename = os.path.basename(src_path)
    base, ext = os.path.splitext(filename)
    out_filename = f"{base} (done){ext}"
    out_path = os.path.join(os.path.dirname(src_path), out_filename)

    if log_func:
        log_func(f"\n[{filename}] Bắt đầu xử lý...")

    t0 = time.time()
    temp_path = os.path.join(SCRATCH_DIR, f"temp_{filename}")

    try:
        # Sao chép file để tránh lock
        res = ctypes.windll.kernel32.CopyFileW(src_path, temp_path, False)
        if not res:
            if log_func:
                log_func(f"  ✗ Không thể copy file (đang mở?): {filename}")
            return None

        # Định nghĩa styles
        font_white_bold = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        font_normal = Font(name='Arial', size=9)
        fill_header = PatternFill(fill_type='solid', start_color='1F4E78', end_color='1F4E78')
        
        fill_het_han = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
        fill_nghi_van = PatternFill(fill_type='solid', start_color='FF6600', end_color='FF6600')
        fill_non_moving = PatternFill(fill_type='solid', start_color='FFA500', end_color='FFA500')
        fill_slow_moving = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
        font_slow = Font(name='Arial', size=9, color='000000')
        
        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')

        # Màu cho từng phân loại
        phan_loai_style = {
            'Hết hạn': (fill_het_han, Font(name='Arial', size=9, bold=True, color='FFFFFF')),
            'Nghi vấn tồn ảo': (fill_nghi_van, Font(name='Arial', size=9, bold=True, color='FFFFFF')),
            'Non-moving': (fill_non_moving, Font(name='Arial', size=9, bold=True, color='000000')),
            'Slow moving': (fill_slow_moving, font_slow),
        }

        wb_out = openpyxl.Workbook(write_only=True)
        ws_out = wb_out.create_sheet(title='Dữ liệu kiểm tra')

        rows_written = 0
        rows_filtered = 0
        total_rows_read = 0

        with zipfile.ZipFile(temp_path, 'r') as z:
            # Tìm sheet dữ liệu chính (bỏ qua hidden và RC)
            wb_data = z.read('xl/workbook.xml')
            root_wb = ET.fromstring(wb_data)
            ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

            sheets_info = []
            for sheet_elem in root_wb.findall('.//ns:sheet', ns):
                sheets_info.append((
                    sheet_elem.attrib.get('name'),
                    sheet_elem.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                ))

            data_sheets = [s for s in sheets_info
                           if s[0] not in ['_com.sap.ip.bi.xl.hiddensheet', 'RC']]
            if not data_sheets:
                if log_func:
                    log_func(f"  ✗ Không tìm thấy sheet dữ liệu trong: {filename}")
                return None

            _, target_rid = data_sheets[0]

            # Tìm đường dẫn file worksheet
            rels_data = z.read('xl/_rels/workbook.xml.rels')
            root_rels = ET.fromstring(rels_data)
            r_ns = {'rels': 'http://schemas.openxmlformats.org/package/2006/relationships'}
            sheet_path_in_zip = None
            for rel in root_rels.findall('.//rels:Relationship', r_ns):
                if rel.attrib.get('Id') == target_rid:
                    sheet_path_in_zip = f"xl/{rel.attrib.get('Target')}"
                    break

            if not sheet_path_in_zip:
                if log_func:
                    log_func(f"  ✗ Không tìm thấy path worksheet: {filename}")
                return None

            # Load shared strings
            shared_strings = []
            if 'xl/sharedStrings.xml' in z.namelist():
                sst_file = z.open('xl/sharedStrings.xml')
                context_sst = ET.iterparse(sst_file, events=('start', 'end'))
                for event, elem in context_sst:
                    if event == 'end' and elem.tag.endswith('}si'):
                        text = "".join([
                            t.text for t in elem.findall(
                                './/{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')
                            if t.text is not None
                        ])
                        shared_strings.append(text)
                        elem.clear()

            # Tối ưu hóa hàm convert số thực
            def get_cell_value(cell_elem):
                t_type = cell_elem.attrib.get('t')
                v_elem = cell_elem.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                if v_elem is None:
                    return None
                val_text = v_elem.text
                if val_text is None:
                    return None
                if t_type == 's':
                    idx = int(val_text)
                    return shared_strings[idx] if idx < len(shared_strings) else ''
                try:
                    if '.' in val_text:
                        return float(val_text)
                    return int(val_text)
                except ValueError:
                    return val_text

            # Stream parse worksheet XML
            sheet_file = z.open(sheet_path_in_zip)
            context = ET.iterparse(sheet_file, events=('start', 'end'))

            current_row_data = {}
            current_row_idx = None
            offset = 1  # Mặc định lệch 1 cột (có cột map)

            for event, elem in context:
                if event == 'start' and elem.tag.endswith('}row'):
                    current_row_data = {}
                    r_attr = elem.attrib.get('r')
                    current_row_idx = int(r_attr) if r_attr and r_attr.isdigit() else None

                elif event == 'end' and elem.tag.endswith('}c'):
                    r_ref = elem.attrib.get('r')
                    if r_ref:
                        col_end = 0
                        for ch in r_ref:
                            if ch.isdigit():
                                break
                            col_end += 1
                        current_row_data[r_ref[:col_end]] = get_cell_value(elem)

                elif event == 'end' and elem.tag.endswith('}row'):
                    if current_row_idx is None:
                        elem.clear()
                        continue

                    total_rows_read += 1

                    if current_row_idx == 1:
                        # Row 1: Nhãn tham chiếu
                        row_list = [None] * 31
                        labels_row1 = {
                            'N': 'Opening Quantity', 'O': 'GR Quantity', 'P': 'GI Quantity',
                            'Q': 'Closing Quantity', 'R': 'Closing Value', 'S': 'Revenue',
                            'T': 'COGS', 'U': 'D-15'
                        }
                        for col_l, label in labels_row1.items():
                            row_list[col_to_idx(col_l) - 1] = label
                        row_list[col_to_idx('Y') - 1] = milestone_date
                        row_list[col_to_idx('AF') - 1] = ref_amount

                        cells_out = []
                        for val in row_list:
                            cell = WriteOnlyCell(ws_out, value=val)
                            cell.font = font_white_bold
                            cell.fill = fill_header
                            cell.alignment = align_center
                            cell.border = border_thin
                            cells_out.append(cell)
                        ws_out.append(cells_out)

                    elif current_row_idx == 2:
                        # Row 2: Header + Xác định offset
                        offset = 1
                        for col_l, val in current_row_data.items():
                            if str(val).strip() == 'RSM':
                                offset = col_to_idx(col_l)
                                break
                        if log_func:
                            log_func(f"  → Dò tìm cấu trúc cột: offset = {offset} (cột {'B' if offset == 1 else 'A'})")

                        row_list = [None] * 31
                        headers = {
                            'B': 'RSM', 'C': 'ASM', 'D': 'Store', 'E': 'Store',
                            'F': 'MCH2 - Department', 'G': 'MCH2 - Department',
                            'H': 'MCH3 - Category', 'I': 'MCH3 - Category',
                            'J': 'MCH4', 'K': 'Article', 'L': 'Article',
                            'M': 'Created on', 'N': 'Tồn đầu kỳ (D-90)',
                            'O': 'Nhập trong kỳ', 'P': 'Xuất trong kỳ',
                            'Q': 'Tồn cuối kỳ (D)', 'R': 'Giá trị tồn',
                            'S': 'Doanh thu', 'T': 'Giá vốn', 'U': 'Tồn D-15',
                            'V': 'Giá tồn D-15', 'W': 'Last GR', 'X': 'Last Sale',
                            'Y': 'Số ngày ko nhập', 'Z': 'Date tham khảo',
                            'AA': 'DIO (D-15)', 'AB': 'DIO/Date', 'AC': 'Note',
                            'AD': 'Note check slow', 'AE': 'note hết hạn',
                            'AF': 'Phân loại kiểm tra'
                        }
                        for col_l, label in headers.items():
                            row_list[col_to_idx(col_l) - 1] = label

                        cells_out = []
                        for val in row_list:
                            cell = WriteOnlyCell(ws_out, value=val)
                            cell.font = font_white_bold
                            cell.fill = fill_header
                            cell.alignment = align_center
                            cell.border = border_thin
                            cells_out.append(cell)
                        ws_out.append(cells_out)

                    else:
                        # Row 3+: Dữ liệu
                        actual_d = get_actual_col('D', offset)
                        actual_e = get_actual_col('E', offset)
                        store_code = str(current_row_data.get(actual_d, '')).strip()
                        store_name = str(current_row_data.get(actual_e, '')).strip()

                        if 'result' in store_code.lower() or 'result' in store_name.lower():
                            rows_filtered += 1
                            elem.clear()
                            continue

                        # Chuẩn hóa (nạp thêm Z cũ từ dữ liệu gốc)
                        norm_row = {}
                        cols_to_norm = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                                        'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Z']
                        for col_std in cols_to_norm:
                            col_act = get_actual_col(col_std, offset)
                            norm_row[col_std] = current_row_data.get(col_act)

                        computed = compute_row(norm_row, milestone_date, ref_amount, shelf_life_map)
                        phan_loai = computed.get('AF')

                        if phan_loai is None or str(phan_loai).strip() == '':
                            rows_filtered += 1
                            elem.clear()
                            continue

                        # Build row_list
                        row_list = [None] * 31
                        # Sao chép các cột từ B đến X (trừ Z cũ vì Z sẽ được cập nhật từ computed)
                        for col_l in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                                      'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']:
                            row_list[col_to_idx(col_l) - 1] = norm_row.get(col_l)

                        for col_l, key in [('Y', 'Y'), ('Z', 'Z'), ('AA', 'AA'),
                                            ('AB', 'AB'), ('AC', 'AC'), ('AD', 'AD'),
                                            ('AE', 'AE'), ('AF', 'AF')]:
                            row_list[col_to_idx(col_l) - 1] = computed.get(key)

                        # Tối ưu ghi file: Định dạng (format) hiển thị cho các cột số, ngày và tỷ lệ
                        cells_out = []
                        for col_out_idx in range(31):
                            col_letter = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
                                          'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                                          'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB',
                                          'AC', 'AD', 'AE', 'AF'][col_out_idx]
                            val = row_list[col_out_idx]

                            if col_letter in ('M', 'W', 'X') and val is not None:
                                dt = excel_date_to_datetime(val) if not (col_letter in ('W', 'X') and str(val).strip() == '>90') else None
                                if dt:
                                    cell = WriteOnlyCell(ws_out, value=dt)
                                    cell.number_format = 'dd/mm/yyyy'
                                    cell.font = font_normal
                                    cell.alignment = align_center
                                else:
                                    cell = WriteOnlyCell(ws_out, value=val)
                                    cell.font = font_normal
                                    cell.alignment = align_center
                                cells_out.append(cell)

                            elif col_letter in ('N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'Y', 'Z') and val is not None:
                                try:
                                    if col_letter == 'Y' and str(val).strip() == '>90':
                                        cell = WriteOnlyCell(ws_out, value=val)
                                        cell.font = font_normal
                                        cell.alignment = align_right
                                    else:
                                        num_val = int(float(val))
                                        cell = WriteOnlyCell(ws_out, value=num_val)
                                        cell.number_format = '#,##0'
                                        cell.font = font_normal
                                        cell.alignment = align_right
                                except (ValueError, TypeError):
                                    cell = WriteOnlyCell(ws_out, value=val)
                                    cell.font = font_normal
                                    cell.alignment = align_right
                                cells_out.append(cell)

                            elif col_letter in ('AA', 'AB') and val is not None:
                                try:
                                    num_val = float(val)
                                    cell = WriteOnlyCell(ws_out, value=num_val)
                                    cell.number_format = '0.00'
                                    cell.font = font_normal
                                    cell.alignment = align_right
                                except (ValueError, TypeError):
                                    cell = WriteOnlyCell(ws_out, value=val)
                                    cell.font = font_normal
                                    cell.alignment = align_right
                                cells_out.append(cell)

                            elif col_letter == 'AF':
                                fill_af, font_af = phan_loai_style.get(phan_loai, (None, font_normal))
                                cell = WriteOnlyCell(ws_out, value=val)
                                cell.font = font_af
                                if fill_af:
                                    cell.fill = fill_af
                                cell.alignment = align_center
                                cells_out.append(cell)

                            else:
                                # Các cột chuỗi thường (B, C, D, E, F, G, H, I, J, K, L, AC, AD, AE) ghi thô
                                cells_out.append(val)

                        ws_out.append(cells_out)
                        rows_written += 1

                    elem.clear()

        wb_out.save(out_path)
        elapsed = time.time() - t0
        if log_func:
            log_func(f"  ✓ Hoàn thành: {rows_written:,} hàng xuất, {rows_filtered:,} hàng lọc bỏ")
            log_func(f"  → Thời gian: {elapsed:.1f}s | File: {out_filename}")
        return out_path

    except Exception as e:
        if log_func:
            log_func(f"  ✗ Lỗi xử lý {filename}: {e}")
            import traceback
            log_func(traceback.format_exc())
        return None
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass


# ===================================================================
# GIAO DIỆN NGƯỜI DÙNG (GUI)
# ===================================================================
class App(tk.Tk):
    """
    Giao diện chính của Tool Đổ Dữ Liệu.
    Các chức năng:
      - Chọn nhiều file Excel đầu vào
      - Chọn file XLSB tra cứu
      - Nhập ngày mốc và số tiền tham chiếu
      - Chạy xử lý và hiển thị log
    """

    # Màu sắc giao diện (dark theme)
    BG_DARK = '#1a1d2e'
    BG_CARD = '#252840'
    BG_INPUT = '#1e2235'
    ACCENT = '#4f8ef7'
    ACCENT_HOVER = '#6ba3ff'
    SUCCESS = '#4caf50'
    WARNING = '#ff9800'
    ERROR = '#f44336'
    TEXT = '#e8eaed'
    TEXT_DIM = '#9aa0b8'
    BORDER = '#353a5e'

    def __init__(self):
        super().__init__()
        self.title(f"{APP_TITLE} v{APP_VERSION}")
        self.geometry("980x720")
        self.minsize(800, 600)
        self.configure(bg=self.BG_DARK)

        # Biến trạng thái
        self.selected_files: list[str] = []
        self.xlsb_path = tk.StringVar(value=DEFAULT_XLSB_PATH)
        self.milestone_date_str = tk.StringVar(value=datetime.now().strftime('%d/%m/%Y'))
        self.ref_amount_str = tk.StringVar(value='100,000')
        self.is_running = False

        self._build_ui()
        self._center_window()

    def _center_window(self):
        self.update_idletasks()
        w = self.winfo_width()
        h = self.winfo_height()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _build_ui(self):
        """Xây dựng toàn bộ giao diện."""
        # === Header ===
        header = tk.Frame(self, bg='#131629', pady=12)
        header.pack(fill='x')

        tk.Label(
            header, text="📊  Tool Đổ Dữ Liệu Phân Tích Tồn Kho",
            font=('Segoe UI', 16, 'bold'), bg='#131629', fg=self.TEXT
        ).pack(side='left', padx=20)

        tk.Label(
            header, text=f"v{APP_VERSION}", font=('Segoe UI', 9),
            bg='#131629', fg=self.TEXT_DIM
        ).pack(side='right', padx=20)

        # === Footer: nút chạy ===
        footer = tk.Frame(self, bg='#131629', pady=10)
        footer.pack(fill='x', side='bottom')

        self.run_btn = tk.Button(
            footer,
            text="▶  BẮT ĐẦU XỬ LÝ",
            font=('Segoe UI', 12, 'bold'),
            bg=self.ACCENT, fg='white',
            activebackground=self.ACCENT_HOVER,
            activeforeground='white',
            cursor='hand2',
            relief='flat',
            padx=30, pady=8,
            command=self._on_run
        )
        self.run_btn.pack(side='left', padx=20)

        self.status_lbl = tk.Label(
            footer, text="Sẵn sàng.",
            font=('Segoe UI', 10), bg='#131629', fg=self.TEXT_DIM
        )
        self.status_lbl.pack(side='left', padx=10)

        # === Main content ===
        content = tk.Frame(self, bg=self.BG_DARK)
        content.pack(fill='both', expand=True, padx=15, pady=10)

        # Left panel: cấu hình
        left = tk.Frame(content, bg=self.BG_DARK)
        left.pack(side='left', fill='y', padx=(0, 10))

        # Right panel: log
        right = tk.Frame(content, bg=self.BG_DARK)
        right.pack(side='right', fill='both', expand=True)

        self._build_config_panel(left)
        self._build_log_panel(right)

    def _card(self, parent, title: str) -> tk.Frame:
        """Tạo card với tiêu đề."""
        outer = tk.Frame(parent, bg=self.BG_CARD, pady=0)
        outer.pack(fill='x', pady=5)
        tk.Label(
            outer, text=title, font=('Segoe UI', 10, 'bold'),
            bg=self.BG_CARD, fg=self.ACCENT, anchor='w', padx=12, pady=6
        ).pack(fill='x')
        tk.Frame(outer, bg=self.BORDER, height=1).pack(fill='x', padx=12)
        inner = tk.Frame(outer, bg=self.BG_CARD, padx=12, pady=10)
        inner.pack(fill='x')
        return inner

    def _build_config_panel(self, parent):
        """Panel bên trái: cấu hình tham số."""
        parent.configure(width=360)

        # --- Card 1: File đầu vào ---
        c1 = self._card(parent, "1. File Excel đầu vào (.xlsx)")

        self.file_listbox = tk.Listbox(
            c1, font=('Segoe UI', 9), bg=self.BG_INPUT, fg=self.TEXT,
            selectbackground=self.ACCENT, selectforeground='white',
            height=6, relief='flat', bd=0,
            highlightthickness=1, highlightcolor=self.BORDER,
            highlightbackground=self.BORDER
        )
        self.file_listbox.pack(fill='x', pady=(0, 6))

        btn_frame = tk.Frame(c1, bg=self.BG_CARD)
        btn_frame.pack(fill='x')
        self._btn(btn_frame, "📂 Thêm file", self._add_files).pack(side='left')
        self._btn(btn_frame, "✖ Xóa chọn", self._remove_selected, color='#e53935').pack(side='left', padx=4)
        self._btn(btn_frame, "🗑 Xóa tất cả", self._clear_files, color='#555').pack(side='left')

        self.file_count_lbl = tk.Label(
            c1, text="Chưa chọn file nào.",
            font=('Segoe UI', 8), bg=self.BG_CARD, fg=self.TEXT_DIM, anchor='w'
        )
        self.file_count_lbl.pack(fill='x', pady=(4, 0))

        # --- Card 2: File XLSB tra cứu ---
        c2 = self._card(parent, "2. File tra cứu hạn sử dụng (.xlsb)")

        xlsb_frame = tk.Frame(c2, bg=self.BG_CARD)
        xlsb_frame.pack(fill='x')

        self.xlsb_entry = tk.Entry(
            xlsb_frame, textvariable=self.xlsb_path,
            font=('Segoe UI', 8), bg=self.BG_INPUT, fg=self.TEXT,
            relief='flat', bd=0, insertbackground=self.TEXT,
            highlightthickness=1, highlightcolor=self.BORDER,
            highlightbackground=self.BORDER
        )
        self.xlsb_entry.pack(side='left', fill='x', expand=True, pady=2)

        self._btn(xlsb_frame, "📂", self._browse_xlsb).pack(side='right', padx=(4, 0))

        # --- Card 3: Tham số ---
        c3 = self._card(parent, "3. Tham số tính toán")

        self._field(c3, "Ngày mốc (Y1):", self.milestone_date_str,
                    tip="Định dạng: DD/MM/YYYY")
        self.ref_amount_entry = self._field(
            c3, "Số tiền tham chiếu (AF1):", self.ref_amount_str,
            tip="Đơn vị: VNĐ (ví dụ: 100,000)"
        )
        self.ref_amount_entry.bind("<KeyRelease>", self._format_ref_amount)
        self.ref_amount_entry.bind("<FocusOut>", self._format_ref_amount)

        # --- Card 4: Ghi chú ---
        c4 = self._card(parent, "📋 Ghi chú phân loại")
        notes = [
            ("🔴", "Hết hạn", "Last GR + Date tham khảo < Ngày mốc"),
            ("🟠", "Nghi vấn tồn ảo", "Không nhập, không bán, có tồn"),
            ("🟡", "Non-moving", "Doanh thu = 0 (90 ngày)"),
            ("⚪", "Slow moving", "DIO/Date ≥ 5 và Giá trị tồn ≥ AF1"),
        ]
        for icon, label, desc in notes:
            row = tk.Frame(c4, bg=self.BG_CARD)
            row.pack(fill='x', pady=1)
            tk.Label(row, text=icon, bg=self.BG_CARD, font=('Segoe UI', 10)).pack(side='left')
            tk.Label(row, text=f" {label}:", font=('Segoe UI', 8, 'bold'),
                     bg=self.BG_CARD, fg=self.TEXT, width=18, anchor='w').pack(side='left')
            tk.Label(row, text=desc, font=('Segoe UI', 8),
                     bg=self.BG_CARD, fg=self.TEXT_DIM, anchor='w').pack(side='left')

    def _build_log_panel(self, parent):
        """Panel bên phải: log output."""
        tk.Label(
            parent, text="Nhật ký xử lý",
            font=('Segoe UI', 11, 'bold'), bg=self.BG_DARK, fg=self.TEXT, anchor='w'
        ).pack(fill='x', pady=(0, 4))

        log_frame = tk.Frame(parent, bg=self.BG_CARD, bd=0)
        log_frame.pack(fill='both', expand=True)

        self.log_text = tk.Text(
            log_frame, font=('Consolas', 9), bg=self.BG_INPUT, fg=self.TEXT,
            relief='flat', bd=0, insertbackground=self.TEXT,
            highlightthickness=0, state='disabled'
        )
        self.log_text.pack(side='left', fill='both', expand=True, padx=(1, 0), pady=1)

        scrollbar = ttk.Scrollbar(log_frame, orient='vertical', command=self.log_text.yview)
        scrollbar.pack(side='right', fill='y')
        self.log_text.configure(yscrollcommand=scrollbar.set)

        # Định nghĩa tags màu cho log
        self.log_text.tag_config('info', foreground=self.TEXT)
        self.log_text.tag_config('success', foreground=self.SUCCESS)
        self.log_text.tag_config('warning', foreground=self.WARNING)
        self.log_text.tag_config('error', foreground=self.ERROR)
        self.log_text.tag_config('highlight', foreground=self.ACCENT, font=('Consolas', 9, 'bold'))

    def _btn(self, parent, text: str, cmd, color=None) -> tk.Button:
        """Helper tạo button."""
        bg_color = color if color else self.ACCENT
        active_bg = self.ACCENT_HOVER if not color else color
        btn = tk.Button(
            parent, text=text, font=('Segoe UI', 8, 'bold'),
            bg=bg_color, fg='white', activebackground=active_bg,
            activeforeground='white', relief='flat', cursor='hand2',
            padx=10, pady=4, command=cmd
        )
        return btn

    def _field(self, parent, label: str, var: tk.StringVar, tip: str = ''):
        frame = tk.Frame(parent, bg=self.BG_CARD)
        frame.pack(fill='x', pady=3)
        tk.Label(
            frame, text=label, font=('Segoe UI', 9),
            bg=self.BG_CARD, fg=self.TEXT, anchor='w', width=25
        ).pack(side='left')
        entry = tk.Entry(
            frame, textvariable=var, font=('Segoe UI', 9),
            bg=self.BG_INPUT, fg=self.TEXT, relief='flat',
            insertbackground=self.TEXT,
            highlightthickness=1, highlightcolor=self.ACCENT,
            highlightbackground=self.BORDER
        )
        entry.pack(side='left', fill='x', expand=True)
        if tip:
            tk.Label(
                parent, text=tip, font=('Segoe UI', 7),
                bg=self.BG_CARD, fg=self.TEXT_DIM, anchor='e'
            ).pack(fill='x')
        return entry

    def _format_ref_amount(self, event=None):
        """Tự động thêm dấu phẩy phân tách hàng nghìn khi người dùng gõ phím."""
        try:
            cursor_pos = self.ref_amount_entry.index(tk.INSERT)
        except Exception:
            cursor_pos = 0

        raw = self.ref_amount_str.get()
        # Đếm số dấu phẩy trước con trỏ trước khi định dạng
        commas_before = raw[:cursor_pos].count(',')
        
        # Chỉ lấy chữ số
        clean = "".join([c for c in raw if c.isdigit()])
        if clean:
            try:
                formatted = f"{int(clean):,}"
                self.ref_amount_str.set(formatted)
                
                # Tính toán lại vị trí con trỏ để tránh bị nhảy về cuối ô nhập liệu
                digits_before = cursor_pos - commas_before
                new_pos = 0
                digit_count = 0
                for char in formatted:
                    if digit_count == digits_before:
                        break
                    if char.isdigit():
                        digit_count += 1
                    new_pos += 1
                
                self.ref_amount_entry.icursor(new_pos)
            except (ValueError, TypeError):
                pass
        else:
            self.ref_amount_str.set("")

    # --- Actions ---
    def _add_files(self):
        files = filedialog.askopenfilenames(
            title="Chọn file báo cáo tồn kho",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        for f in files:
            norm_f = os.path.abspath(f)
            if norm_f not in self.selected_files:
                self.selected_files.append(norm_f)
                self.file_listbox.insert(tk.END, os.path.basename(norm_f))
        self._update_file_count()

    def _remove_selected(self):
        selected_indices = list(self.file_listbox.curselection())
        for idx in reversed(selected_indices):
            self.selected_files.pop(idx)
            self.file_listbox.delete(idx)
        self._update_file_count()

    def _clear_files(self):
        self.selected_files.clear()
        self.file_listbox.delete(0, tk.END)
        self._update_file_count()

    def _update_file_count(self):
        count = len(self.selected_files)
        if count == 0:
            self.file_count_lbl.config(text="Chưa chọn file nào.", fg=self.TEXT_DIM)
        else:
            self.file_count_lbl.config(text=f"Đã chọn {count} file Excel.", fg=self.SUCCESS)

    def _browse_xlsb(self):
        f = filedialog.askopenfilename(
            title="Chọn file tra cứu XLSB",
            filetypes=[("XLSB Files", "*.xlsb")]
        )
        if f:
            self.xlsb_path.set(os.path.abspath(f))

    def log(self, message: str, tag: str = 'info'):
        """Ghi log vào khung text."""
        self.log_text.config(state='normal')
        self.log_text.insert(tk.END, message + "\n", tag)
        self.log_text.see(tk.END)
        self.log_text.config(state='disabled')
        self.update_idletasks()

    def _on_run(self):
        if self.is_running:
            return
        
        # Kiểm tra đầu vào
        if not self.selected_files:
            messagebox.showerror("Lỗi", "Vui lòng chọn ít nhất 1 file Excel đầu vào!")
            return

        xlsb = self.xlsb_path.get().strip()
        if not os.path.exists(xlsb):
            messagebox.showerror("Lỗi", f"Không tìm thấy file tra cứu XLSB tại:\n{xlsb}")
            return

        # Parse ngày mốc
        date_str = self.milestone_date_str.get().strip()
        try:
            milestone_date = datetime.strptime(date_str, '%d/%m/%Y')
        except ValueError:
            messagebox.showerror("Lỗi", "Ngày mốc không đúng định dạng DD/MM/YYYY!\nVí dụ: 01/07/2026")
            return

        # Parse số tiền tham chiếu (loại bỏ dấu phẩy/chấm phân tách hàng nghìn)
        ref_str = self.ref_amount_str.get().strip().replace(',', '').replace('.', '')
        try:
            ref_amount = float(ref_str)
        except ValueError:
            messagebox.showerror("Lỗi", "Số tiền tham chiếu phải là một số hợp lệ!")
            return

        # Vô hiệu hóa nút và bắt đầu chạy thread ngầm
        self.is_running = True
        self.run_btn.config(state='disabled', text="⏳  ĐANG XỬ LÝ...")
        self.status_lbl.config(text="Đang xử lý dữ liệu...", fg=self.WARNING)
        self.log_text.config(state='normal')
        self.log_text.delete('1.0', tk.END)
        self.log_text.config(state='disabled')

        def worker():
            t_start = time.time()
            self.log(f"=== BẮT ĐẦU TIẾN TRÌNH XỬ LÝ (v{APP_VERSION}) ===", 'highlight')
            self.log(f"Ngày mốc: {date_str} | Số tiền tham chiếu: {ref_amount:,.0f} VNĐ\n")
            
            # 1. Nạp XLSB
            shelf_life_map = load_shelf_life_map(xlsb, lambda msg, t='info': self.log(msg, t))
            
            # 2. Xử lý từng file
            success_count = 0
            for idx, file in enumerate(self.selected_files):
                filename = os.path.basename(file)
                self.log(f"\n[{idx+1}/{len(self.selected_files)}] Đang xử lý: {filename}...", 'highlight')
                
                try:
                    out = process_excel_file(
                        file, shelf_life_map, milestone_date, ref_amount,
                        log_func=lambda msg, t='info': self.log(msg, t)
                    )
                    if out:
                        success_count += 1
                except Exception as ex:
                    self.log(f"  ✗ Lỗi không xác định: {ex}", 'error')

            elapsed = time.time() - t_start
            self.log(f"\n=== TIẾN TRÌNH HOÀN TẤT ===", 'highlight')
            self.log(f"Đã xử lý thành công {success_count}/{len(self.selected_files)} file.", 'success')
            self.log(f"Tổng thời gian thực hiện: {elapsed:.1f} giây.\n", 'success')
            
            # Mở thư mục chứa file done đầu tiên nếu thành công
            if success_count > 0 and self.selected_files:
                first_dir = os.path.dirname(self.selected_files[0])
                try:
                    os.startfile(first_dir)
                except Exception:
                    pass

            # Khôi phục trạng thái giao diện
            self.is_running = False
            self.run_btn.config(state='normal', text="▶  BẮT ĐẦU XỬ LÝ")
            self.status_lbl.config(text="Đã hoàn tất xử lý.", fg=self.SUCCESS)

        # Chạy thread ngầm để không đơ GUI
        Thread(target=worker, daemon=True).start()


# ===================================================================
# KÍCH HOẠT ỨNG DỤNG
# ===================================================================
if __name__ == '__main__':
    # Đăng ký DPI để giao diện hiển thị sắc nét trên màn hình Windows độ phân giải cao
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass

    app = App()
    app.mainloop()
