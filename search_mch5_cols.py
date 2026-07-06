import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

path = r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu\Bùi Anh Tuấn.xlsx"
try:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = wb['Chi tiết data']
    
    found = 0
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i < 2:
            continue
        row_str = str(row)
        if "Bánh bao" in row_str or "Đậu hũ" in row_str:
            print(f"Row {i} has match:")
            for idx, val in enumerate(row):
                if val is not None and ("Bánh bao" in str(val) or "Đậu hũ" in str(val) or val in ["Bánh bao", "Đậu hũ"]):
                    print(f"  Col {idx} ({openpyxl.utils.get_column_letter(idx+1)}): {val}")
            found += 1
            if found >= 5:
                break
except Exception as e:
    print(f"Error: {e}")
