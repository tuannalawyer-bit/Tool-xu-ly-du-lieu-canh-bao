import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

path = r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu\Bùi Anh Tuấn.xlsx"
try:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = wb['Chi tiết data']
    
    # We want to check for the first 100 rows where column AF is not empty, and print all columns.
    # We also want to check if any row has more than 32 columns.
    max_cols = 0
    sample_rows = []
    
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if len(row) > max_cols:
            max_cols = len(row)
        
        # Check if column AF has data and grab first few
        val = row[31] if len(row) > 31 else None
        if val is not None and str(val).strip() != "" and len(sample_rows) < 5 and i >= 2:
            sample_rows.append((i, row))
            
    print(f"Max columns in sheet: {max_cols}")
    print("Sample rows with check data:")
    for r_idx, row in sample_rows:
        print(f"Row {r_idx}:")
        for c_idx, val in enumerate(row):
            print(f"  Col {c_idx} ({openpyxl.utils.get_column_letter(c_idx+1)}): {val}")
            
except Exception as e:
    print(f"Error: {e}")
