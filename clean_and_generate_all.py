import os
import sys
import time
import zipfile
import xml.etree.ElementTree as ET
import ctypes
import concurrent.futures
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell import WriteOnlyCell
from pyxlsb import open_workbook
from datetime import datetime

# Force unbuffered output
sys.stdout.reconfigure(encoding='utf-8')

folder = r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu"
scratch_dir = r"C:\Users\Tuan\.gemini\antigravity\brain\b053fc36-8fe6-4316-bba8-ac8fef252caf\scratch"
xlsb_path = os.path.join(folder, "raw mẫu", "updated_data_moi.xlsb")
template_path = os.path.join(scratch_dir, "dashboard_win_template.html")

# Define files for all chains
chains = {
    "rural": [
        "Lê Thị Hồng Thu.xlsx",
        "Hoàng Nguyễn Tú Anh.xlsx",
        "Bùi Anh Tuấn.xlsx",
        "Lạc Nhật Minh.xlsx"
    ],
    "urban": [
        "Phạm Thị Ngọc Xuyến.xlsx",
        "Lê Duy Đức.xlsx",
        "Vương Phi Sơn.xlsx",
        "Trần Thị Diệp.xlsx",
        "Đỗ Khắc Chức.xlsx",
        "NGuyễn Văn Tuấn.xlsx"
    ],
    "win": [
        "Nguyễn Đức Thiên Ân.xlsx",
        "Đỗ Thị Thanh Loan.xlsx",
        "Lê Văn Trí.xlsx",
        "Nguyễn minh Trang.xlsx"
    ]
}

def normalize_rsm_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    name_lower = name_str.lower()
    if name_lower == "nguyễn văn tuấn":
        return "Nguyễn Văn Tuấn"
    if name_lower == "nguyễn minh trang":
        return "Nguyễn Minh Trang"
    return name_str

def col_to_idx(col):
    val = 0
    for char in col:
        val = val * 26 + (ord(char.upper()) - 64)
    return val - 1

# ==========================================
# STEP 1: Merge Excel Files (excl. 'Result')
# ==========================================
def merge_excel_files():
    print("\n==================================================", flush=True)
    print("STEP 1: Re-merging Excel files (filtering out 'Result' rows)...", flush=True)
    print("==================================================", flush=True)
    
    for chain_name, file_list in chains.items():
        t_start = time.time()
        output_filename = f"Tong_hop_Kiem_tra_{chain_name.capitalize()}.xlsx"
        output_path = os.path.join(folder, output_filename)
        print(f"\nProcessing Excel Merge for {chain_name.upper()} -> {output_filename}", flush=True)
        
        wb_out = openpyxl.Workbook(write_only=True)
        ws_out = wb_out.create_sheet(title="Dữ liệu kiểm tra")
        
        headers_written = False
        total_copied = 0
        total_skipped_result = 0
        
        for idx, filename in enumerate(file_list):
            src_path = os.path.join(folder, filename)
            temp_path = os.path.join(scratch_dir, f"temp_split_merge_{filename}")
            
            print(f"  [{idx+1}/{len(file_list)}] Reading {filename}...", flush=True)
            res = ctypes.windll.kernel32.CopyFileW(src_path, temp_path, False)
            if not res:
                print(f"    Error: Copy failed for {filename}", flush=True)
                continue
                
            try:
                with zipfile.ZipFile(temp_path, 'r') as z:
                    # Find worksheet XML path
                    wb_data = z.read('xl/workbook.xml')
                    root_wb = ET.fromstring(wb_data)
                    ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    
                    sheets = []
                    for sheet in root_wb.findall('.//ns:sheet', ns):
                        sheets.append((sheet.attrib.get('name'), sheet.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')))
                        
                    sheet_name, target_rid = [s for s in sheets if s[0] not in ['_com.sap.ip.bi.xl.hiddensheet', 'RC']][0]
                    
                    rels_data = z.read('xl/_rels/workbook.xml.rels')
                    root_rels = ET.fromstring(rels_data)
                    r_ns = {'rels': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                    
                    sheet_path_in_zip = None
                    for rel in root_rels.findall('.//rels:Relationship', r_ns):
                        if rel.attrib.get('Id') == target_rid:
                            sheet_path_in_zip = f"xl/{rel.attrib.get('Target')}"
                            break
                            
                    # Load shared strings
                    shared_strings = []
                    if 'xl/sharedStrings.xml' in z.namelist():
                        sst_file = z.open('xl/sharedStrings.xml')
                        context_sst = ET.iterparse(sst_file, events=('start', 'end'))
                        for event, elem in context_sst:
                            if event == 'end' and elem.tag.endswith('}si'):
                                text = "".join([t.text for t in elem.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t') if t.text is not None])
                                shared_strings.append(text)
                                elem.clear()
                                
                    # Stream parse worksheet XML
                    sheet_file = z.open(sheet_path_in_zip)
                    context = ET.iterparse(sheet_file, events=('start', 'end'))
                    
                    current_row = {}
                    current_row_idx = None
                    
                    def get_val(cell_elem):
                        t_type = cell_elem.attrib.get('t')
                        v_elem = cell_elem.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                        if v_elem is not None and v_elem.text is not None:
                            val_text = v_elem.text
                            if t_type == 's':
                                idx = int(val_text)
                                if idx < len(shared_strings):
                                    return shared_strings[idx]
                            if val_text.isdigit():
                                return int(val_text)
                            try:
                                return float(val_text)
                            except:
                                return val_text
                        return None
                        
                    for event, elem in context:
                        if event == 'start' and elem.tag.endswith('}row'):
                            current_row = {}
                            r_attr = elem.attrib.get('r')
                            current_row_idx = int(r_attr) if r_attr and r_attr.isdigit() else None
                            
                        elif event == 'end' and elem.tag.endswith('}c'):
                            r_ref = elem.attrib.get('r')
                            if r_ref:
                                digit_idx = 0
                                for char in r_ref:
                                    if char.isdigit():
                                        break
                                    digit_idx += 1
                                current_row[r_ref[:digit_idx]] = get_val(elem)
                                
                        elif event == 'end' and elem.tag.endswith('}row'):
                            if current_row_idx:
                                if current_row_idx <= 2:
                                    if not headers_written:
                                        row_list = [None] * 32
                                        for col_letter, val in current_row.items():
                                            idx = col_to_idx(col_letter)
                                            if idx < 32:
                                                row_list[idx] = val
                                        ws_out.append(row_list)
                                else:
                                    af_val = current_row.get('AF')
                                    if af_val is not None:
                                        af_str = str(af_val).strip()
                                        if af_str != "" and af_str.upper() not in ["N/A", "#N/A"]:
                                            # Check 'Result' in store columns
                                            store_code = str(current_row.get('D', '')).strip()
                                            store_name = str(current_row.get('E', '')).strip()
                                            
                                            if "result" in store_code.lower() or "result" in store_name.lower():
                                                total_skipped_result += 1
                                                continue
                                                
                                            row_list = [None] * 32
                                            for col_letter, val in current_row.items():
                                                idx = col_to_idx(col_letter)
                                                if idx < 32:
                                                    row_list[idx] = val
                                                    
                                            row_list[1] = normalize_rsm_name(row_list[1])
                                            ws_out.append(row_list)
                                            total_copied += 1
                            elem.clear()
                            
                    if not headers_written:
                        headers_written = True
                        
            except Exception as e:
                print(f"    Error processing {filename}: {e}", flush=True)
            finally:
                if os.path.exists(temp_path):
                    try:
                        os.remove(temp_path)
                    except:
                        pass
                        
        print("  Saving workbook...", flush=True)
        wb_out.save(output_path)
        print(f"  Completed. Rows Copied: {total_copied:,}, Skipped 'Result': {total_skipped_result:,}, Time: {time.time()-t_start:.1f}s", flush=True)


# ==========================================
# STEP 2: Prepare JSON Data (excl. 'Result')
# ==========================================
def load_master_articles():
    print("\nLoading master articles from XLSB...", flush=True)
    t0 = time.time()
    mapping = {}
    try:
        with open_workbook(xlsb_path) as wb:
            with wb.get_sheet('master article') as sheet:
                for i, row in enumerate(sheet.rows()):
                    if i == 0:
                        continue
                    prod_id = row[0].v
                    mch5_desc = row[22].v if len(row) > 22 else None
                    if prod_id:
                        prod_id_str = str(int(prod_id)) if isinstance(prod_id, float) else str(prod_id).strip()
                        mapping[prod_id_str] = str(mch5_desc).strip() if mch5_desc else "Khác"
        print(f"Loaded {len(mapping)} articles in {time.time()-t0:.2f}s", flush=True)
    except Exception as e:
        print(f"Error loading master articles: {e}", flush=True)
    return mapping

def process_file_raw(args):
    filename, product_mch5_map, chain_name = args
    src_path = os.path.join(folder, filename)
    temp_path = os.path.join(scratch_dir, f"temp_raw_{chain_name}_{filename}")
    
    t_start = time.time()
    try:
        res = ctypes.windll.kernel32.CopyFileW(src_path, temp_path, False)
        if not res:
            return {"filename": filename, "status": "error", "message": "Copy failed"}
            
        with zipfile.ZipFile(temp_path, 'r') as z:
            wb_data = z.read('xl/workbook.xml')
            root_wb = ET.fromstring(wb_data)
            ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            
            sheets = []
            for sheet in root_wb.findall('.//ns:sheet', ns):
                sheets.append((sheet.attrib.get('name'), sheet.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')))
                
            sheet_name, target_rid = [s for s in sheets if s[0] not in ['_com.sap.ip.bi.xl.hiddensheet', 'RC']][0]
            
            rels_data = z.read('xl/_rels/workbook.xml.rels')
            root_rels = ET.fromstring(rels_data)
            r_ns = {'rels': 'http://schemas.openxmlformats.org/package/2006/relationships'}
            
            sheet_path_in_zip = None
            for rel in root_rels.findall('.//rels:Relationship', r_ns):
                if rel.attrib.get('Id') == target_rid:
                    sheet_path_in_zip = f"xl/{rel.attrib.get('Target')}"
                    break
                    
            shared_strings = []
            if 'xl/sharedStrings.xml' in z.namelist():
                sst_file = z.open('xl/sharedStrings.xml')
                context_sst = ET.iterparse(sst_file, events=('start', 'end'))
                for event, elem in context_sst:
                    if event == 'end' and elem.tag.endswith('}si'):
                        text = "".join([t.text for t in elem.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t') if t.text is not None])
                        shared_strings.append(text)
                        elem.clear()
            
            sheet_file = z.open(sheet_path_in_zip)
            context = ET.iterparse(sheet_file, events=('start', 'end'))
            
            rows_data = []
            current_row = {}
            current_row_idx = None
            
            def get_val(cell_elem):
                t_type = cell_elem.attrib.get('t')
                v_elem = cell_elem.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                if v_elem is not None and v_elem.text is not None:
                    val_text = v_elem.text
                    if t_type == 's':
                        idx = int(val_text)
                        if idx < len(shared_strings):
                            return shared_strings[idx]
                    return val_text
                return None
                
            for event, elem in context:
                if event == 'start' and elem.tag.endswith('}row'):
                    current_row = {}
                    r_attr = elem.attrib.get('r')
                    current_row_idx = int(r_attr) if r_attr and r_attr.isdigit() else None
                    
                elif event == 'end' and elem.tag.endswith('}c'):
                    r_ref = elem.attrib.get('r')
                    if r_ref:
                        digit_idx = 0
                        for char in r_ref:
                            if char.isdigit():
                                break
                            digit_idx += 1
                        current_row[r_ref[:digit_idx]] = get_val(elem)
                        
                elif event == 'end' and elem.tag.endswith('}row'):
                    if current_row_idx and current_row_idx >= 3:
                        af_val = current_row.get('AF')
                        if af_val is not None:
                            af_str = str(af_val).strip()
                            if af_str != "" and af_str.upper() not in ["N/A", "#N/A"]:
                                store_code = str(current_row.get('D', '')).strip()
                                store_name = str(current_row.get('E', '')).strip()
                                
                                # Skip Result rows
                                if "result" in store_code.lower() or "result" in store_name.lower():
                                    continue
                                    
                                rsm = normalize_rsm_name(current_row.get('B'))
                                asm = str(current_row.get('C', '')).strip()
                                mch2 = str(current_row.get('G', '')).strip()
                                mch3 = str(current_row.get('I', '')).strip()
                                mch4 = str(current_row.get('J', '')).strip()
                                art_code = str(current_row.get('K', '')).strip()
                                art_name = str(current_row.get('L', '')).strip()
                                
                                if art_code.endswith('.0'):
                                    art_code = art_code[:-2]
                                    
                                mch5 = product_mch5_map.get(art_code, "Khác")
                                qty_str = current_row.get('Q', '0')
                                val_str = current_row.get('R', '0')
                                
                                try:
                                    qty = float(qty_str) if qty_str else 0.0
                                except:
                                    qty = 0.0
                                try:
                                    val = float(val_str) if val_str else 0.0
                                except:
                                    val = 0.0
                                    
                                rows_data.append({
                                    "rsm": rsm, "asm": asm, "store_code": store_code, "store_name": store_name,
                                    "mch2": mch2, "mch3": mch3, "mch4": mch4, "mch5": mch5,
                                    "article_code": art_code, "article_name": art_name,
                                    "qty": qty, "value": val, "check_class": af_str
                                })
                    elem.clear()
            return {"filename": filename, "status": "success", "data": rows_data, "time_taken": time.time() - t_start}
    except Exception as e:
        return {"filename": filename, "status": "error", "message": str(e)}
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass

def prepare_json_data(product_mch5_map):
    print("\n==================================================", flush=True)
    print("STEP 2: Preparing V2 JSON Data (excluding 'Result' rows)...", flush=True)
    print("==================================================", flush=True)
    
    workers = min(os.cpu_count() or 4, 4)
    
    for chain_name, file_list in chains.items():
        print(f"\nProcessing JSON data for {chain_name.upper()}...", flush=True)
        t_start = time.time()
        
        all_rows = []
        with concurrent.futures.ProcessPoolExecutor(max_workers=workers) as executor:
            args_list = [(f, product_mch5_map, chain_name) for f in file_list]
            futures = {executor.submit(process_file_raw, args): args[0] for args in args_list}
            for future in concurrent.futures.as_completed(futures):
                f = futures[future]
                res = future.result()
                if res["status"] == "success":
                    all_rows.extend(res["data"])
                    print(f"  Processed {f} in {res['time_taken']:.1f}s: rows={len(res['data']):,}", flush=True)
                else:
                    print(f"  Error processing {f}: {res['message']}", flush=True)
                    
        print(f"  Total rows extracted: {len(all_rows):,}", flush=True)
        
        # Compress using V2 structure (Dictionary encoding for articles)
        rsms = set()
        asms = set()
        stores = {}
        mch2s = set()
        mch3s = set()
        mch4s = set()
        mch5s = set()
        classes = set()
        articles = {}
        
        for r in all_rows:
            rsms.add(r["rsm"])
            asms.add(r["asm"])
            stores[r["store_code"]] = r["store_name"]
            mch2s.add(r["mch2"])
            mch3s.add(r["mch3"])
            mch4s.add(r["mch4"])
            mch5s.add(r["mch5"])
            classes.add(r["check_class"])
            
            art_key = (r["article_code"], r["article_name"])
            if art_key not in articles:
                articles[art_key] = len(articles)
                
        rsm_list = sorted(list(rsms))
        asm_list = sorted(list(asms))
        store_list = [[code, name] for code, name in sorted(stores.items())]
        mch2_list = sorted(list(mch2s))
        mch3_list = sorted(list(mch3s))
        mch4_list = sorted(list(mch4s))
        mch5_list = sorted(list(mch5s))
        class_list = sorted(list(classes))
        
        article_list = [None] * len(articles)
        for (code, name), idx in articles.items():
            article_list[idx] = [code, name]
            
        rsm_map = {name: idx for idx, name in enumerate(rsm_list)}
        asm_map = {name: idx for idx, name in enumerate(asm_list)}
        store_map = {item[0]: idx for idx, item in enumerate(store_list)}
        mch2_map = {name: idx for idx, name in enumerate(mch2_list)}
        mch3_map = {name: idx for idx, name in enumerate(mch3_list)}
        mch4_map = {name: idx for idx, name in enumerate(mch4_list)}
        mch5_map = {name: idx for idx, name in enumerate(mch5_list)}
        class_map = {name: idx for idx, name in enumerate(class_list)}
        
        compressed_rows = []
        for r in all_rows:
            art_key = (r["article_code"], r["article_name"])
            row = [
                store_map[r["store_code"]],
                rsm_map[r["rsm"]],
                asm_map[r["asm"]],
                mch2_map[r["mch2"]],
                mch3_map[r["mch3"]],
                mch4_map[r["mch4"]],
                mch5_map[r["mch5"]],
                articles[art_key],          # article_idx
                class_map[r["check_class"]], # class_idx
                round(r["qty"]),
                int(r["value"])
            ]
            compressed_rows.append(row)
            
        out_dict = {
            "rsm_list": rsm_list,
            "asm_list": asm_list,
            "store_list": store_list,
            "mch2_list": mch2_list,
            "mch3_list": mch3_list,
            "mch4_list": mch4_list,
            "mch5_list": mch5_list,
            "class_list": class_list,
            "article_list": article_list,
            "rows": compressed_rows
        }
        
        out_path = os.path.join(scratch_dir, f"aggregated_{chain_name}_dashboard_data_v2.json")
        with open(out_path, "w", encoding="utf-8") as f_out:
            json.dump(out_dict, f_out, ensure_ascii=False, separators=(',', ':'))
            
        print(f"  Saved JSON to: {out_path} ({os.path.getsize(out_path)/(1024*1024):.2f} MB)", flush=True)
        print(f"  Finished chain {chain_name} in {time.time()-t_start:.1f}s", flush=True)


# ==========================================
# STEP 3: Generate HTML files
# ==========================================
def generate_html_files():
    print("\n==================================================", flush=True)
    print("STEP 3: Re-generating HTML Dashboards...", flush=True)
    print("==================================================", flush=True)
    
    chains_meta = [
        ("win", "Win", "dashboard_win.html"),
        ("rural", "Rural", "dashboard_rural.html"),
        ("urban", "Urban", "dashboard_urban.html")
    ]
    
    for name_lower, title, html_name in chains_meta:
        t0 = time.time()
        json_path = os.path.join(scratch_dir, f"aggregated_{name_lower}_dashboard_data_v2.json")
        output_html_path = os.path.join(folder, html_name)
        
        print(f"Generating HTML for {title} -> {html_name}...", flush=True)
        
        if not os.path.exists(json_path):
            print(f"  Error: JSON not found for {title}", flush=True)
            continue
            
        with open(json_path, 'r', encoding='utf-8') as f:
            db_data = json.load(f)
            
        with open(template_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
            
        # Title and Badge updates
        html_content = html_content.replace(
            "<title>Báo Cáo Đối Soát Dữ Liệu Kiểm Tra - Chuỗi WIN</title>",
            f"<title>Báo Cáo Đối Soát Dữ Liệu Kiểm Tra - Chuỗi {title}</title>"
        )
        html_content = html_content.replace(
            '<span class="px-2.5 py-0.5 bg-indigo-500/10 text-indigo-400 border border-indigo-500/20 text-xs font-bold rounded-full">Chuỗi WIN</span>',
            f'<span class="px-2.5 py-0.5 bg-indigo-500/10 text-indigo-400 border border-indigo-500/20 text-xs font-bold rounded-full">Chuỗi {title}</span>'
        )
        html_content = html_content.replace(
            '"Bao_cao_chi_tiet_kiem_tra_kho_Win.csv"',
            f'"Bao_cao_chi_tiet_kiem_tra_kho_{title}.csv"'
        )
        
        js_injection = f"const dbData = {json.dumps(db_data, ensure_ascii=False, separators=(',', ':'))};\n"
        js_injection += "const rawData = dbData.rows.map(r => ({\n"
        js_injection += "    store_code: dbData.store_list[r[0]][0],\n"
        js_injection += "    store_name: dbData.store_list[r[0]][1],\n"
        js_injection += "    rsm: dbData.rsm_list[r[1]],\n"
        js_injection += "    rsm_idx: r[1],\n"
        js_injection += "    asm: dbData.asm_list[r[2]],\n"
        js_injection += "    asm_idx: r[2],\n"
        js_injection += "    mch2: dbData.mch2_list[r[3]],\n"
        js_injection += "    mch2_idx: r[3],\n"
        js_injection += "    mch3: dbData.mch3_list[r[4]],\n"
        js_injection += "    mch3_idx: r[4],\n"
        js_injection += "    mch4: dbData.mch4_list[r[5]],\n"
        js_injection += "    mch4_idx: r[5],\n"
        js_injection += "    mch5: dbData.mch5_list[r[6]],\n"
        js_injection += "    mch5_idx: r[6],\n"
        js_injection += "    article_code: dbData.article_list[r[7]][0],\n"
        js_injection += "    article_name: dbData.article_list[r[7]][1],\n"
        js_injection += "    check_class: dbData.class_list[r[8]],\n"
        js_injection += "    class_idx: r[8],\n"
        js_injection += "    qty: r[9],\n"
        js_injection += "    value: r[10]\n"
        js_injection += "}));"
        
        html_content = html_content.replace('// DATA_PLACEHOLDER', js_injection)
        
        with open(output_html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
            
        print(f"  Generated. File size: {os.path.getsize(output_html_path)/(1024*1024):.2f} MB, Time: {time.time()-t0:.1f}s", flush=True)


# ==========================================
# STEP 4: Format Excel Files
# ==========================================
def format_excel_files():
    print("\n==================================================", flush=True)
    print("STEP 4: Styling & Formatting Merged Excel files...", flush=True)
    print("==================================================", flush=True)
    
    font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    fill_header = PatternFill(fill_type='solid', start_color='1F4E78', end_color='1F4E78')
    align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    
    for chain_name in ["Win", "Rural", "Urban"]:
        t_start = time.time()
        filename = f"Tong_hop_Kiem_tra_{chain_name}.xlsx"
        filepath = os.path.join(folder, filename)
        temppath = os.path.join(folder, f"temp_format_{filename}")
        
        print(f"\nFormatting: {filename}...", flush=True)
        if not os.path.exists(filepath):
            print(f"  Error: {filepath} not found.", flush=True)
            continue
            
        if os.path.exists(temppath):
            try:
                os.remove(temppath)
            except:
                pass
                
        wb_in = openpyxl.load_workbook(filepath, read_only=True)
        ws_in = wb_in.active
        
        wb_out = openpyxl.Workbook(write_only=True)
        ws_out = wb_out.create_sheet(title=ws_in.title)
        ws_out.views.sheetView[0].showGridLines = True
        
        col_widths = {}
        row_count = 0
        
        for row_idx, row in enumerate(ws_in.iter_rows(values_only=True), 1):
            row_count += 1
            if row_idx % 100000 == 0:
                print(f"  Formatted {row_idx:,} rows...", flush=True)
                
            cells_out = []
            
            for col_idx, val in enumerate(row):
                if col_idx == 0:
                    continue # Skip Column A
                    
                new_col_idx = col_idx - 1
                
                # Format headers (row 1 & 2)
                if row_idx <= 2:
                    if row_idx == 1 and new_col_idx in (23, 30):
                        val = None
                    cell = WriteOnlyCell(ws_out, value=val)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_header
                    cell.border = border_thin
                    cells_out.append(cell)
                else:
                    # Data cells formatting (only wrap for formats to keep it fast)
                    if new_col_idx in (11, 21, 22): # Date
                        date_val = parse_date_value(val)
                        if isinstance(date_val, (datetime, float)) or hasattr(date_val, 'strftime'):
                            cell = WriteOnlyCell(ws_out, value=date_val)
                            cell.number_format = 'dd/mm/yyyy'
                            cells_out.append(cell)
                        else:
                            cells_out.append(date_val)
                            
                    elif new_col_idx in (16, 17, 18, 20): # Currency
                        try:
                            num_val = float(val) if val is not None else None
                            if num_val is not None:
                                cell = WriteOnlyCell(ws_out, value=num_val)
                                cell.number_format = '#,##0'
                                cells_out.append(cell)
                            else:
                                cells_out.append(None)
                        except ValueError:
                            cells_out.append(val)
                            
                    elif new_col_idx in (25, 26): # Floats
                        try:
                            num_val = float(val) if val is not None else None
                            if num_val is not None:
                                cell = WriteOnlyCell(ws_out, value=num_val)
                                cell.number_format = '0.0'
                                cells_out.append(cell)
                            else:
                                cells_out.append(None)
                        except ValueError:
                            cells_out.append(val)
                            
                    elif new_col_idx in (12, 13, 14, 15, 19, 23, 24): # Int or float
                        try:
                            num_val = float(val) if val is not None else None
                            if num_val is not None:
                                if num_val == int(num_val):
                                    num_val = int(num_val)
                                    cell = WriteOnlyCell(ws_out, value=num_val)
                                    cell.number_format = '#,##0'
                                else:
                                    cell = WriteOnlyCell(ws_out, value=num_val)
                                    cell.number_format = '0.0'
                                cells_out.append(cell)
                            else:
                                cells_out.append(None)
                        except ValueError:
                            cells_out.append(val)
                    else:
                        cells_out.append(val)
                
                # Column widths (first 2000 rows only)
                if row_idx <= 2000:
                    val_str = str(val) if val is not None else ""
                    str_len = len(val_str)
                    if new_col_idx not in col_widths:
                        col_widths[new_col_idx] = str_len
                    else:
                        col_widths[new_col_idx] = max(col_widths[new_col_idx], str_len)
                        
            ws_out.append(cells_out)
            
        wb_in.close()
        
        # Apply column widths
        for col_idx, max_len in col_widths.items():
            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
            ws_out.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        wb_out.save(temppath)
        
        if os.path.exists(filepath):
            os.remove(filepath)
        os.rename(temppath, filepath)
        print(f"  Finished formatting in {time.time()-t_start:.1f}s. Rows: {row_count:,}", flush=True)


# ==========================================
# MAIN CONTROL PIPELINE
# ==========================================
def main():
    t_start = time.time()
    
    # 1. Merge Excel files (filtering out SAP result rows)
    merge_excel_files()
    
    # 2. Extract and prepare JSON data for dashboards
    product_mch5_map = load_master_articles()
    prepare_json_data(product_mch5_map)
    
    # 3. Generate HTML dashboards
    generate_html_files()
    
    # 4. Format/style the Excel files
    format_excel_files()
    
    print(f"\n==================================================", flush=True)
    print(f"PIPELINE COMPLETE: Cleaned & formatted all files successfully in {time.time()-t_start:.1f}s!", flush=True)
    print("==================================================", flush=True)

if __name__ == '__main__':
    main()
