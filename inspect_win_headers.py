import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

folder = r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu"
win_file = os.path.join(folder, "Tong_hop_Kiem_tra_Win.xlsx")

wb = openpyxl.load_workbook(win_file, read_only=True)
sheet = wb.active
for r_idx, row in enumerate(sheet.iter_rows(max_row=3, values_only=True)):
    print(f"Row {r_idx+1}:")
    for c_idx, val in enumerate(row):
        col_letter = openpyxl.utils.get_column_letter(c_idx+1)
        if val is not None:
            print(f"  {col_letter}: {val}")
wb.close()
