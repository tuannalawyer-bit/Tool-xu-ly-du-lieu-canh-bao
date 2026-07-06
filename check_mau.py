import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

path = r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu\raw mẫu\Tổng hợp mẫu.xlsx"
try:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print("Sheets in Tổng hợp mẫu.xlsx:", wb.sheetnames)
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        print(f"Sheet: {sheetname}")
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= 5:
                break
            print(f"  Row {i}: {row[:15]}")
except Exception as e:
    print(f"Error: {e}")
