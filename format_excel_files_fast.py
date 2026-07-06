import os
import sys
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell import WriteOnlyCell
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

folder = r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu"
files_to_format = [
    "Tong_hop_Kiem_tra_Win.xlsx",
    "Tong_hop_Kiem_tra_Rural.xlsx",
    "Tong_hop_Kiem_tra_Urban.xlsx"
]

def parse_date_value(val):
    if val is None:
        return None
    val_str = str(val).strip()
    if not val_str:
        return None
    for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    try:
        float_val = float(val_str)
        if 35000 < float_val < 60000:
            return float_val
    except ValueError:
        pass
    return val

def format_excel_file(filename):
    t_start = time.time()
    filepath = os.path.join(folder, filename)
    temppath = os.path.join(folder, f"temp_format_{filename}")
    
    print(f"\nProcessing: {filename}...")
    
    if not os.path.exists(filepath):
        print(f"  Error: {filepath} not found.")
        return
        
    # Clean up old temp file if exists
    if os.path.exists(temppath):
        try:
            os.remove(temppath)
        except:
            pass
            
    wb_in = openpyxl.load_workbook(filepath, read_only=True)
    ws_in = wb_in.active
    
    wb_out = openpyxl.Workbook(write_only=True)
    ws_out = wb_out.create_sheet(title=ws_in.title)
    
    # Enable grid lines explicitly in the output sheet
    ws_out.views.sheetView[0].showGridLines = True
    
    # Styles for header
    font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    fill_header = PatternFill(fill_type='solid', start_color='1F4E78', end_color='1F4E78') # Steel blue
    align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    col_widths = {}
    row_count = 0
    
    # Process rows
    for row_idx_1based, row in enumerate(ws_in.iter_rows(values_only=True), 1):
        row_count += 1
        if row_idx_1based % 100000 == 0:
            print(f"  Processed {row_idx_1based:,} rows...")
            
        cells_out = []
        
        # We skip Column A (index 0) which is "map"
        for col_idx_0based, val in enumerate(row):
            if col_idx_0based == 0:
                continue # Skip Column A
                
            new_col_idx = col_idx_0based - 1 # 0-based index for the output sheet
            
            # Format row 1 and 2 as headers
            if row_idx_1based <= 2:
                # Clear Y1 and AF1 (which are index 23 and 30 in the new 0-based column mapping)
                if row_idx_1based == 1 and new_col_idx in (23, 30):
                    val = None
                    
                cell = WriteOnlyCell(ws_out, value=val)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_header
                cell.border = border_thin
                cells_out.append(cell)
                
            else:
                # Data rows formatting
                # We only wrap in WriteOnlyCell if we need to apply a number format.
                # Otherwise, we just append the raw value which is 10x faster!
                
                # 1. Dates: Created on (11), Last GR (21), Last Sale (22)
                if new_col_idx in (11, 21, 22):
                    date_val = parse_date_value(val)
                    if isinstance(date_val, (datetime, float)) or hasattr(date_val, 'strftime'):
                        cell = WriteOnlyCell(ws_out, value=date_val)
                        cell.number_format = 'dd/mm/yyyy'
                        cells_out.append(cell)
                    else:
                        cells_out.append(date_val)
                        
                # 2. Currency: Giá trị tồn (16), Doanh thu (17), Giá vốn (18), Giá tồn D-15 (20)
                elif new_col_idx in (16, 17, 18, 20):
                    try:
                        num_val = float(val) if val is not None else None
                        if num_val is not None:
                            cell = WriteOnlyCell(ws_out, value=num_val)
                            cell.number_format = '#,##0'
                            cells_out.append(cell)
                        else:
                            cells_out.append(None)
                    except ValueError:
                        cells_out.append(val)
                        
                # 3. Float Decimals: DIO (25), DIO/Date (26)
                elif new_col_idx in (25, 26):
                    try:
                        num_val = float(val) if val is not None else None
                        if num_val is not None:
                            cell = WriteOnlyCell(ws_out, value=num_val)
                            cell.number_format = '0.0'
                            cells_out.append(cell)
                        else:
                            cells_out.append(None)
                    except ValueError:
                        cells_out.append(val)
                        
                # 4. Standard Integers (Quantities etc.):
                elif new_col_idx in (12, 13, 14, 15, 19, 23, 24):
                    try:
                        num_val = float(val) if val is not None else None
                        if num_val is not None:
                            if num_val == int(num_val):
                                num_val = int(num_val)
                                cell = WriteOnlyCell(ws_out, value=num_val)
                                cell.number_format = '#,##0'
                            else:
                                cell = WriteOnlyCell(ws_out, value=num_val)
                                cell.number_format = '0.0'
                            cells_out.append(cell)
                        else:
                            cells_out.append(None)
                    except ValueError:
                        cells_out.append(val)
                else:
                    # Text columns, store codes, article codes, etc. - write raw values!
                    cells_out.append(val)
            
            # Compute width for column (measure first 2000 rows to avoid slow performance)
            if row_idx_1based <= 2000:
                val_str = str(val) if val is not None else ""
                str_len = len(val_str)
                if new_col_idx not in col_widths:
                    col_widths[new_col_idx] = str_len
                else:
                    col_widths[new_col_idx] = max(col_widths[new_col_idx], str_len)
                    
        ws_out.append(cells_out)
        
    wb_in.close()
    
    # Apply column widths
    print("  Applying column widths...")
    for col_idx, max_len in col_widths.items():
        col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
        ws_out.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    print("  Saving file...")
    wb_out.save(temppath)
    
    # Replace original file with formatted file
    if os.path.exists(filepath):
        os.remove(filepath)
    os.rename(temppath, filepath)
    
    print(f"  Completed in {time.time()-t_start:.1f}s. Total rows: {row_count:,}")

def main():
    t_all = time.time()
    for f in files_to_format:
        format_excel_file(f)
    print(f"\nAll files formatted successfully in {time.time()-t_all:.1f}s!")

if __name__ == '__main__':
    main()
