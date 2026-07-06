import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

folder = r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu"
files = [f for f in os.listdir(folder) if f.endswith(".xlsx")]

path = os.path.join(folder, files[0])
print(f"File: {files[0]}")
try:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = wb['Chi tiết data']
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 1:
            print("Headers:")
            for idx, val in enumerate(row):
                print(f"Col {idx} ({openpyxl.utils.get_column_letter(idx+1)}): {val}")
            break
except Exception as e:
    print(f"Error: {e}")
