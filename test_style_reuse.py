import openpyxl
from openpyxl.cell import WriteOnlyCell
import time

wb = openpyxl.Workbook(write_only=True)
ws = wb.create_sheet()

# Method 1: New cell every time
t0 = time.time()
for i in range(20000):
    row = []
    for j in range(10):
        cell = WriteOnlyCell(ws, value=i*j)
        cell.number_format = '#,##0'
        row.append(cell)
    ws.append(row)
print(f"Time with new cells: {time.time() - t0:.2f}s")

# Method 2: Reused cells
wb2 = openpyxl.Workbook(write_only=True)
ws2 = wb2.create_sheet()

cell_template = WriteOnlyCell(ws2)
cell_template.number_format = '#,##0'

t0 = time.time()
for i in range(20000):
    row = []
    for j in range(10):
        cell_template.value = i*j
        row.append(cell_template)
    ws2.append(row)
print(f"Time with reused cells: {time.time() - t0:.2f}s")
