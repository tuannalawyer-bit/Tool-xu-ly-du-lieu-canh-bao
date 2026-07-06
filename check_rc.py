import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

path = r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu\Bùi Anh Tuấn.xlsx"
try:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if 'RC' in wb.sheetnames:
        sheet = wb['RC']
        print("Sheet 'RC' rows:")
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= 10:
                break
            print(f"Row {i}: {row[:15]}")
    else:
        print("Sheet 'RC' not found")
except Exception as e:
    print(f"Error: {e}")
