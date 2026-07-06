import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

path = r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu\raw mẫu\6361 last sale last GR.xlsx"
try:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if 'Data TCX' in wb.sheetnames:
        sheet = wb['Data TCX']
        print("Sheet 'Data TCX' rows:")
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= 10:
                break
            print(f"Row {i}: {row[:15]}")
    else:
        print("Sheet 'Data TCX' not found")
except Exception as e:
    print(f"Error: {e}")
