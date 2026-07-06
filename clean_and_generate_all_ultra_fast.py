import os
import sys
import time
import zipfile
import xml.etree.ElementTree as ET
import ctypes
import concurrent.futures
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell import WriteOnlyCell
from pyxlsb import open_workbook
from datetime import datetime
import shutil

# Force unbuffered output
sys.stdout.reconfigure(encoding='utf-8')

folder = r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu"
scratch_dir = r"C:\Users\Tuan\.gemini\antigravity\brain\b053fc36-8fe6-4316-bba8-ac8fef252caf\scratch"
xlsb_path = os.path.join(folder, "raw mẫu", "updated_data_moi.xlsb")
template_path = os.path.join(scratch_dir, "dashboard_win_template.html")

# Define files for all chains
chains = {
    "win": [
        "Nguyễn Đức Thiên Ân.xlsx",
        "Đỗ Thị Thanh Loan.xlsx",
        "Lê Văn Trí.xlsx",
        "Nguyễn minh Trang.xlsx"
    ],
    "rural": [
        "Lê Thị Hồng Thu.xlsx",
        "Hoàng Nguyễn Tú Anh.xlsx",
        "Bùi Anh Tuấn.xlsx",
        "Lạc Nhật Minh.xlsx"
    ],
    "urban": [
        "Phạm Thị Ngọc Xuyến.xlsx",
        "Lê Duy Đức.xlsx",
        "Vương Phi Sơn.xlsx",
        "Trần Thị Diệp.xlsx",
        "Đỗ Khắc Chức.xlsx",
        "NGuyễn Văn Tuấn.xlsx"
    ]
}

def normalize_rsm_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    name_lower = name_str.lower()
    if name_lower == "nguyễn văn tuấn":
        return "Nguyễn Văn Tuấn"
    if name_lower == "nguyễn minh trang":
        return "Nguyễn Minh Trang"
    return name_str

def col_to_idx(col):
    val = 0
    for char in col:
        val = val * 26 + (ord(char.upper()) - 64)
    return val - 1

def parse_date_value(val):
    if val is None:
        return None
    val_str = str(val).strip()
    if not val_str:
        return None
    for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    try:
        float_val = float(val_str)
        if 35000 < float_val < 60000:
            return float_val
    except ValueError:
        pass
    return val

def load_master_articles():
    print("\nLoading master articles from XLSB...", flush=True)
    t0 = time.time()
    mapping = {}
    try:
        with open_workbook(xlsb_path) as wb:
            with wb.get_sheet('master article') as sheet:
                for i, row in enumerate(sheet.rows()):
                    if i == 0:
                        continue
                    prod_id = row[0].v
                    mch5_desc = row[22].v if len(row) > 22 else None
                    if prod_id:
                        prod_id_str = str(int(prod_id)) if isinstance(prod_id, float) else str(prod_id).strip()
                        mapping[prod_id_str] = str(mch5_desc).strip() if mch5_desc else "Khác"
        print(f"Loaded {len(mapping)} articles in {time.time()-t0:.2f}s", flush=True)
    except Exception as e:
        print(f"Error loading master articles: {e}", flush=True)
    return mapping

# ==========================================
# PARALLEL EXTRACTOR WORKER
# ==========================================
def extract_file_data(args):
    filename, product_mch5_map, chain_name = args
    src_path = os.path.join(folder, filename)
    temp_path = os.path.join(scratch_dir, f"temp_extract_{chain_name}_{filename}")
    
    t_start = time.time()
    try:
        res = ctypes.windll.kernel32.CopyFileW(src_path, temp_path, False)
        if not res:
            return {"filename": filename, "status": "error", "message": "Copy failed"}
            
        with zipfile.ZipFile(temp_path, 'r') as z:
            wb_data = z.read('xl/workbook.xml')
            root_wb = ET.fromstring(wb_data)
            ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            
            sheets = []
            for sheet in root_wb.findall('.//ns:sheet', ns):
                sheets.append((sheet.attrib.get('name'), sheet.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')))
                
            sheet_name, target_rid = [s for s in sheets if s[0] not in ['_com.sap.ip.bi.xl.hiddensheet', 'RC']][0]
            
            rels_data = z.read('xl/_rels/workbook.xml.rels')
            root_rels = ET.fromstring(rels_data)
            r_ns = {'rels': 'http://schemas.openxmlformats.org/package/2006/relationships'}
            
            sheet_path_in_zip = None
            for rel in root_rels.findall('.//rels:Relationship', r_ns):
                if rel.attrib.get('Id') == target_rid:
                    sheet_path_in_zip = f"xl/{rel.attrib.get('Target')}"
                    break
                    
            shared_strings = []
            if 'xl/sharedStrings.xml' in z.namelist():
                sst_file = z.open('xl/sharedStrings.xml')
                context_sst = ET.iterparse(sst_file, events=('start', 'end'))
                for event, elem in context_sst:
                    if event == 'end' and elem.tag.endswith('}si'):
                        text = "".join([t.text for t in elem.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t') if t.text is not None])
                        shared_strings.append(text)
                        elem.clear()
            
            sheet_file = z.open(sheet_path_in_zip)
            context = ET.iterparse(sheet_file, events=('start', 'end'))
            
            rows_list = []
            current_row = {}
            current_row_idx = None
            
            def get_val(cell_elem):
                t_type = cell_elem.attrib.get('t')
                v_elem = cell_elem.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                if v_elem is not None and v_elem.text is not None:
                    val_text = v_elem.text
                    if t_type == 's':
                        idx = int(val_text)
                        if idx < len(shared_strings):
                            return shared_strings[idx]
                    if val_text.isdigit():
                        return int(val_text)
                    try:
                        return float(val_text)
                    except:
                        return val_text
                return None
                
            for event, elem in context:
                if event == 'start' and elem.tag.endswith('}row'):
                    current_row = {}
                    r_attr = elem.attrib.get('r')
                    current_row_idx = int(r_attr) if r_attr and r_attr.isdigit() else None
                    
                elif event == 'end' and elem.tag.endswith('}c'):
                    r_ref = elem.attrib.get('r')
                    if r_ref:
                        digit_idx = 0
                        for char in r_ref:
                            if char.isdigit():
                                break
                            digit_idx += 1
                        current_row[r_ref[:digit_idx]] = get_val(elem)
                        
                elif event == 'end' and elem.tag.endswith('}row'):
                    if current_row_idx and current_row_idx >= 3:
                        af_val = current_row.get('AF')
                        if af_val is not None:
                            af_str = str(af_val).strip()
                            if af_str != "" and af_str.upper() not in ["N/A", "#N/A"]:
                                store_code = str(current_row.get('D', '')).strip()
                                store_name = str(current_row.get('E', '')).strip()
                                
                                # Skip Result rows
                                if "result" in store_code.lower() or "result" in store_name.lower():
                                    continue
                                    
                                row_list = [None] * 32
                                for col_letter, val in current_row.items():
                                    idx = col_to_idx(col_letter)
                                    if idx < 32:
                                        row_list[idx] = val
                                        
                                row_list[1] = normalize_rsm_name(row_list[1])
                                rows_list.append(row_list)
                    elem.clear()
            return {"filename": filename, "status": "success", "data": rows_list, "time_taken": time.time() - t_start}
    except Exception as e:
        return {"filename": filename, "status": "error", "message": str(e)}
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass

# ==========================================
# PIPELINE STEP PROCESSING
# ==========================================
def process_chain(chain_name, product_mch5_map, workers):
    print(f"\n==================================================", flush=True)
    print(f"PROCESSING CHAIN: {chain_name.upper()}", flush=True)
    print(f"==================================================", flush=True)
    t_start = time.time()
    
    # 1. Parallel extract data
    file_list = chains[chain_name]
    print(f"Extracting data in parallel from {len(file_list)} files...", flush=True)
    
    all_data_rows = []
    with concurrent.futures.ProcessPoolExecutor(max_workers=workers) as executor:
        args_list = [(f, product_mch5_map, chain_name) for f in file_list]
        futures = {executor.submit(extract_file_data, args): args[0] for args in args_list}
        for future in concurrent.futures.as_completed(futures):
            f = futures[future]
            res = future.result()
            if res["status"] == "success":
                all_data_rows.extend(res["data"])
                print(f"  Processed {f} in {res['time_taken']:.1f}s: rows={len(res['data']):,}", flush=True)
            else:
                print(f"  Error processing {f}: {res['message']}", flush=True)
                
    total_rows = len(all_data_rows)
    print(f"Total clean rows extracted: {total_rows:,}", flush=True)
    
    # 2. Write Formatted Merged Excel File (ONE-PASS WRITE & FORMAT - LOCAL WRITES)
    print("Writing formatted merged Excel file directly to local temp...", flush=True)
    t_excel = time.time()
    
    excel_filename = f"Tong_hop_Kiem_tra_{chain_name.capitalize()}.xlsx"
    excel_filepath = os.path.join(folder, excel_filename)
    excel_temppath = os.path.join(scratch_dir, f"temp_write_{excel_filename}")
    
    if os.path.exists(excel_temppath):
        try: os.remove(excel_temppath)
        except: pass
        
    wb_out = openpyxl.Workbook(write_only=True)
    ws_out = wb_out.create_sheet(title="Dữ liệu kiểm tra")
    ws_out.views.sheetView[0].showGridLines = True
    
    # Header styles
    font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    fill_header = PatternFill(fill_type='solid', start_color='1F4E78', end_color='1F4E78')
    align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Row headers
    header_row_1 = [None] * 32
    header_row_1[13] = "Opening Quantity"
    header_row_1[14] = "GR Quantity"
    header_row_1[15] = "GI Quantity"
    header_row_1[16] = "Closing Quantity"
    header_row_1[17] = "Closing Value"
    header_row_1[18] = "Revenue"
    header_row_1[19] = "COGS"
    header_row_1[20] = "D-15"
    header_row_1[24] = "46204"
    header_row_1[31] = "100000"

    header_row_2 = [
        "map", "RSM", "ASM", "Store", "Store", "MCH2 - Department", "MCH2 - Department",
        "MCH3 - Category", "MCH3 - Category", "MCH4", "Article", "Article", "Created on",
        "Tồn đầu kỳ (D-90)", "Nhập trong kỳ", "Xuất trong kỳ", "Tồn cuối kỳ (D)",
        "Giá trị tồn", "Doanh thu", "Giá vốn", "Tồn D-15", "Giá tồn D-15",
        "Last GR", "Last Sale", "Số ngày ko nhập", "Date tham khảo", "DIO (D-15)",
        "DIO/Date", "Note", "Note check slow", "note hết hạn", "Phân loại kiểm tra"
    ]
    
    # Write Row 1
    cells_out = []
    for col_idx in range(1, 32):
        new_col_idx = col_idx - 1
        val = header_row_1[col_idx]
        if new_col_idx in (23, 30):
            val = None
        cell = WriteOnlyCell(ws_out, value=val)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_thin
        cells_out.append(cell)
    ws_out.append(cells_out)
    
    # Write Row 2
    cells_out = []
    for col_idx in range(1, 32):
        val = header_row_2[col_idx]
        cell = WriteOnlyCell(ws_out, value=val)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_thin
        cells_out.append(cell)
    ws_out.append(cells_out)
    
    col_widths = {}
    
    # Write Data Rows
    for row_idx_0based, row in enumerate(all_data_rows):
        cells_out = []
        for col_idx, val in enumerate(row):
            if col_idx == 0:
                continue # Skip Column A
                
            new_col_idx = col_idx - 1
            
            # Format/wrapping logic
            if new_col_idx in (11, 21, 22): # Dates
                date_val = parse_date_value(val)
                if isinstance(date_val, (datetime, float)) or hasattr(date_val, 'strftime'):
                    cell = WriteOnlyCell(ws_out, value=date_val)
                    cell.number_format = 'dd/mm/yyyy'
                    cells_out.append(cell)
                else:
                    cells_out.append(date_val)
                    
            elif new_col_idx in (16, 17, 18, 20): # Currency
                try:
                    num_val = float(val) if val is not None else None
                    if num_val is not None:
                        cell = WriteOnlyCell(ws_out, value=num_val)
                        cell.number_format = '#,##0'
                        cells_out.append(cell)
                    else:
                        cells_out.append(None)
                except ValueError:
                    cells_out.append(val)
                    
            elif new_col_idx in (25, 26): # Float Decimals
                try:
                    num_val = float(val) if val is not None else None
                    if num_val is not None:
                        cell = WriteOnlyCell(ws_out, value=num_val)
                        cell.number_format = '0.0'
                        cells_out.append(cell)
                    else:
                        cells_out.append(None)
                except ValueError:
                    cells_out.append(val)
                    
            elif new_col_idx in (12, 13, 14, 15, 19, 23, 24): # Int or float
                try:
                    num_val = float(val) if val is not None else None
                    if num_val is not None:
                        if num_val == int(num_val):
                            num_val = int(num_val)
                            cell = WriteOnlyCell(ws_out, value=num_val)
                            cell.number_format = '#,##0'
                        else:
                            cell = WriteOnlyCell(ws_out, value=num_val)
                            cell.number_format = '0.0'
                        cells_out.append(cell)
                    else:
                        cells_out.append(None)
                except ValueError:
                    cells_out.append(val)
            else:
                cells_out.append(val)
                
            # Measure widths (first 2000 rows only)
            if row_idx_0based <= 2000:
                val_str = str(val) if val is not None else ""
                str_len = len(val_str)
                if new_col_idx not in col_widths:
                    col_widths[new_col_idx] = str_len
                else:
                    col_widths[new_col_idx] = max(col_widths[new_col_idx], str_len)
                    
        ws_out.append(cells_out)
        
    # Apply column widths
    for col_idx, max_len in col_widths.items():
        col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
        ws_out.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Save formatted Excel locally first
    wb_out.save(excel_temppath)
    
    # Safe copy to destination OneDrive folder with retry on lock
    print(f"  Local file saved, copying to {excel_filepath}...", flush=True)
    copied = False
    for attempt in range(5):
        try:
            if os.path.exists(excel_filepath):
                os.remove(excel_filepath)
            shutil.copy2(excel_temppath, excel_filepath)
            copied = True
            print(f"  Formatted Excel saved and synced to: {excel_filepath} ({os.path.getsize(excel_filepath)/(1024*1024):.2f} MB)", flush=True)
            break
        except Exception as e:
            print(f"  Attempt {attempt+1} to write {excel_filepath} failed: {e}. Retrying in 3s...", flush=True)
            time.sleep(3)
            
    if not copied:
        print(f"  ERROR: Could not save final file to OneDrive folder. Local file preserved at {excel_temppath}", flush=True)
        
    # Clean up local temp file
    if copied and os.path.exists(excel_temppath):
        try: os.remove(excel_temppath)
        except: pass
        
    print(f"  Finished Excel merged write in {time.time()-t_excel:.1f}s", flush=True)
    
    # 3. Prepare Dashboard JSON Data
    print("Preparing JSON v2 dashboard data from memory...", flush=True)
    t_json = time.time()
    
    rsms = set()
    asms = set()
    stores = {}
    mch2s = set()
    mch3s = set()
    mch4s = set()
    mch5s = set()
    classes = set()
    articles = {}
    
    for row in all_data_rows:
        rsms.add(row[1])
        asms.add(row[2])
        stores[row[3]] = row[4]
        mch2s.add(row[6])
        mch3s.add(row[8])
        mch4s.add(row[9])
        
        art_code = str(row[10]).strip()
        art_name = str(row[11]).strip()
        mch5 = product_mch5_map.get(art_code, "Khác")
        mch5s.add(mch5)
        classes.add(row[31])
        
        art_key = (art_code, art_name)
        if art_key not in articles:
            articles[art_key] = len(articles)
            
    rsm_list = sorted(list(rsms))
    asm_list = sorted(list(asms))
    store_list = [[code, name] for code, name in sorted(stores.items())]
    mch2_list = sorted(list(mch2s))
    mch3_list = sorted(list(mch3s))
    mch4_list = sorted(list(mch4s))
    mch5_list = sorted(list(mch5s))
    class_list = sorted(list(classes))
    
    article_list = [None] * len(articles)
    for (code, name), idx in articles.items():
        article_list[idx] = [code, name]
        
    rsm_map = {name: idx for idx, name in enumerate(rsm_list)}
    asm_map = {name: idx for idx, name in enumerate(asm_list)}
    store_map = {item[0]: idx for idx, item in enumerate(store_list)}
    mch2_map = {name: idx for idx, name in enumerate(mch2_list)}
    mch3_map = {name: idx for idx, name in enumerate(mch3_list)}
    mch4_map = {name: idx for idx, name in enumerate(mch4_list)}
    mch5_map = {name: idx for idx, name in enumerate(mch5_list)}
    class_map = {name: idx for idx, name in enumerate(class_list)}
    
    compressed_rows = []
    for row in all_data_rows:
        art_code = str(row[10]).strip()
        art_name = str(row[11]).strip()
        art_key = (art_code, art_name)
        mch5 = product_mch5_map.get(art_code, "Khác")
        
        qty_val = float(row[16]) if row[16] is not None else 0.0
        val_val = float(row[17]) if row[17] is not None else 0.0
        
        comp_row = [
            store_map[row[3]],
            rsm_map[row[1]],
            asm_map[row[2]],
            mch2_map[row[6]],
            mch3_map[row[8]],
            mch4_map[row[9]],
            mch5_map[mch5],
            articles[art_key],          # article_idx
            class_map[row[31]],         # class_idx
            round(qty_val),
            int(val_val)
        ]
        compressed_rows.append(comp_row)
        
    out_dict = {
        "rsm_list": rsm_list,
        "asm_list": asm_list,
        "store_list": store_list,
        "mch2_list": mch2_list,
        "mch3_list": mch3_list,
        "mch4_list": mch4_list,
        "mch5_list": mch5_list,
        "class_list": class_list,
        "article_list": article_list,
        "rows": compressed_rows
    }
    
    json_filename = f"aggregated_{chain_name}_dashboard_data_v2.json"
    json_path = os.path.join(scratch_dir, json_filename)
    with open(json_path, "w", encoding="utf-8") as f_out:
        json.dump(out_dict, f_out, ensure_ascii=False, separators=(',', ':'))
        
    print(f"  JSON v2 saved: {json_path} ({os.path.getsize(json_path)/(1024*1024):.2f} MB) in {time.time()-t_json:.1f}s", flush=True)
    
    # 4. Generate HTML Dashboard file
    print("Generating HTML dashboard...", flush=True)
    t_html = time.time()
    
    html_filename = f"dashboard_{chain_name}.html"
    output_html_path = os.path.join(folder, html_filename)
    
    with open(template_path, 'r', encoding='utf-8') as f:
        html_content = f.read()
        
    # Title and Badge updates
    title = chain_name.capitalize()
    html_content = html_content.replace(
        "<title>Báo Cáo Đối Soát Dữ Liệu Kiểm Tra - Chuỗi WIN</title>",
        f"<title>Báo Cáo Đối Soát Dữ Liệu Kiểm Tra - Chuỗi {title}</title>"
    )
    html_content = html_content.replace(
        '<span class="px-2.5 py-0.5 bg-indigo-500/10 text-indigo-400 border border-indigo-500/20 text-xs font-bold rounded-full">Chuỗi WIN</span>',
        f'<span class="px-2.5 py-0.5 bg-indigo-500/10 text-indigo-400 border border-indigo-500/20 text-xs font-bold rounded-full">Chuỗi {title}</span>'
    )
    html_content = html_content.replace(
        '"Bao_cao_chi_tiet_kiem_tra_kho_Win.csv"',
        f'"Bao_cao_chi_tiet_kiem_tra_kho_{title}.csv"'
    )
    
    js_injection = f"const dbData = {json.dumps(out_dict, ensure_ascii=False, separators=(',', ':'))};\n"
    js_injection += "const rawData = dbData.rows.map(r => ({\n"
    js_injection += "    store_code: dbData.store_list[r[0]][0],\n"
    js_injection += "    store_name: dbData.store_list[r[0]][1],\n"
    js_injection += "    rsm: dbData.rsm_list[r[1]],\n"
    js_injection += "    rsm_idx: r[1],\n"
    js_injection += "    asm: dbData.asm_list[r[2]],\n"
    js_injection += "    asm_idx: r[2],\n"
    js_injection += "    mch2: dbData.mch2_list[r[3]],\n"
    js_injection += "    mch2_idx: r[3],\n"
    js_injection += "    mch3: dbData.mch3_list[r[4]],\n"
    js_injection += "    mch3_idx: r[4],\n"
    js_injection += "    mch4: dbData.mch4_list[r[5]],\n"
    js_injection += "    mch4_idx: r[5],\n"
    js_injection += "    mch5: dbData.mch5_list[r[6]],\n"
    js_injection += "    mch5_idx: r[6],\n"
    js_injection += "    article_code: dbData.article_list[r[7]][0],\n"
    js_injection += "    article_name: dbData.article_list[r[7]][1],\n"
    js_injection += "    check_class: dbData.class_list[r[8]],\n"
    js_injection += "    class_idx: r[8],\n"
    js_injection += "    qty: r[9],\n"
    js_injection += "    value: r[10]\n"
    js_injection += "}));"
    
    html_content = html_content.replace('// DATA_PLACEHOLDER', js_injection)
    
    # Safe write dashboard HTML
    with open(output_html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
        
    print(f"  HTML Dashboard saved: {output_html_path} ({os.path.getsize(output_html_path)/(1024*1024):.2f} MB) in {time.time()-t_html:.1f}s", flush=True)
    print(f"Finished processing chain {chain_name.upper()} in {time.time()-t_start:.1f}s!", flush=True)

# ==========================================
# MAIN PIPELINE CONTROLLER
# ==========================================
def main():
    t_all = time.time()
    
    # 1. Load product catalog mapping once
    product_mch5_map = load_master_articles()
    
    # Determine cpu cores to allocate workers
    workers = min(os.cpu_count() or 4, 4)
    
    # Process Win, Rural, and Urban chains one by one
    for chain_name in ["win", "rural", "urban"]:
        process_chain(chain_name, product_mch5_map, workers)
        
    print(f"\n==================================================", flush=True)
    print(f"ULTRA-FAST PIPELINE COMPLETE: Cleaned & formatted all chains successfully in {time.time()-t_all:.1f}s!", flush=True)
    print("==================================================", flush=True)

if __name__ == '__main__':
    main()
