import openpyxl
import os
import sys
import time
import ctypes

sys.stdout.reconfigure(encoding='utf-8')

folder = r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu"
scratch_dir = r"C:\Users\Tuan\.gemini\antigravity\brain\b053fc36-8fe6-4316-bba8-ac8fef252caf\scratch"
output_filename = "Tong_hop_du_lieu_kiem_tra.xlsx"
output_path = os.path.join(folder, output_filename)

def clean_sheet_title(title):
    # Excel sheet title limit is 31 chars
    # Remove chars like :, \, /, ?, *, [, ]
    for char in [':', '\\', '/', '?', '*', '[', ']']:
        title = title.replace(char, '')
    return title[:31]

def main():
    t_start = time.time()
    
    files = sorted([f for f in os.listdir(folder) if f.endswith(".xlsx") and f != output_filename])
    print(f"Found {len(files)} files to merge.")
    
    # Initialize a write-only workbook
    wb_out = openpyxl.Workbook(write_only=True)
    
    for idx, filename in enumerate(files):
        t_file_start = time.time()
        # Determine the RSM name from filename (without extension)
        rsm_name = os.path.splitext(filename)[0]
        sheet_title = clean_sheet_title(rsm_name)
        print(f"[{idx+1}/{len(files)}] Processing '{filename}' -> Sheet '{sheet_title}'...")
        
        src_path = os.path.join(folder, filename)
        temp_path = os.path.join(scratch_dir, f"temp_merge_{filename}")
        
        # Check if the file is locked / needs ctypes copying
        is_copied = False
        try:
            # Try opening the file to see if it is locked
            with open(src_path, 'rb') as f:
                pass
            path_to_open = src_path
        except PermissionError:
            # If locked, copy it to scratch folder using Win32 API
            print(f"  File is locked, copying to temp location...")
            res = ctypes.windll.kernel32.CopyFileW(src_path, temp_path, False)
            if not res:
                err = ctypes.windll.kernel32.GetLastError()
                print(f"  Error: Copy failed with Win32 error code {err}")
                continue
            path_to_open = temp_path
            is_copied = True
            
        try:
            wb_in = openpyxl.load_workbook(path_to_open, read_only=True, data_only=True)
            candidates = [s for s in wb_in.sheetnames if s not in ['_com.sap.ip.bi.xl.hiddensheet', 'RC']]
            if not candidates:
                print(f"  Error: No data sheets found in {filename}")
                wb_in.close()
                continue
            
            sheet = wb_in[candidates[0]]
            
            # Create a new sheet in the output workbook
            ws_out = wb_out.create_sheet(title=sheet_title)
            
            total_copied = 0
            # Read rows
            for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if r_idx == 0 or r_idx == 1:
                    # Append headers (Row 0 and Row 1)
                    ws_out.append(row)
                else:
                    # Check column AF (index 31)
                    val = row[31] if len(row) > 31 else None
                    if val is not None and str(val).strip() != "":
                        ws_out.append(row)
                        total_copied += 1
            
            wb_in.close()
            t_file_end = time.time()
            print(f"  Completed! Appended {total_copied} rows. Time taken: {t_file_end-t_file_start:.2f}s")
            
        except Exception as e:
            print(f"  Error processing {filename}: {e}")
        finally:
            if is_copied and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except:
                    pass
                    
    print("Saving the merged workbook...")
    t_save_start = time.time()
    wb_out.save(output_path)
    t_end = time.time()
    print(f"Successfully saved merged file to: {output_path}")
    print(f"Total time taken: {t_end-t_start:.2f}s (Save time: {t_end-t_save_start:.2f}s)")

if __name__ == '__main__':
    main()
