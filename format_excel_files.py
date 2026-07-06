import os
import sys
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell import WriteOnlyCell
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

folder = r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu"
files_to_format = [
    "Tong_hop_Kiem_tra_Win.xlsx",
    "Tong_hop_Kiem_tra_Rural.xlsx",
    "Tong_hop_Kiem_tra_Urban.xlsx"
]

def parse_date_value(val):
    if val is None:
        return None
    val_str = str(val).strip()
    if not val_str:
        return None
    for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    # Try parsing serial date if it's numeric and large (Excel dates are generally > 40000 currently)
    try:
        float_val = float(val_str)
        if 35000 < float_val < 60000:
            # Let openpyxl write it as float, excel will format it
            return float_val
    except ValueError:
        pass
    return val

def format_excel_file(filename):
    t_start = time.time()
    filepath = os.path.join(folder, filename)
    temppath = os.path.join(folder, f"temp_format_{filename}")
    
    print(f"\nProcessing: {filename}...")
    
    if not os.path.exists(filepath):
        print(f"  Error: {filepath} not found.")
        return
        
    wb_in = openpyxl.load_workbook(filepath, read_only=True)
    ws_in = wb_in.active
    
    wb_out = openpyxl.Workbook(write_only=True)
    ws_out = wb_out.create_sheet(title=ws_in.title)
    
    # Styles definition
    font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    fill_header = PatternFill(fill_type='solid', start_color='1F4E78', end_color='1F4E78') # Steel blue
    align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    font_data = Font(name='Arial', size=10)
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    
    col_widths = {}
    
    # We will read rows and write cell objects
    row_count = 0
    
    for row_idx_1based, row in enumerate(ws_in.iter_rows(values_only=True), 1):
        row_count += 1
        if row_idx_1based % 50000 == 0:
            print(f"  Processed {row_idx_1based} rows...")
            
        cells_out = []
        
        # We skip Column A (index 0) which is "map"
        for col_idx_0based, val in enumerate(row):
            if col_idx_0based == 0:
                continue # Skip Column A
                
            new_col_idx = col_idx_0based - 1 # 0-based index for the output sheet
            
            # Format row 1 and 2 as headers
            if row_idx_1based <= 2:
                # Clear Y1 and AF1 (which are index 23 and 30 in the new 0-based column mapping)
                if row_idx_1based == 1 and new_col_idx in (23, 30):
                    val = None
                    
                cell = WriteOnlyCell(ws_out, value=val)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_header
                cell.border = border_thin
                
            else:
                # Data rows formatting
                cell = WriteOnlyCell(ws_out, value=val)
                cell.font = font_data
                cell.border = border_thin
                
                # Determine formats based on column indices (new index)
                # Left aligned text columns:
                # RSM (0), ASM (1), Store Name (3), MCH2 Name (5), MCH3 Name (7), MCH4 (8), Article Name (10), Note/check columns (27, 28, 29)
                if new_col_idx in (0, 1, 3, 5, 7, 8, 10, 27, 28, 29):
                    cell.alignment = align_left
                    
                # Center aligned: Phân loại kiểm tra (30)
                elif new_col_idx == 30:
                    cell.alignment = align_center
                    
                # Right aligned / Formatted columns:
                else:
                    cell.alignment = align_right
                    
                    # 1. Dates: Created on (11), Last GR (21), Last Sale (22)
                    if new_col_idx in (11, 21, 22):
                        cell.value = parse_date_value(val)
                        # Set date number format only if it's a date or float
                        if isinstance(cell.value, (datetime, float)) or hasattr(cell.value, 'strftime'):
                            cell.number_format = 'dd/mm/yyyy'
                            cell.alignment = align_center
                            
                    # 2. Currency: Giá trị tồn (16), Doanh thu (17), Giá vốn (18), Giá tồn D-15 (20)
                    elif new_col_idx in (16, 17, 18, 20):
                        try:
                            cell.value = float(val) if val is not None else None
                            cell.number_format = '#,##0'
                        except ValueError:
                            pass
                            
                    # 3. Float Decimals: DIO (25), DIO/Date (26)
                    elif new_col_idx in (25, 26):
                        try:
                            cell.value = float(val) if val is not None else None
                            cell.number_format = '0.0'
                        except ValueError:
                            pass
                            
                    # 4. Standard Integers (Quantities etc.):
                    elif new_col_idx in (2, 4, 6, 9, 12, 13, 14, 15, 19, 23, 24):
                        try:
                            cell.value = float(val) if val is not None else None
                            # if it's an integer, format without decimals, else with 1 decimal
                            if cell.value is not None:
                                if cell.value == int(cell.value):
                                    cell.value = int(cell.value)
                                    cell.number_format = '#,##0'
                                else:
                                    cell.number_format = '0.0'
                        except ValueError:
                            pass
            
            # Compute width for column (measure first 2000 rows to avoid slow performance)
            if row_idx_1based <= 2000:
                val_str = str(val) if val is not None else ""
                # Give numeric columns a bit more space
                str_len = len(val_str)
                if new_col_idx not in col_widths:
                    col_widths[new_col_idx] = str_len
                else:
                    col_widths[new_col_idx] = max(col_widths[new_col_idx], str_len)
                    
            cells_out.append(cell)
            
        ws_out.append(cells_out)
        
    wb_in.close()
    
    # Apply column widths
    print("  Applying column widths...")
    for col_idx, max_len in col_widths.items():
        col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
        # Add safety margin
        ws_out.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    print("  Saving file...")
    wb_out.save(temppath)
    
    # Replace original file with formatted file
    if os.path.exists(filepath):
        os.remove(filepath)
    os.rename(temppath, filepath)
    
    print(f"  Completed in {time.time()-t_start:.1f}s. Total rows: {row_count:,}")

def main():
    t_all = time.time()
    for f in files_to_format:
        format_excel_file(f)
    print(f"\nAll files formatted successfully in {time.time()-t_all:.1f}s!")

if __name__ == '__main__':
    main()
