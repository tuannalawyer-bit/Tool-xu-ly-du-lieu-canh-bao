import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

folders = [
    r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu",
    r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu\raw mẫu"
]

for folder in folders:
    files = [f for f in os.listdir(folder) if f.endswith(".xlsx")]
    for f in files:
        path = os.path.join(folder, f)
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            for sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                for i, row in enumerate(sheet.iter_rows(values_only=True)):
                    if i >= 5:
                        break
                    row_str = str(row)
                    if "MCH5" in row_str or "mch5" in row_str:
                        print(f"Match found in {f} -> sheet '{sheetname}' -> Row {i}: {row[:15]}")
        except Exception as e:
            pass
