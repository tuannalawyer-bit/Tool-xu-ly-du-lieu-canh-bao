import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

folder = r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu\raw mẫu"
files = [f for f in os.listdir(folder) if f.endswith((".xlsx", ".xlsb"))]

for f in files:
    path = os.path.join(folder, f)
    print(f"File: {f}")
    if f.endswith(".xlsb"):
        # We might need pyxlsb to read xlsb, let's see if we can use openpyxl or check if openpyxl supports it.
        # openpyxl does NOT support xlsb, we'll try to check it or use pyxlsb
        print("  XLSB file - openpyxl does not support. Skip for a moment.")
        continue
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        print(f"  Sheets: {wb.sheetnames}")
        sheet = wb.active
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 1:
                print(f"  Row 1: {row[:15]}")
                break
    except Exception as e:
        print(f"  Error: {e}")
