# -*- coding: utf-8 -*-
"""
TOOL ĐỔ DỮ LIỆU PHÂN TÍCH TỒN KHO & TỔNG QUAN CHUỖI
Phiên bản: 1.5.3
Mô tả: Tích hợp hai chức năng:
       1. Phân tích chi tiết từng cửa hàng (từ file SAP gốc sang file *(done).xlsx và html chi tiết)
       2. Tổng hợp báo cáo Tổng quan chuỗi (từ nhiều file *(done).xlsx thành báo cáo HTML gộp)
Tối ưu hóa:
       - Cache pickle danh mục sản phẩm: lần 2+ nạp < 1 giây thay vì 2-4 phút
       - csv.reader thay DictReader: nhanh hơn ~30% khi đọc file CSV lớn
       - Tab 1 (.xlsx): bỏ bước giải nén/parse XML thô dư thừa khi dò dòng header
       - Tab 2: đọc song song nhiều file *(done).xlsx bằng ThreadPoolExecutor
"""

import os
import sys
import time
import ctypes
import json
import pickle
import hashlib
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, date
from threading import Thread
from concurrent.futures import ThreadPoolExecutor, as_completed
import tempfile
import csv
# Force UTF-8 output
if sys.stdout is not None:
    sys.stdout.reconfigure(encoding='utf-8')

# ===================================================================
# CẤU HÌNH MẶC ĐỊNH & THƯ VIỆN OFFLINE
# ===================================================================
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Thêm các thư mục chứa thư viện offline vào sys.path trước khi import openpyxl
for sub_dir in ["libs", os.path.join("Tool_PhanTichTonKho_Portable", "libs")]:
    p = os.path.join(BASE_DIR, sub_dir)
    if os.path.exists(p) and p not in sys.path:
        sys.path.insert(0, p)

import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def find_default_file(filename):
    p = os.path.abspath(os.path.join(BASE_DIR, filename))
    if os.path.exists(p):
        return p
    p = os.path.abspath(os.path.join(BASE_DIR, "..", filename))
    if os.path.exists(p):
        return p
    p = os.path.abspath(os.path.join(BASE_DIR, "..", "Last GR.SALE", filename))
    if os.path.exists(p):
        return p
    p = os.path.abspath(os.path.join(BASE_DIR, "..", "..", "Last GR.SALE", filename))
    if os.path.exists(p):
        return p
    return os.path.abspath(os.path.join(BASE_DIR, "..", "Last GR.SALE", filename))

DEFAULT_MASTER_PATH = find_default_file("Master article 7.7.xlsx")
DEFAULT_STORE_PATH = find_default_file("Store List.xlsx")
DEFAULT_TEMPLATE_NAME = "dashboard_tong_quan_template.html"

APP_VERSION = "1.5.6"
APP_TITLE = "Tool Phân Tích Tồn Kho & Tổng Quan Chuỗi"

LAST_GR_SALE_DIR = os.path.dirname(DEFAULT_MASTER_PATH)
if os.path.exists(LAST_GR_SALE_DIR):
    SCRATCH_DIR = os.path.join(LAST_GR_SALE_DIR, "temp")
    os.makedirs(SCRATCH_DIR, exist_ok=True)
else:
    SCRATCH_DIR = tempfile.mkdtemp(prefix="tool_do_du_lieu_")

# ===================================================================
# HÀM TRỢ GIÚP DÙNG CHUNG
# ===================================================================
def col_to_idx(col_letter: str) -> int:
    val = 0
    for char in col_letter.upper():
        val = val * 26 + (ord(char) - 64)
    return val - 1

def get_column_letter(col_idx: int) -> str:
    result = []
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result.append(chr(65 + remainder))
    return ''.join(reversed(result))

def safe_float(val) -> float | None:
    if val is None or val == '':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def excel_date_to_datetime(val) -> datetime | None:
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        try:
            return datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(val))
        except (ValueError, OverflowError):
            return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    
    val_str = str(val).strip()
    try:
        return datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(float(val_str)))
    except (ValueError, OverflowError):
        pass
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d.%m.%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(val_str, fmt)
        except ValueError:
            continue
    return None

# ===================================================================
# PHẦN 1: LOGIC XỬ LÝ CHI TIẾT
# ===================================================================
def load_store_details(store_path: str, log_func=None) -> dict:
    store_map = {}
    if not os.path.exists(store_path):
        if log_func:
            log_func(f"  [CẢNH BÁO] Không tìm thấy file Store List tại: {store_path}")
        return store_map

    if log_func:
        log_func(f"  Đang đọc danh sách siêu thị từ: {os.path.basename(store_path)}")

    try:
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, f"temp_{os.path.basename(store_path)}")
        res = ctypes.windll.kernel32.CopyFileW(store_path, temp_path, False)
        if not res:
            if log_func:
                log_func("  ✗ Không thể copy file Store List (đang mở?)")
            return store_map

        wb = openpyxl.load_workbook(temp_path, read_only=True)
        
        def parse_wm_sheet(ws_obj):
            headers = next(ws_obj.iter_rows(values_only=True))
            idx_code = 0
            idx_name = 1
            idx_gdv = 2
            if headers:
                for idx, val in enumerate(headers):
                    if val:
                        val_str = str(val).strip().lower()
                        if any(x in val_str for x in ('code', 'store', 'mã', 'ma ch')) and not any(x in val_str for x in ('name', 'tên')):
                            idx_code = idx
                        elif any(x in val_str for x in ('name', 'tên', 'ten ch')):
                            idx_name = idx
                        elif any(x in val_str for x in ('gđv', 'gđm', 'gdv', 'gdm', 'giám sát', 'giam sat', 'rsm')):
                            idx_gdv = idx
            return idx_code, idx_name, idx_gdv

        sheet_count = 0
        for sheetname in wb.sheetnames:
            if sheetname == '_com.sap.ip.bi.xl.hiddensheet':
                continue
            ws = wb[sheetname]
            sheet_count += 1
            
            sheet_type = 'WM+'
            sheet_lower = sheetname.lower()
            if 'wm' in sheet_lower and '+' not in sheet_lower:
                sheet_type = 'WM'
                
            if sheet_type == 'WM':
                idx_code, idx_name, idx_gdv = parse_wm_sheet(ws)
                is_first = True
                for row in ws.iter_rows(values_only=True):
                    if is_first:
                        is_first = False
                        continue
                    if len(row) > max(idx_code, idx_name, idx_gdv):
                        store_id = str(row[idx_code]).strip() if row[idx_code] is not None else ''
                        store_name = str(row[idx_name]).strip() if row[idx_name] is not None else ''
                        gdv = str(row[idx_gdv]).strip() if row[idx_gdv] is not None else ''
                        if store_id:
                            if store_id.endswith('.0'):
                                store_id = store_id[:-2]
                            # Lưu vào map: (GĐV/GĐM, ASM=StoreName, StoreName, sheet_type)
                            store_map[store_id] = (gdv, store_name, store_name, 'WM')
            else:
                is_first = True
                for row in ws.iter_rows(values_only=True):
                    if is_first:
                        is_first = False
                        continue
                    if len(row) >= 4:
                        rsm = str(row[0]).strip() if row[0] is not None else ''
                        asm = str(row[1]).strip() if row[1] is not None else ''
                        store_id = str(row[2]).strip() if row[2] is not None else ''
                        store_name = str(row[3]).strip() if row[3] is not None else ''
                        if store_id:
                            if store_id.endswith('.0'):
                                store_id = store_id[:-2]
                            # Lưu vào map: (RSM, ASM, StoreName, sheet_type)
                            store_map[store_id] = (rsm, asm, store_name, 'WM+')
                
        wb.close()
        try:
            os.remove(temp_path)
        except Exception:
            pass
        if log_func:
            log_func(f"  ✓ Đã nạp xong {len(store_map):,} siêu thị/cửa hàng từ {sheet_count} sheet.")
    except Exception as e:
        if log_func:
            log_func(f"  ✗ Lỗi khi đọc file Store List: {e}")
    return store_map

# Phiên bản schema của product_map — TĂNG mỗi khi đổi cấu trúc tuple lưu trong
# product_map (thêm/bớt trường), để cache pickle cũ tự động bị bỏ qua thay vì
# nạp nhầm dữ liệu thiếu trường mới (vd: thiếu mch4_desc gây cột MCH4 Desc rỗng).
CACHE_SCHEMA_VERSION = 2

def _get_master_cache_path(master_path: str) -> str:
    """Trả về đường dẫn file cache pickle cạnh file master (kèm phiên bản schema)."""
    master_dir = os.path.dirname(master_path)
    master_base = os.path.splitext(os.path.basename(master_path))[0]
    return os.path.join(master_dir, f".{master_base}_cache_v{CACHE_SCHEMA_VERSION}.pkl")

def _is_cache_valid(master_path: str, cache_path: str) -> bool:
    """Kiểm tra cache còn hợp lệ (file master chưa thay đổi)."""
    if not os.path.exists(cache_path):
        return False
    try:
        master_mtime = os.path.getmtime(master_path)
        cache_mtime = os.path.getmtime(cache_path)
        return cache_mtime >= master_mtime
    except Exception:
        return False

def load_product_details(master_path: str, log_func=None, force_reload: bool = False) -> dict:
    """Nạp danh mục sản phẩm từ Master Article xlsx.
    Lần đầu: đọc Excel và lưu cache pickle.
    Lần sau: nạp pickle trong < 1 giây nếu file master không thay đổi."""
    product_map = {}
    if not os.path.exists(master_path):
        if log_func:
            log_func(f"  [CẢNH BÁO] Không tìm thấy file Master Article tại: {master_path}")
        return product_map

    cache_path = _get_master_cache_path(master_path)

    # --- Thử load từ cache pickle ---
    if not force_reload and _is_cache_valid(master_path, cache_path):
        try:
            t0 = time.time()
            with open(cache_path, 'rb') as f:
                product_map = pickle.load(f)
            t1 = time.time()
            if log_func:
                log_func(f"  ✓ Nạp cache danh mục: {len(product_map):,} sản phẩm | Thời gian: {t1 - t0:.2f}s (cache pickle)")
            return product_map
        except Exception:
            pass  # Cache lỗi → đọc lại Excel

    if log_func:
        log_func(f"  Đang đọc danh mục sản phẩm từ: {os.path.basename(master_path)} (lần đầu, sẽ cache lại...)")

    t0 = time.time()
    try:
        import openpyxl
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, f"temp_{os.path.basename(master_path)}")
        res = ctypes.windll.kernel32.CopyFileW(master_path, temp_path, False)
        if not res:
            if log_func:
                log_func("  ✗ Không thể copy file Master (đang mở?)")
            return product_map

        count = 0
        wb = openpyxl.load_workbook(temp_path, read_only=True)
        ws = wb['Export'] if 'Export' in wb.sheetnames else wb.active
        
        is_first = True
        for row in ws.iter_rows(values_only=True):
            if is_first:
                is_first = False
                continue
            if not row or len(row) == 0:
                continue
                
            prod_id = str(row[0]).strip() if row[0] is not None else ''
            if not prod_id:
                continue
            if prod_id.endswith('.0'):
                prod_id = prod_id[:-2]
                
            prod_name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            
            date_val = row[6] if len(row) > 6 and row[6] is not None else None
            try:
                date_num = float(date_val) if date_val is not None else 0.0
            except (ValueError, TypeError):
                date_num = 0.0
                
            mch2_id = str(row[15]).strip() if len(row) > 15 and row[15] is not None else ''
            mch2_desc = str(row[16]).strip() if len(row) > 16 and row[16] is not None else ''
            mch3_id = str(row[17]).strip() if len(row) > 17 and row[17] is not None else ''
            mch3_desc = str(row[18]).strip() if len(row) > 18 and row[18] is not None else ''
            mch4_id = str(row[19]).strip() if len(row) > 19 and row[19] is not None else ''
            mch4_desc = str(row[20]).strip() if len(row) > 20 and row[20] is not None else ''
            mch5_desc = str(row[22]).strip() if len(row) > 22 and row[22] is not None else 'Khác'
            
            shelf_life = None
            if len(row) > 25 and row[25] is not None:
                try:
                    shelf_life = int(float(row[25]))
                except (ValueError, TypeError):
                    pass
            
            # Lưu theo định dạng chuẩn (mch4_desc nối cuối để không xê dịch index cũ):
            # (prod_name, mch2_id, mch2_desc, mch3_id, mch3_desc, mch4_id, shelf_life, mch5_desc, date_num, mch4_desc)
            product_map[prod_id] = (prod_name, mch2_id, mch2_desc, mch3_id, mch3_desc, mch4_id, shelf_life, mch5_desc, date_num, mch4_desc)
            count += 1
            
        wb.close()
        try:
            os.remove(temp_path)
        except Exception:
            pass
        t1 = time.time()
        if log_func:
            log_func(f"  ✓ Đọc xong danh mục sản phẩm: {count:,} sản phẩm | Thời gian: {t1 - t0:.1f}s")

        # --- Lưu cache pickle ---
        try:
            with open(cache_path, 'wb') as f:
                pickle.dump(product_map, f, protocol=pickle.HIGHEST_PROTOCOL)
            if log_func:
                log_func(f"  💾 Đã lưu cache → lần sau nạp < 1 giây")
        except Exception as e:
            if log_func:
                log_func(f"  [CẢNH BÁO] Không thể lưu cache: {e}")

    except Exception as e:
        if log_func:
            log_func(f"  ✗ Lỗi khi đọc file Master: {e}")
    return product_map

def generate_html_dashboard(out_path: str, rows_data: list, product_map: dict, log_func=None, target_date=None, chain_title: str = "") -> bool:
    try:
        t0 = time.time()
        current_dir = os.path.dirname(out_path)
        # Ưu tiên template mới nhất trong thư mục ứng dụng (BASE_DIR); chỉ dùng
        # bản nằm cạnh file dữ liệu khi BASE_DIR không có, tránh dùng template cũ.
        template_path = os.path.join(BASE_DIR, "dashboard_chi_tiet_template.html")
        if not os.path.exists(template_path):
            template_path = os.path.join(current_dir, "dashboard_chi_tiet_template.html")

        if not os.path.exists(template_path):
            if log_func:
                log_func("  ✗ Không tìm thấy file template 'dashboard_chi_tiet_template.html'. Bỏ qua bước sinh Dashboard!")
            return False

        sample_len = len(rows_data[0]) if rows_data else 27
        is_excel = sample_len in (30, 31)
        has_asm = sample_len in (27, 31)
        
        idx_rsm = 0
        if has_asm:
            idx_asm = 1
            idx_store_code = 2
            idx_store_name = 3
            idx_mch2_desc = 5
            idx_mch3_desc = 7
            idx_mch4_id = 8
            idx_art_code = 9
            idx_art_name = 10
            
            if is_excel:
                idx_qty = 15
                idx_val = 16
                idx_class = 30
            else:
                idx_qty = 14
                idx_val = 15
                idx_class = 26
        else:
            idx_asm = -1
            idx_store_code = 1
            idx_store_name = 2
            idx_mch2_desc = 4
            idx_mch3_desc = 6
            idx_mch4_id = 7
            idx_art_code = 8
            idx_art_name = 9
            
            if is_excel:
                idx_qty = 14
                idx_val = 15
                idx_class = 29
            else:
                idx_qty = 13
                idx_val = 14
                idx_class = 25

        rsms = set()
        asms = set()
        stores = {}
        mch2s = set()
        mch3s = set()
        mch4s = set()
        mch5s = set()
        classes = set()
        articles = {}
        
        for r in rows_data:
            rsm = r[idx_rsm] or ''
            asm = r[idx_asm] or '' if has_asm else ''
            store_code = r[idx_store_code] or ''
            store_name = r[idx_store_name] or ''
            mch2_desc = r[idx_mch2_desc] or ''
            mch3_desc = r[idx_mch3_desc] or ''
            mch4_id = r[idx_mch4_id] or ''
            art_code = r[idx_art_code] or ''
            art_name = r[idx_art_name] or ''
            check_class = r[idx_class] or ''
            
            rsms.add(rsm)
            if has_asm and asm:
                asms.add(asm)
            stores[store_code] = store_name
            mch2s.add(mch2_desc)
            mch3s.add(mch3_desc)
            mch4s.add(mch4_id)
            classes.add(check_class)
            
            prod_info = product_map.get(art_code)
            mch5 = 'Khác'
            if prod_info and len(prod_info) > 7:
                mch5 = prod_info[7] or 'Khác'
            mch5s.add(mch5)
            
            art_key = (art_code, art_name)
            if art_key not in articles:
                articles[art_key] = len(articles)
                
        rsm_list = sorted(list(rsms))
        asm_list = sorted(list(asms)) if has_asm else []
        store_list = [[code, name] for code, name in sorted(stores.items())]
        mch2_list = sorted(list(mch2s))
        mch3_list = sorted(list(mch3s))
        mch4_list = sorted(list(mch4s))
        mch5_list = sorted(list(mch5s))
        class_list = sorted(list(classes))
        
        article_list = [None] * len(articles)
        for (code, name), idx in articles.items():
            article_list[idx] = [code, name]
            
        rsm_map = {name: idx for idx, name in enumerate(rsm_list)}
        asm_map = {name: idx for idx, name in enumerate(asm_list)}
        store_map = {item[0]: idx for idx, item in enumerate(store_list)}
        mch2_map = {name: idx for idx, name in enumerate(mch2_list)}
        mch3_map = {name: idx for idx, name in enumerate(mch3_list)}
        mch4_map = {name: idx for idx, name in enumerate(mch4_list)}
        mch5_map = {name: idx for idx, name in enumerate(mch5_list)}
        class_map = {name: idx for idx, name in enumerate(class_list)}
        
        compressed_rows = []
        for r in rows_data:
            rsm = r[idx_rsm] or ''
            asm = r[idx_asm] or '' if has_asm else ''
            store_code = r[idx_store_code] or ''
            mch2_desc = r[idx_mch2_desc] or ''
            mch3_desc = r[idx_mch3_desc] or ''
            mch4_id = r[idx_mch4_id] or ''
            art_code = r[idx_art_code] or ''
            art_name = r[idx_art_name] or ''
            qty_val = safe_float(r[idx_qty]) or 0.0
            val_val = safe_float(r[idx_val]) or 0.0
            check_class = r[idx_class] or ''
            
            prod_info = product_map.get(art_code)
            mch5 = 'Khác'
            if prod_info and len(prod_info) > 7:
                mch5 = prod_info[7] or 'Khác'
                
            art_key = (art_code, art_name)
            
            comp_row = [
                store_map[store_code],
                rsm_map[rsm],
                asm_map.get(asm, -1),
                mch2_map[mch2_desc],
                mch3_map[mch3_desc],
                mch4_map[mch4_id],
                mch5_map[mch5],
                articles[art_key],
                class_map[check_class],
                round(qty_val),
                int(val_val)
            ]
            compressed_rows.append(comp_row)
            
        has_wm = any(r[1] == r[3] for r in rows_data if len(r) >= 4)
        out_dict = {
            "rsm_list": rsm_list,
            "asm_list": asm_list,
            "store_list": store_list,
            "mch2_list": mch2_list,
            "mch3_list": mch3_list,
            "mch4_list": mch4_list,
            "mch5_list": mch5_list,
            "class_list": class_list,
            "article_list": article_list,
            "has_wm": has_wm,
            "rows": compressed_rows
        }
        
        with open(template_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
            
        filename = os.path.basename(out_path)
        title = "WIN"
        if "rural" in filename.lower():
            title = "Rural"
        elif "urban" in filename.lower():
            title = "Urban"
            
        display_title = chain_title if chain_title else f"Chuỗi {title}"
        
        html_content = html_content.replace(
            "<title>Báo Cáo Đối Soát Dữ Liệu Kiểm Tra - Chuỗi WIN</title>",
            f"<title>Báo Cáo Đối Soát Dữ Liệu Kiểm Tra - {display_title}</title>"
        )
        html_content = html_content.replace(
            '<span class="px-2.5 py-0.5 bg-indigo-500/10 text-indigo-400 border border-indigo-500/20 text-xs font-bold rounded-full">Chuỗi WIN</span>',
            f'<span class="px-2.5 py-0.5 bg-indigo-500/10 text-indigo-400 border border-indigo-500/20 text-xs font-bold rounded-full">{display_title}</span>'
        )
        
        if target_date:
            if isinstance(target_date, str):
                today_str = target_date
            else:
                today_str = target_date.strftime("%d/%m/%Y")
        else:
            today_str = datetime.now().strftime("%d/%m/%Y")
        html_content = html_content.replace(
            "Cập nhật: 01/07/2026",
            f"Cập nhật: {today_str}"
        )
        
        js_injection = f"const dbData = {json.dumps(out_dict, ensure_ascii=False, separators=(',', ':'))};\n"
        js_injection += "const rawData = dbData.rows.map(r => ({\n"
        js_injection += "    store_idx: r[0],\n"
        js_injection += "    store_code: dbData.store_list[r[0]][0],\n"
        js_injection += "    store_name: dbData.store_list[r[0]][1],\n"
        js_injection += "    rsm: dbData.rsm_list[r[1]],\n"
        js_injection += "    rsm_idx: r[1],\n"
        js_injection += "    asm: r[2] !== -1 ? dbData.asm_list[r[2]] : '',\n"
        js_injection += "    asm_idx: r[2],\n"
        js_injection += "    mch2: dbData.mch2_list[r[3]],\n"
        js_injection += "    mch2_idx: r[3],\n"
        js_injection += "    mch3: dbData.mch3_list[r[4]],\n"
        js_injection += "    mch3_idx: r[4],\n"
        js_injection += "    mch4: dbData.mch4_list[r[5]],\n"
        js_injection += "    mch4_idx: r[5],\n"
        js_injection += "    mch5: dbData.mch5_list[r[6]],\n"
        js_injection += "    mch5_idx: r[6],\n"
        js_injection += "    article_code: dbData.article_list[r[7]][0],\n"
        js_injection += "    article_name: dbData.article_list[r[7]][1],\n"
        js_injection += "    check_class: dbData.class_list[r[8]],\n"
        js_injection += "    class_idx: r[8],\n"
        js_injection += "    qty: r[9],\n"
        js_injection += "    value: r[10]\n"
        js_injection += "}));"
        
        html_content = html_content.replace('// DATA_PLACEHOLDER', js_injection)
        
        out_html_path = out_path.replace(".xlsx", "_dashboard.html")
        with open(out_html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
            
        elapsed = time.time() - t0
        if log_func:
            log_func(f"  ✓ Đã sinh Dashboard: {os.path.basename(out_html_path)} | Thời gian: {elapsed:.1f}s")
        return True
    except Exception as e:
        if log_func:
            log_func(f"  ✗ Lỗi sinh Dashboard: {e}")
        return False

# ===================================================================
# LOGIC TÍNH TOÁN NGHIỆP VỤ & XỬ LÝ CSV
# ===================================================================
def compute_row(row_dict: dict, milestone_date: datetime, ref_amount: float,
                shelf_life_map: dict) -> dict:
    result = {}
    
    art_code_raw = row_dict.get('K')
    if art_code_raw is None:
        art_code_raw = ''
    art_code_str = str(art_code_raw).strip()
    if art_code_str.endswith('.0'):
        art_code_str = art_code_str[:-2]

    ton_dau_ky = safe_float(row_dict.get('N')) or 0
    nhap_trong_ky = safe_float(row_dict.get('O')) or 0
    xuat_trong_ky = safe_float(row_dict.get('P')) or 0
    ton_cuoi_ky = safe_float(row_dict.get('Q')) or 0
    gia_tri_ton = safe_float(row_dict.get('R')) or 0
    doanh_thu = safe_float(row_dict.get('S')) or 0
    gia_von = safe_float(row_dict.get('T')) or 0
    ton_d15 = safe_float(row_dict.get('U')) or 0
    gia_ton_d15 = safe_float(row_dict.get('V')) or 0

    last_gr_raw = row_dict.get('W')
    last_sale_raw = row_dict.get('X')

    last_gr_is_empty = (last_gr_raw is None or str(last_gr_raw).strip() == '' or str(last_gr_raw).strip().lower() == 'none')
    last_sale_is_empty = (last_sale_raw is None or str(last_sale_raw).strip() == '' or str(last_sale_raw).strip().lower() == 'none')

    last_gr_gt90 = (str(last_gr_raw).strip() == '>90' or last_gr_is_empty)
    last_sale_gt90 = (str(last_sale_raw).strip() == '>90' or last_sale_is_empty)

    last_gr = None if last_gr_gt90 else excel_date_to_datetime(last_gr_raw)
    last_sale = None if last_sale_gt90 else excel_date_to_datetime(last_sale_raw)

    result['W'] = '>90' if last_gr_gt90 else last_gr_raw
    result['X'] = '>90' if last_sale_gt90 else last_sale_raw

    if last_gr_gt90 or last_gr is None or milestone_date is None:
        result['Y'] = '>90'
        y_days = 91
    else:
        delta = milestone_date - last_gr
        diff_days = delta.days
        if diff_days > 90:
            result['Y'] = '>90'
            y_days = 91
        else:
            result['Y'] = diff_days
            y_days = diff_days

    z_shelf_life = shelf_life_map.get(art_code_str)
    if z_shelf_life is None:
        orig_z = row_dict.get('Z')
        if orig_z == '#N/A' or orig_z is None:
            z_shelf_life = '#N/A'
        else:
            z_shelf_life = safe_float(orig_z)
    result['Z'] = z_shelf_life

    if xuat_trong_ky == 0:
        aa_val = 9999
    else:
        aa_val = (ton_d15 / xuat_trong_ky) * 90
    result['AA'] = aa_val

    if z_shelf_life == '#N/A':
        result['AB'] = '#N/A'
        result['AD'] = '#N/A'
        result['AE'] = '#N/A'
        result['AF'] = '#N/A'
        ac_note = ''
        if last_sale_gt90:
            if last_gr_gt90:
                ac_note = 'không giao dịch >90 ngày'
            else:
                ac_note = 'Không sale 90 ngày'
        result['AC'] = ac_note if ac_note else None
        return result

    if aa_val is not None and z_shelf_life is not None and z_shelf_life > 0:
        result['AB'] = aa_val / z_shelf_life
    else:
        result['AB'] = None
    ab_val = result['AB']

    la_nghi_van = False
    la_non_moving = False
    ac_note = ''

    if last_sale_gt90:
        if last_gr_gt90:
            ac_note = 'không giao dịch >90 ngày'
            la_nghi_van = True
        else:
            ac_note = 'Không sale 90 ngày'
            la_non_moving = True
    result['AC'] = ac_note if ac_note else None

    if ab_val is not None and ab_val > 5 and gia_tri_ton > ref_amount:
        result['AD'] = 'Check'
    else:
        result['AD'] = None

    ae_note = ''
    if z_shelf_life is not None:
        if last_gr_gt90 or y_days == 91:
            if z_shelf_life < 90:
                ae_note = 'Hết hạn'
        elif y_days is not None:
            if y_days > z_shelf_life:
                ae_note = 'Hết hạn'
    result['AE'] = ae_note if ae_note else None

    phan_loai = None
    if ae_note:
        phan_loai = 'Hết hạn'
    elif la_nghi_van:
        phan_loai = 'Nghi vấn tồn ảo'
    elif la_non_moving:
        phan_loai = 'Non-moving'
    elif result['AD'] == 'Check':
        phan_loai = 'Slow moving'

    result['AF'] = phan_loai
    return result

def process_csv_file(src_path: str, store_map: dict, product_map: dict, milestone_date: datetime,
                     ref_amount: float, log_func=None, chain_title: str = "", shelf_life_map: dict = None) -> str | None:
    filename = os.path.basename(src_path)
    base, ext = os.path.splitext(filename)
    out_filename = f"{base} (done).xlsx"
    out_path = os.path.join(os.path.dirname(src_path), out_filename)

    if os.path.exists(out_path):
        try:
            with open(out_path, 'a'):
                pass
        except PermissionError:
            if log_func:
                log_func(f"  ✗ [LỖI KHÓA FILE] File kết quả '{out_filename}' đang mở. Vui lòng đóng và chạy lại!", 'error')
            return None

    if log_func:
        log_func(f"  Bắt đầu xử lý (CSV)...")

    t0 = time.time()
    temp_path = os.path.join(SCRATCH_DIR, f"temp_{filename}")

    try:
        res = ctypes.windll.kernel32.CopyFileW(src_path, temp_path, False)
        if not res:
            if log_func:
                log_func(f"  ✗ Không thể copy file: {filename}")
            return None

        font_white_bold = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        font_normal = Font(name='Arial', size=9)
        fill_header = PatternFill(fill_type='solid', start_color='1F4E78', end_color='1F4E78')
        
        fill_het_han = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
        fill_nghi_van = PatternFill(fill_type='solid', start_color='FF6600', end_color='FF6600')
        fill_non_moving = PatternFill(fill_type='solid', start_color='FFA500', end_color='FFA500')
        fill_slow_moving = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
        font_slow = Font(name='Arial', size=9, color='000000')
        
        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        phan_loai_style = {
            'Hết hạn': (fill_het_han, Font(name='Arial', size=9, bold=True, color='FFFFFF')),
            'Nghi vấn tồn ảo': (fill_nghi_van, Font(name='Arial', size=9, bold=True, color='FFFFFF')),
            'Non-moving': (fill_non_moving, Font(name='Arial', size=9, bold=True, color='000000')),
            'Slow moving': (fill_slow_moving, font_slow),
        }

        wb_out = openpyxl.Workbook(write_only=True)
        ws_out = wb_out.create_sheet(title='Dữ liệu kiểm tra')

        if shelf_life_map is None:
            shelf_life_map = {k: v[6] for k, v in product_map.items()}

        rows_written = 0
        rows_filtered = 0
        written_rows = []

        with open(temp_path, 'r', encoding='utf-8', errors='ignore') as f:
            first_line = f.readline()
            delimiter = ';' if ';' in first_line else ','

        # --- Đọc nhanh bằng csv.reader + tra cứu chỉ mục từ header ---
        with open(temp_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f, delimiter=delimiter)
            header = next(reader, None)
            if header is None:
                return None
            col_idx = {h.strip(): i for i, h in enumerate(header)}

            def _get(row, col_name, default=''):
                i = col_idx.get(col_name)
                if i is None or i >= len(row):
                    return default
                v = row[i]
                return v if v is not None else default

            for row in reader:
                store_id = _get(row, 'STORE_ID').strip()
                if store_id.endswith('.0'):
                    store_id = store_id[:-2]
                    
                prod_id = _get(row, 'PRODUCT_ID').strip()
                if prod_id.endswith('.0'):
                    prod_id = prod_id[:-2]

                store_name_raw = _get(row, 'STORE_NAME').strip()
                if 'result' in store_id.lower() or 'result' in store_name_raw.lower():
                    rows_filtered += 1
                    continue

                store_info = store_map.get(store_id)
                if store_info:
                    rsm_name, asm_name, store_name, sheet_type = store_info[:4]
                else:
                    store_name = store_name_raw
                    if store_id.isdigit():
                        rsm_name = _get(row, 'GDV') or _get(row, 'RSM')
                        asm_name = store_name_raw
                        sheet_type = 'WM'
                    else:
                        rsm_name = _get(row, 'GDV') or _get(row, 'RSM')
                        asm_name = _get(row, 'ASM')
                        sheet_type = 'WM+'
                
                prod_info = product_map.get(prod_id, (_get(row, 'PRODUCT_NAME'), '', '', '', '', '', None, 'Khác'))
                prod_name, mch2_id, mch2_desc, mch3_id, mch3_desc, mch4_id, *rest = prod_info
                # rest = [shelf_life, mch5_desc, date_num, mch4_desc]
                mch5_desc = rest[1] if len(rest) > 1 and rest[1] else 'Khác'
                mch4_desc = rest[3] if len(rest) > 3 and rest[3] else ''

                if mch2_id == '101' or not mch2_id or not mch3_id or str(mch2_id).strip() == '' or str(mch3_id).strip() == '':
                    rows_filtered += 1
                    continue

                norm_row = {
                    'K': prod_id,
                    'N': _get(row, 'CLOSING_STOCK_QUANTITY_D90'),
                    'O': _get(row, 'GR_QTY_LAST90_D'),
                    'P': _get(row, 'GI_QTY_LAST90_D'),
                    'Q': _get(row, 'CLOSING_STOCK_QUANTITY_LASTDAY'),
                    'R': _get(row, 'CLOSING_STOCK_VALUE_LASTDAY'),
                    'S': 0,
                    'T': 0,
                    'U': _get(row, 'CLOSING_STOCK_QUANTITY_D15'),
                    'V': 0,
                    'W': _get(row, 'MAX_GR_DATE'),
                    'X': _get(row, 'MAX_SALE_DATE'),
                    'Z': _get(row, 'Z')
                }

                computed = compute_row(norm_row, milestone_date, ref_amount, shelf_life_map)
                phan_loai = computed.get('AF')

                if phan_loai is None or str(phan_loai).strip() == '':
                    rows_filtered += 1
                    continue

                row_list = [None] * 27
                row_list[0] = rsm_name
                row_list[1] = asm_name
                row_list[2] = store_id
                row_list[3] = store_name
                row_list[4] = mch2_id
                row_list[5] = mch2_desc
                row_list[6] = mch3_id
                row_list[7] = mch3_desc
                row_list[8] = mch4_desc
                row_list[9] = prod_id
                row_list[10] = prod_name
                row_list[11] = norm_row['N']
                row_list[12] = norm_row['O']
                row_list[13] = norm_row['P']
                row_list[14] = norm_row['Q']
                row_list[15] = norm_row['R']
                row_list[16] = norm_row['U']
                row_list[17] = computed['W']
                row_list[18] = computed['X']
                row_list[19] = computed['Y']
                row_list[20] = computed['Z']
                row_list[21] = computed['AA']
                row_list[22] = computed['AB']
                row_list[23] = computed['AC']
                row_list[24] = computed['AD']
                row_list[25] = computed['AE']
                row_list[26] = computed['AF']

                written_rows.append((row_list, phan_loai, sheet_type, mch5_desc))

        has_wm = any(item[2] == 'WM' for item in written_rows)

        headers = [
            'RSM / GĐV/GĐM' if has_wm else 'RSM', 'ASM', 'Store Code', 'Store Name',
            'MCH2 ID', 'MCH2 Desc', 'MCH3 ID', 'MCH3 Desc', 'MCH4 Desc',
            'Article Code', 'Article Name',
            'Tồn đầu kỳ (D-90)', 'Nhập trong kỳ', 'Xuất trong kỳ',
            'Tồn cuối kỳ (D)', 'Giá trị tồn', 'Tồn D-15',
            'Last GR', 'Last Sale', 'Số ngày ko nhập', 'Date tham khảo',
            'DIO (D-15)', 'DIO/Date', 'Note', 'Note check slow', 'note hết hạn',
            'Phân loại kiểm tra', 'MCH5 Desc'
        ]
            
        cells_out2 = []
        for val in headers:
            cell = WriteOnlyCell(ws_out, value=val)
            cell.font = font_white_bold
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin
            cells_out2.append(cell)
        ws_out.append(cells_out2)
        
        dashboard_rows = []
        for row_list, phan_loai, sheet_type, mch5_desc in written_rows:
            cells_temp = []
            for col_idx in range(27):
                val = row_list[col_idx]
                if val is None or val == '':
                    cells_temp.append(None)
                    continue

                if col_idx in (17, 18):
                    dt = excel_date_to_datetime(val) if not (val == '>90') else None
                    if dt:
                        cell = WriteOnlyCell(ws_out, value=dt)
                        cell.number_format = 'dd/mm/yyyy'
                        cells_temp.append(cell)
                    else:
                        cells_temp.append(val)
                elif col_idx in (11, 12, 13, 14, 15, 16, 19, 20):
                    if col_idx == 19 and str(val).strip() == '>90':
                        cells_temp.append(val)
                    else:
                        try:
                            num_val = int(float(val))
                            cell = WriteOnlyCell(ws_out, value=num_val)
                            cell.number_format = '#,##0'
                            cells_temp.append(cell)
                        except (ValueError, TypeError):
                            cells_temp.append(val)
                elif col_idx in (21, 22):
                    try:
                        num_val = float(val)
                        cell = WriteOnlyCell(ws_out, value=num_val)
                        cell.number_format = '0.00'
                        cells_temp.append(cell)
                    except (ValueError, TypeError):
                        cells_temp.append(val)
                elif col_idx == 26:
                    fill_af, font_af = phan_loai_style.get(phan_loai, (None, None))
                    cell = WriteOnlyCell(ws_out, value=val)
                    if font_af:
                        cell.font = font_af
                    if fill_af:
                        cell.fill = fill_af
                    cells_temp.append(cell)
                else:
                    cells_temp.append(val)

            # Cột cuối: Nhóm hàng (MCH5) — chỉ ghi ra file, không đưa vào dashboard_rows (Tab 1)
            cells_temp.append(mch5_desc if mch5_desc else None)

            ws_out.append(cells_temp)
            rows_written += 1
            dashboard_rows.append(row_list)

        wb_out.save(out_path)
        if rows_written > 0:
            generate_html_dashboard(out_path, dashboard_rows, product_map, log_func, milestone_date, chain_title)

        elapsed = time.time() - t0
        if log_func:
            log_func(f"  ✓ Hoàn thành: {rows_written:,} hàng xuất, {rows_filtered:,} hàng lọc bỏ")
            log_func(f"  → Thời gian: {elapsed:.1f}s | File: {out_filename}")
        return out_path

    except Exception as e:
        if log_func:
            if isinstance(e, PermissionError):
                log_func(f"  ✗ [LỖI KHÓA FILE] Không thể ghi đè. Vui lòng đóng và chạy lại!", 'error')
            else:
                log_func(f"  ✗ Lỗi xử lý {filename}: {e}")
                import traceback
                log_func(traceback.format_exc())
        return None
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass

def process_excel_file(src_path: str, store_map: dict, product_map: dict, milestone_date: datetime,
                       ref_amount: float, log_func=None, chain_title: str = "", shelf_life_map: dict = None) -> str | None:
    filename = os.path.basename(src_path)
    base, ext = os.path.splitext(filename)
    if ext.lower() == '.csv':
        return process_csv_file(src_path, store_map, product_map, milestone_date, ref_amount, log_func, chain_title, shelf_life_map)
        
    out_dir = os.path.dirname(src_path)
    out_filename = filename.replace(".xlsx", " (done).xlsx")
    out_path = os.path.join(out_dir, out_filename)
    
    t0 = time.time()

    try:
        if log_func:
            log_func(f"  Bước 1: Quét & Tính toán các chỉ số tồn kho...")

        wb_in = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
        sheet_in = wb_in.worksheets[0]

        rsm_row_idx = -1
        for row_num, row in enumerate(sheet_in.iter_rows(max_row=500, values_only=True), start=1):
            if row and any(isinstance(v, str) and v.strip() == 'RSM' for v in row):
                rsm_row_idx = row_num
                break

        if rsm_row_idx == -1:
            if log_func:
                log_func("  ✗ Không tìm thấy tiêu đề 'RSM' trong file gốc. Vui lòng kiểm tra lại cấu trúc file!", "error")
            return None

        rows_filtered = 0
        rows_written = 0
        
        wb_out = openpyxl.Workbook(write_only=True)
        ws_out = wb_out.create_sheet("Dữ liệu kiểm tra")
        
        font_white_bold = Font(name='Segoe UI', size=9, bold=True, color='FFFFFF')
        font_normal = Font(name='Segoe UI', size=9)
        fill_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_right = Alignment(horizontal='right', vertical='center')
        
        border_thin = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        
        fill_exp = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        font_exp = Font(name='Segoe UI', size=9, color='9C0006', bold=True)
        
        fill_virtual = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        font_virtual = Font(name='Segoe UI', size=9, color='9C6500', bold=True)
        
        fill_nonmoving = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        font_nonmoving = Font(name='Segoe UI', size=9, color='1F4E78', bold=True)
        
        fill_slow = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        font_slow = Font(name='Segoe UI', size=9, color='375623', bold=True)
        
        phan_loai_style = {
            'Hết hạn': (fill_exp, font_exp),
            'Nghi vấn tồn ảo': (fill_virtual, font_virtual),
            'Non-moving': (fill_nonmoving, font_nonmoving),
            'Slow moving': (fill_slow, font_slow)
        }
        
        headers_row_in = next(sheet_in.iter_rows(max_row=rsm_row_idx, min_row=rsm_row_idx, values_only=True))
        h_map = {str(h).strip(): idx for idx, h in enumerate(headers_row_in) if h is not None}
        
        c_rsm = h_map.get('RSM', 0)
        c_asm = h_map.get('ASM')
        c_store_code = h_map.get('Store Code', 1)
        c_store_name = h_map.get('Store Name', 2)
        c_art_code = h_map.get('Article Code', 8)
        c_art_name = h_map.get('Article Name', 9)
        c_last_gr = h_map.get('Last GR', 21)
        c_last_sale = h_map.get('Last Sale', 22)
        c_dt = h_map.get('Doanh thu', 17)
        c_qty = h_map.get('Tồn cuối kỳ (D)', 13)
        c_val = h_map.get('Giá trị tồn', 14)
        c_val_d15 = h_map.get('Giá tồn D-15', 20)
        c_created_on = h_map.get('Created on')
        c_ton_dau_ky = h_map.get('Tồn đầu kỳ (D-90)')
        c_nhap_trong_ky = h_map.get('Nhập trong kỳ')
        c_xuat_trong_ky = h_map.get('Xuất trong kỳ')
        c_gia_von = h_map.get('Giá vốn')
        c_ton_d15 = h_map.get('Tồn D-15')

        def _row_val(row, idx_c):
            return row[idx_c] if idx_c is not None and idx_c < len(row) else None

        written_rows = []
        row_idx_iter = 0
        
        for row in sheet_in.iter_rows(values_only=True):
            row_idx_iter += 1
            if row_idx_iter <= rsm_row_idx:
                continue
            
            if not row or all(v is None for v in row):
                continue
            
            store_code = str(row[c_store_code]).strip() if row[c_store_code] is not None else ''
            store_name = str(row[c_store_name]).strip() if row[c_store_name] is not None else ''
            art_code = str(row[c_art_code]).strip() if row[c_art_code] is not None else ''
            art_name = str(row[c_art_name]).strip() if row[c_art_name] is not None else ''
            
            last_gr_val = row[c_last_gr]
            last_sale_val = row[c_last_sale]
            
            val_dt = safe_float(row[c_dt]) or 0.0
            val_qty = safe_float(row[c_qty]) or 0.0
            val_val = safe_float(row[c_val]) or 0.0
            val_val_d15 = safe_float(row[c_val_d15]) or 0.0
            
            last_gr_dt = excel_date_to_datetime(last_gr_val)
            last_sale_dt = excel_date_to_datetime(last_sale_val)
            
            if last_gr_val == ">90" or last_gr_val is None or last_gr_val == "":
                days_no_gr = ">90"
            else:
                if last_gr_dt:
                    diff = (milestone_date - last_gr_dt).days
                    days_no_gr = ">90" if diff > 90 else diff
                else:
                    days_no_gr = ">90"
            
            p_info = product_map.get(art_code)
            # date_num nằm ở index 8 của tuple product_map (KHÔNG dùng [-1] vì đã
            # nối thêm mch4_desc ở index 9 -> [-1] sẽ trỏ nhầm sang chuỗi mch4_desc).
            ref_date = p_info[8] if p_info and len(p_info) > 8 else 0.0
            
            val_dt_90 = val_dt * 90
            val_dio_d15 = 9999.0 if val_dt_90 == 0.0 else (val_val_d15 / val_dt_90)
            val_dio_date = 9999.0 if ref_date == 0.0 else (val_val_d15 / ref_date)
            
            note = ""
            if last_sale_val == ">90" and last_gr_val == ">90":
                note = "không giao dịch >90 ngày"
            elif last_sale_val == ">90" and last_gr_dt:
                note = "Không sale 90 ngày"
                
            note_check_slow = ""
            if val_dio_date > 5.0 and val_val > ref_amount:
                note_check_slow = "Check"
                
            note_expired = ""
            if days_no_gr == ">90":
                if ref_date < 90:
                    note_expired = "Hết hạn"
            else:
                if isinstance(days_no_gr, (int, float)) and days_no_gr > ref_date:
                    note_expired = "Hết hạn"
                    
            phan_loai = ""
            if note_expired == "Hết hạn":
                phan_loai = "Hết hạn"
            elif note == "không giao dịch >90 ngày":
                phan_loai = "Nghi vấn tồn ảo"
            elif note == "Không sale 90 ngày":
                phan_loai = "Non-moving"
            elif note_check_slow == "Check":
                phan_loai = "Slow moving"
                
            if phan_loai == "":
                rows_filtered += 1
                continue
                
            computed = {
                'Y': days_no_gr,
                'Z': ref_date,
                'AA': val_dio_d15,
                'AB': val_dio_date,
                'AC': note,
                'AD': note_check_slow,
                'AE': note_expired,
                'AF': phan_loai
            }
            
            mch2_id = mch2_desc = mch3_id = mch3_desc = mch4_desc = ''
            mch5_desc = 'Khác'
            if p_info:
                mch2_id = p_info[1] or ''
                mch2_desc = p_info[2] or ''
                mch3_id = p_info[3] or ''
                mch3_desc = p_info[4] or ''
                mch4_desc = (p_info[9] if len(p_info) > 9 and p_info[9] else '')
                mch5_desc = (p_info[7] if len(p_info) > 7 and p_info[7] else 'Khác')

            row_list = [None] * 31
            row_list[0] = _row_val(row, c_rsm)             # B - RSM (giá trị thô, sẽ được ghi đè bên dưới sau khi tra store_map)
            row_list[1] = _row_val(row, c_asm)              # C - ASM (giá trị thô, sẽ được ghi đè bên dưới)
            row_list[2] = store_code                        # D - Store Code
            row_list[3] = store_name                        # E - Store Name
            row_list[4] = mch2_id                           # F - MCH2 ID
            row_list[5] = mch2_desc                         # G - MCH2 Desc
            row_list[6] = mch3_id                           # H - MCH3 ID
            row_list[7] = mch3_desc                         # I - MCH3 Desc
            row_list[8] = mch4_desc                         # J - MCH4 Desc
            row_list[9] = art_code                          # K - Article Code
            row_list[10] = art_name                         # L - Article Name
            row_list[11] = _row_val(row, c_created_on)      # M - Created on
            row_list[12] = _row_val(row, c_ton_dau_ky)      # N - Tồn đầu kỳ (D-90)
            row_list[13] = _row_val(row, c_nhap_trong_ky)   # O - Nhập trong kỳ
            row_list[14] = _row_val(row, c_xuat_trong_ky)   # P - Xuất trong kỳ
            row_list[15] = val_qty                          # Q - Tồn cuối kỳ (D)
            row_list[16] = val_val                          # R - Giá trị tồn
            row_list[17] = val_dt                           # S - Doanh thu
            row_list[18] = _row_val(row, c_gia_von)         # T - Giá vốn
            row_list[19] = _row_val(row, c_ton_d15)         # U - Tồn D-15
            row_list[20] = val_val_d15                      # V - Giá tồn D-15

            store_info = store_map.get(store_code)
            if store_info:
                rsm_name, asm_name, store_name, sheet_type = store_info[:4]
            else:
                store_name = store_name
                if store_code.isdigit():
                    rsm_name = row_list[0] or ''
                    asm_name = store_name
                    sheet_type = 'WM'
                else:
                    rsm_name = row_list[0] or ''
                    asm_name = row_list[1] or ''
                    sheet_type = 'WM+'
                    
            row_list[0] = rsm_name
            row_list[1] = asm_name
            row_list[3] = store_name
            
            for col_l, key in [('W', 'W'), ('X', 'X'), ('Y', 'Y'), ('Z', 'Z'), ('AA', 'AA'),
                                ('AB', 'AB'), ('AC', 'AC'), ('AD', 'AD'),
                                ('AE', 'AE'), ('AF', 'AF')]:
                row_list[col_to_idx(col_l) - 1] = computed.get(key)

            written_rows.append((row_list, phan_loai, sheet_type, mch5_desc))

        has_wm = any(item[2] == 'WM' for item in written_rows)

        headers = {
            'B': 'RSM' if not has_wm else 'RSM / GĐV/GĐM', 'C': 'ASM', 'D': 'Store Code', 'E': 'Store Name',
            'F': 'MCH2 ID', 'G': 'MCH2 Desc', 'H': 'MCH3 ID', 'I': 'MCH3 Desc',
            'J': 'MCH4 Desc', 'K': 'Article Code', 'L': 'Article Name',
            'M': 'Created on', 'N': 'Tồn đầu kỳ (D-90)',
            'O': 'Nhập trong kỳ', 'P': 'Xuất trong kỳ',
            'Q': 'Tồn cuối kỳ (D)', 'R': 'Giá trị tồn',
            'S': 'Doanh thu', 'T': 'Giá vốn', 'U': 'Tồn D-15',
            'V': 'Giá tồn D-15', 'W': 'Last GR', 'X': 'Last Sale',
            'Y': 'Số ngày ko nhập', 'Z': 'Date tham khảo',
            'AA': 'DIO (D-15)', 'AB': 'DIO/Date', 'AC': 'Note',
            'AD': 'Note check slow', 'AE': 'note hết hạn',
            'AF': 'Phân loại kiểm tra', 'AG': 'MCH5 Desc'
        }
        
        cells_out2 = []
        for col_l, label in sorted(headers.items(), key=lambda x: col_to_idx(x[0])):
            cell = WriteOnlyCell(ws_out, value=label)
            cell.font = font_white_bold
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin
            cells_out2.append(cell)
        ws_out.append(cells_out2)
        
        col_letters_out = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
                           'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                           'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB',
                           'AC', 'AD', 'AE', 'AF']

        dashboard_rows = []
        for row_list, phan_loai, sheet_type, mch5_desc in written_rows:
            cells_temp = []
            for col_out_idx in range(31):
                col_letter = col_letters_out[col_out_idx]
                val = row_list[col_out_idx]

                if val is None or val == '':
                    cells_temp.append(None)
                    continue

                if col_letter in ('M', 'W', 'X'):
                    dt = excel_date_to_datetime(val) if not (col_letter in ('W', 'X') and str(val).strip() == '>90') else None
                    if dt:
                        cell = WriteOnlyCell(ws_out, value=dt)
                        cell.number_format = 'dd/mm/yyyy'
                        cell.font = font_normal
                        cell.alignment = align_center
                    else:
                        cell = WriteOnlyCell(ws_out, value=val)
                        cell.font = font_normal
                        cell.alignment = align_center
                    cells_temp.append(cell)

                elif col_letter in ('N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'Y', 'Z'):
                    try:
                        if col_letter == 'Y' and str(val).strip() == '>90':
                            cell = WriteOnlyCell(ws_out, value=val)
                            cell.font = font_normal
                            cell.alignment = align_right
                        else:
                            num_val = int(float(val))
                            cell = WriteOnlyCell(ws_out, value=num_val)
                            cell.number_format = '#,##0'
                            cell.font = font_normal
                            cell.alignment = align_right
                    except (ValueError, TypeError):
                        cell = WriteOnlyCell(ws_out, value=val)
                        cell.font = font_normal
                        cell.alignment = align_right
                    cells_temp.append(cell)

                elif col_letter in ('AA', 'AB'):
                    try:
                        num_val = float(val)
                        cell = WriteOnlyCell(ws_out, value=num_val)
                        cell.number_format = '0.00'
                        cell.font = font_normal
                        cell.alignment = align_right
                    except (ValueError, TypeError):
                        cell = WriteOnlyCell(ws_out, value=val)
                        cell.font = font_normal
                        cell.alignment = align_right
                    cells_temp.append(cell)

                elif col_letter == 'AF':
                    fill_af, font_af = phan_loai_style.get(phan_loai, (None, font_normal))
                    cell = WriteOnlyCell(ws_out, value=val)
                    cell.font = font_af
                    if fill_af:
                        cell.fill = fill_af
                    cell.alignment = align_center
                    cells_temp.append(cell)

                else:
                    cells_temp.append(val)

            # Cột AG: Nhóm hàng (MCH5) — chỉ ghi ra file, không đưa vào dashboard_rows (Tab 1)
            if mch5_desc:
                cell_mch5 = WriteOnlyCell(ws_out, value=mch5_desc)
                cell_mch5.font = font_normal
                cell_mch5.alignment = align_center
                cells_temp.append(cell_mch5)
            else:
                cells_temp.append(None)

            ws_out.append(cells_temp)
            rows_written += 1
            dashboard_rows.append(row_list)

        wb_out.save(out_path)
        if rows_written > 0:
            generate_html_dashboard(out_path, dashboard_rows, product_map, log_func, milestone_date, chain_title)
        elapsed = time.time() - t0
        if log_func:
            log_func(f"  ✓ Hoàn thành: {rows_written:,} hàng xuất, {rows_filtered:,} hàng lọc bỏ")
            log_func(f"  → Thời gian: {elapsed:.1f}s | File: {out_filename}")
        return out_path
    except Exception as e:
        if log_func:
            log_func(f"  ✗ Lỗi xử lý {filename}: {e}")
            import traceback
            log_func(traceback.format_exc())
        return None

# ===================================================================
# PHẦN 2: LOGIC TỔNG HỢP CHUỖI (OVERVIEW)
# ===================================================================
def _read_single_done_file(path: str) -> tuple[str, list[dict], int, bool, str | None]:
    """Đọc 1 file *(done).xlsx, trả về (filename, rows, row_count, has_wm, error_msg)."""
    filename = os.path.basename(path)
    rows_out = []
    has_wm_local = False
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_name = 'Dữ liệu kiểm tra'
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]

        header_row = None
        header_row_idx = -1
        for r_idx, row in enumerate(sheet.iter_rows(max_row=5, values_only=True), start=1):
            if row and any(h in row for h in ['RSM', 'Store Code', 'Phân loại kiểm tra']):
                header_row = row
                header_row_idx = r_idx
                break

        if header_row is None:
            for r_idx, row in enumerate(sheet.iter_rows(max_row=1, values_only=True), start=1):
                header_row = row
                header_row_idx = r_idx
                break

        if header_row is None:
            wb.close()
            return filename, [], 0, False, f"  ✗ Thất bại: Không tìm thấy tiêu đề cột trong file {filename}"

        header_map = {}
        for c_idx, h in enumerate(header_row):
            if h is not None:
                header_map[str(h).strip()] = c_idx

        idx_rsm = header_map.get('RSM', header_map.get('RSM / GĐV/GĐM', 0))
        idx_asm = header_map.get('ASM', 1)
        idx_store_code = header_map.get('Store Code', 2)
        idx_store_name = header_map.get('Store Name', 3)
        idx_mch2_desc = header_map.get('MCH2 Desc', 5)
        idx_mch3_desc = header_map.get('MCH3 Desc', 7)
        idx_mch4_id = header_map.get('MCH4 Desc', header_map.get('MCH4 ID', 8))
        idx_mch5_desc = header_map.get('MCH5 Desc', header_map.get('Nhóm hàng (MCH5)', -1))
        idx_qty = header_map.get('Tồn cuối kỳ (D)', 14)
        idx_val = header_map.get('Giá trị tồn', 15)
        idx_class = header_map.get('Phân loại kiểm tra', 26)

        row_count = 0
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if r_idx <= header_row_idx:
                continue

            if not row or all(v is None for v in row):
                continue

            rsm = str(row[idx_rsm]).strip() if idx_rsm < len(row) and row[idx_rsm] is not None else ''
            asm = str(row[idx_asm]).strip() if idx_asm < len(row) and row[idx_asm] is not None else ''
            store_code = str(row[idx_store_code]).strip() if idx_store_code < len(row) and row[idx_store_code] is not None else ''
            store_name = str(row[idx_store_name]).strip() if idx_store_name < len(row) and row[idx_store_name] is not None else ''
            mch2 = str(row[idx_mch2_desc]).strip() if idx_mch2_desc < len(row) and row[idx_mch2_desc] is not None else ''
            mch3 = str(row[idx_mch3_desc]).strip() if 0 <= idx_mch3_desc < len(row) and row[idx_mch3_desc] is not None else ''
            mch4 = str(row[idx_mch4_id]).strip() if 0 <= idx_mch4_id < len(row) and row[idx_mch4_id] is not None else ''
            mch5 = str(row[idx_mch5_desc]).strip() if 0 <= idx_mch5_desc < len(row) and row[idx_mch5_desc] is not None else ''
            check_class = str(row[idx_class]).strip() if idx_class < len(row) and row[idx_class] is not None else ''

            try:
                qty = float(row[idx_qty]) if idx_qty < len(row) and row[idx_qty] is not None else 0.0
            except (ValueError, TypeError):
                qty = 0.0

            try:
                val = float(row[idx_val]) if idx_val < len(row) and row[idx_val] is not None else 0.0
            except (ValueError, TypeError):
                val = 0.0

            if not check_class or check_class.upper() in ('N/A', 'NONE', ''):
                continue

            if not has_wm_local and (store_code.isdigit() or 'WM' in store_code.upper()):
                has_wm_local = True

            rows_out.append({
                'rsm': rsm,
                'asm': asm,
                'store_code': store_code,
                'store_name': store_name,
                'mch2': mch2,
                'mch3': mch3,
                'mch4': mch4,
                'mch5': mch5,
                'qty': qty,
                'val': val,
                'check_class': check_class
            })
            row_count += 1

        wb.close()
        return filename, rows_out, row_count, has_wm_local, None

    except Exception as e:
        return filename, [], 0, False, f"  ✗ Lỗi khi đọc file {filename}: {e}"


def read_and_aggregate_files(file_paths, log_func=None) -> tuple[list[dict], datetime | None, bool]:
    all_data = []
    milestone_date = None
    has_wm = False

    total_files = len(file_paths)
    # Đọc các file song song bằng luồng — I/O (đọc đĩa/OneDrive) chiếm phần lớn
    # thời gian nên chạy đa luồng giúp chồng lấp thời gian chờ giữa các file,
    # thay vì đọc tuần tự từng file một.
    max_workers = min(8, max(1, total_files))
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_path = {executor.submit(_read_single_done_file, path): path for path in file_paths}
        done_count = 0
        for future in as_completed(future_to_path):
            done_count += 1
            filename, rows_out, row_count, has_wm_local, err = future.result()
            if log_func:
                log_func(f"  [{done_count}/{total_files}] Đã đọc: {filename}")
            if err:
                if log_func:
                    log_func(err, "error")
                continue
            all_data.extend(rows_out)
            if has_wm_local:
                has_wm = True
            if log_func:
                log_func(f"  ✓ Đã nạp {row_count:,} dòng hợp lệ.")

    if not all_data:
        return [], None, False
        
    if log_func:
        log_func(f"\n💡 Đang gộp dữ liệu để tối ưu hóa dung lượng...")
        
    aggregated = {}
    for item in all_data:
        key = (
            item['rsm'],
            item['asm'],
            item['store_code'],
            item['store_name'],
            item['mch2'],
            item.get('mch3', ''),
            item.get('mch4', ''),
            item.get('mch5', ''),
            item['check_class']
        )
        if key not in aggregated:
            aggregated[key] = {'qty': 0.0, 'val': 0.0, 'count': 0}
        aggregated[key]['qty'] += item['qty']
        aggregated[key]['val'] += item['val']
        aggregated[key]['count'] += 1
        
    result_rows = []
    for key, metrics in aggregated.items():
        result_rows.append({
            'rsm': key[0],
            'asm': key[1],
            'store_code': key[2],
            'store_name': key[3],
            'mch2': key[4],
            'mch3': key[5],
            'mch4': key[6],
            'mch5': key[7],
            'check_class': key[8],
            'qty': metrics['qty'],
            'val': metrics['val'],
            'count': metrics['count']
        })
        
    if log_func:
        log_func(f"  ✓ Gộp dữ liệu hoàn tất. Giảm từ {len(all_data):,} xuống {len(result_rows):,} dòng tổng hợp (Tiết kiệm >99% dung lượng).")
        
    return result_rows, milestone_date, has_wm

def generate_overview_html(out_html_path: str, rows_data: list, has_wm: bool, 
                           log_func=None, target_date=None, chain_title="Tổng Quan Chuỗi WIN") -> bool:
    try:
        t0 = time.time()
        template_path = os.path.join(BASE_DIR, DEFAULT_TEMPLATE_NAME)

        if not os.path.exists(template_path):
            template_path = os.path.join(BASE_DIR, "Tool_PhanTichTonKho_Portable", DEFAULT_TEMPLATE_NAME)
            
        if not os.path.exists(template_path):
            if log_func:
                log_func(f"  ✗ Lỗi: Không tìm thấy file template '{DEFAULT_TEMPLATE_NAME}'", "error")
            return False
            
        rsm_set = set()
        asm_set = set()
        store_map = {}
        mch2_set = set()
        mch3_set = set()
        mch4_set = set()
        mch5_set = set()
        class_set = set()

        for r in rows_data:
            rsm_set.add(r['rsm'])
            if r['asm']:
                asm_set.add(r['asm'])
            store_map[r['store_code']] = r['store_name']
            mch2_set.add(r['mch2'])
            mch3_set.add(r.get('mch3', ''))
            mch4_set.add(r.get('mch4', ''))
            mch5_set.add(r.get('mch5', ''))
            class_set.add(r['check_class'])

        rsm_list = sorted(list(rsm_set))
        asm_list = sorted(list(asm_set))
        store_list = [[code, name] for code, name in sorted(store_map.items())]
        mch2_list = sorted(list(mch2_set))
        mch3_list = sorted(list(mch3_set))
        mch4_list = sorted(list(mch4_set))
        mch5_list = sorted(list(mch5_set))
        class_list = sorted(list(class_set))

        rsm_idx_map = {name: idx for idx, name in enumerate(rsm_list)}
        asm_idx_map = {name: idx for idx, name in enumerate(asm_list)}
        store_idx_map = {code: idx for idx, [code, name] in enumerate(store_list)}
        mch2_idx_map = {name: idx for idx, name in enumerate(mch2_list)}
        mch3_idx_map = {name: idx for idx, name in enumerate(mch3_list)}
        mch4_idx_map = {name: idx for idx, name in enumerate(mch4_list)}
        mch5_idx_map = {name: idx for idx, name in enumerate(mch5_list)}
        class_idx_map = {name: idx for idx, name in enumerate(class_list)}
        
        compressed_rows = []
        for r in rows_data:
            store_idx = store_idx_map.get(r['store_code'], -1)
            rsm_idx = rsm_idx_map.get(r['rsm'], -1)
            asm_idx = asm_idx_map.get(r['asm'], -1)
            mch2_idx = mch2_idx_map.get(r['mch2'], -1)
            class_idx = class_idx_map.get(r['check_class'], -1)
            
            mch3_idx = mch3_idx_map.get(r.get('mch3', ''), -1)
            mch4_idx = mch4_idx_map.get(r.get('mch4', ''), -1)
            mch5_idx = mch5_idx_map.get(r.get('mch5', ''), -1)

            comp_row = [
                store_idx,
                rsm_idx,
                asm_idx,
                mch2_idx,
                class_idx,
                round(r['qty']),
                int(r['val']),
                r['count'],
                mch3_idx,
                mch4_idx,
                mch5_idx
            ]
            compressed_rows.append(comp_row)

        out_dict = {
            "rsm_list": rsm_list,
            "asm_list": asm_list,
            "store_list": store_list,
            "mch2_list": mch2_list,
            "mch3_list": mch3_list,
            "mch4_list": mch4_list,
            "mch5_list": mch5_list,
            "class_list": class_list,
            "has_wm": has_wm,
            "rows": compressed_rows
        }
        
        with open(template_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
            
        html_content = html_content.replace(
            "Báo Cáo Đối Soát Dữ Liệu Kiểm Tra - Tổng Quan Chuỗi",
            f"Báo Cáo Đối Soát Dữ Liệu Kiểm Tra - {chain_title}"
        )
        html_content = html_content.replace(
            "Tổng Quan Chuỗi WIN",
            chain_title
        )
        
        html_content = html_content.replace(
            "Cập nhật: 01/07/2026",
            f"Cập nhật: {target_date or datetime.now().strftime('%d/%m/%Y')}"
        )
        
        js_injection = f"const dbData = {json.dumps(out_dict, ensure_ascii=False, separators=(',', ':'))};\n"
        js_injection += "const rawData = dbData.rows.map(r => ({\n"
        js_injection += "    store_idx: r[0],\n"
        js_injection += "    store_code: dbData.store_list[r[0]][0],\n"
        js_injection += "    store_name: dbData.store_list[r[0]][1],\n"
        js_injection += "    rsm: dbData.rsm_list[r[1]],\n"
        js_injection += "    rsm_idx: r[1],\n"
        js_injection += "    asm: r[2] !== -1 ? dbData.asm_list[r[2]] : '',\n"
        js_injection += "    asm_idx: r[2],\n"
        js_injection += "    mch2: dbData.mch2_list[r[3]],\n"
        js_injection += "    mch2_idx: r[3],\n"
        js_injection += "    check_class: dbData.class_list[r[4]],\n"
        js_injection += "    class_idx: r[4],\n"
        js_injection += "    qty: r[5],\n"
        js_injection += "    value: r[6],\n"
        js_injection += "    count: r[7],\n"
        js_injection += "    mch3: r[8] !== -1 ? dbData.mch3_list[r[8]] : '',\n"
        js_injection += "    mch3_idx: r[8],\n"
        js_injection += "    mch4: r[9] !== -1 ? dbData.mch4_list[r[9]] : '',\n"
        js_injection += "    mch4_idx: r[9],\n"
        js_injection += "    mch5: r[10] !== -1 ? dbData.mch5_list[r[10]] : '',\n"
        js_injection += "    mch5_idx: r[10]\n"
        js_injection += "}));"
        
        html_content = html_content.replace('// DATA_PLACEHOLDER', js_injection)
        
        with open(out_html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
            
        elapsed = time.time() - t0
        if log_func:
            log_func(f"  ✓ Đã sinh file HTML: {os.path.basename(out_html_path)} | Thời gian: {elapsed:.1f}s")
        return True
        
    except Exception as e:
        if log_func:
            log_func(f"  ✗ Thất bại khi sinh file HTML: {e}", "error")
        return False

# ===================================================================
# GIAO DIỆN NGƯỜI DÙNG TÍCH HỢP (GUI)
# ===================================================================
class App(tk.Tk):
    BG_DARK = '#1a1d2e'
    BG_CARD = '#252840'
    BG_INPUT = '#1e2235'
    ACCENT = '#4f8ef7'
    ACCENT_HOVER = '#6ba3ff'
    SUCCESS = '#4caf50'
    WARNING = '#ff9800'
    ERROR = '#f44336'
    TEXT = '#e8eaed'
    TEXT_DIM = '#9aa0b8'
    BORDER = '#353a5e'

    def __init__(self):
        super().__init__()
        self.title(f"{APP_TITLE} v{APP_VERSION}")
        self.geometry("1020x760")
        self.minsize(850, 650)
        self.configure(bg=self.BG_DARK)

        self.selected_files_t1 = []
        self.master_path = tk.StringVar(value=DEFAULT_MASTER_PATH)
        self.store_path = tk.StringVar(value=DEFAULT_STORE_PATH)
        self.milestone_date_str = tk.StringVar(value=datetime.now().strftime('%d/%m/%Y'))
        self.ref_amount_str = tk.StringVar(value='100,000')
        self.chain_title_var_t1 = tk.StringVar(value="Tổng quan chuỗi WIN")
        self.is_running_t1 = False

        self.selected_files_t2 = []
        self.chain_title_var = tk.StringVar(value="Tổng quan chuỗi WIN")
        self.target_date_str = tk.StringVar(value=datetime.now().strftime('%d/%m/%Y'))
        self.is_running_t2 = False

        self._build_ui()
        self._center_window()

    def _center_window(self):
        self.update_idletasks()
        w = self.winfo_width()
        h = self.winfo_height()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _build_ui(self):
        header = tk.Frame(self, bg='#131629', pady=12)
        header.pack(fill='x')

        tk.Label(
            header, text="📈  Báo Cáo Đối Soát & Phân Tích Tồn Kho",
            font=('Segoe UI', 16, 'bold'), bg='#131629', fg=self.TEXT
        ).pack(side='left', padx=20)

        tk.Label(
            header, text=f"v{APP_VERSION}", font=('Segoe UI', 9),
            bg='#131629', fg=self.TEXT_DIM
        ).pack(side='right', padx=20)

        style = ttk.Style()
        style.theme_use('default')
        style.configure('TNotebook', background=self.BG_DARK, borderwidth=0)
        style.configure('TNotebook.Tab', background=self.BG_CARD, foreground=self.TEXT, 
                        padding=[15, 6], font=('Segoe UI', 10, 'bold'), borderwidth=0)
        style.map('TNotebook.Tab', 
                  background=[('selected', self.ACCENT)], 
                  foreground=[('selected', 'white')])

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill='both', expand=True, padx=15, pady=10)

        self.tab1 = tk.Frame(self.notebook, bg=self.BG_DARK)
        self.notebook.add(self.tab1, text="  🔍  Phân Tích Chi Tiết  ")
        self._build_tab1_ui()

        self.tab2 = tk.Frame(self.notebook, bg=self.BG_DARK)
        self.notebook.add(self.tab2, text="  📊  Tổng Quan Chuỗi  ")
        self._build_tab2_ui()

    def _build_tab1_ui(self):
        content = tk.Frame(self.tab1, bg=self.BG_DARK)
        content.pack(fill='both', expand=True, pady=10)

        left = tk.Frame(content, bg=self.BG_DARK, width=380)
        left.pack(side='left', fill='y', padx=(0, 10))
        left.pack_propagate(False)

        right = tk.Frame(content, bg=self.BG_DARK)
        right.pack(side='right', fill='both', expand=True)

        c1 = self._card(left, "1. File dữ liệu SAP đầu vào (.xlsx, .csv)")
        self.file_listbox_t1 = tk.Listbox(
            c1, font=('Segoe UI', 9), bg=self.BG_INPUT, fg=self.TEXT,
            selectbackground=self.ACCENT, selectforeground='white',
            height=5, relief='flat', bd=0, highlightthickness=1, 
            highlightcolor=self.BORDER, highlightbackground=self.BORDER
        )
        self.file_listbox_t1.pack(fill='x', pady=(0, 6))

        btn_frame = tk.Frame(c1, bg=self.BG_CARD)
        btn_frame.pack(fill='x')
        self._btn(btn_frame, "📂 Thêm file", self._add_files_t1).pack(side='left')
        self._btn(btn_frame, "✖ Xóa", self._remove_selected_t1, color='#e53935').pack(side='left', padx=4)
        self._btn(btn_frame, "🗑 Xóa hết", self._clear_files_t1, color='#555').pack(side='left')

        self.file_count_lbl_t1 = tk.Label(
            c1, text="Chưa chọn file nào.",
            font=('Segoe UI', 8), bg=self.BG_CARD, fg=self.TEXT_DIM, anchor='w'
        )
        self.file_count_lbl_t1.pack(fill='x', pady=(4, 0))

        c2 = self._card(left, "2. File danh mục & store list")
        tk.Label(c2, text="Danh mục sản phẩm (Master Article):", font=('Segoe UI', 8), bg=self.BG_CARD, fg=self.TEXT_DIM, anchor='w').pack(fill='x', pady=(0, 2))
        f_master = tk.Frame(c2, bg=self.BG_CARD)
        f_master.pack(fill='x', pady=(0, 6))
        self.master_entry = tk.Entry(f_master, textvariable=self.master_path, font=('Segoe UI', 8), bg=self.BG_INPUT, fg=self.TEXT, relief='flat', highlightthickness=1, highlightcolor=self.BORDER, highlightbackground=self.BORDER)
        self.master_entry.pack(side='left', fill='x', expand=True)
        self._btn(f_master, "📂", self._browse_master).pack(side='right', padx=(4, 0))

        tk.Label(c2, text="Danh sách cửa hàng (Store List):", font=('Segoe UI', 8), bg=self.BG_CARD, fg=self.TEXT_DIM, anchor='w').pack(fill='x', pady=(0, 2))
        f_store = tk.Frame(c2, bg=self.BG_CARD)
        f_store.pack(fill='x')
        self.store_entry = tk.Entry(f_store, textvariable=self.store_path, font=('Segoe UI', 8), bg=self.BG_INPUT, fg=self.TEXT, relief='flat', highlightthickness=1, highlightcolor=self.BORDER, highlightbackground=self.BORDER)
        self.store_entry.pack(side='left', fill='x', expand=True)
        self._btn(f_store, "📂", self._browse_store).pack(side='right', padx=(4, 0))

        c3 = self._card(left, "3. Tham số phân tích")
        self._field(c3, "Ngày mốc (Y1):", self.milestone_date_str, tip="Định dạng: DD/MM/YYYY")
        self.ref_amount_entry = self._field(c3, "Số tiền tham chiếu (AF1):", self.ref_amount_str, tip="Đơn vị: VNĐ (ví dụ: 100,000)")
        self.ref_amount_entry.bind("<KeyRelease>", self._format_ref_amount)
        
        tk.Label(c3, text="Chuỗi báo cáo hiển thị:", font=('Segoe UI', 9), bg=self.BG_CARD, fg=self.TEXT, anchor='w').pack(fill='x', pady=(6, 2))
        self.chain_combo_t1 = ttk.Combobox(
            c3, textvariable=self.chain_title_var_t1, 
            values=["Tổng quan chuỗi WIN", "Tổng quan chuỗi Rural", "Tổng quan chuỗi Urban", "Tổng quan chuỗi winmart"],
            state="readonly", font=('Segoe UI', 9)
        )
        self.chain_combo_t1.pack(fill='x', pady=(0, 2))
        self.chain_combo_t1.current(0)

        log_title_frame = tk.Frame(right, bg=self.BG_DARK)
        log_title_frame.pack(fill='x', pady=(0, 4))
        tk.Label(log_title_frame, text="Nhật ký xử lý chi tiết", font=('Segoe UI', 11, 'bold'), bg=self.BG_DARK, fg=self.TEXT).pack(side='left')
        
        self.run_btn_t1 = tk.Button(
            log_title_frame, text="▶  BẮT ĐẦU XỬ LÝ CHI TIẾT", font=('Segoe UI', 10, 'bold'),
            bg=self.ACCENT, fg='white', activebackground=self.ACCENT_HOVER, activeforeground='white',
            cursor='hand2', relief='flat', padx=15, pady=4, command=self._on_run_t1
        )
        self.run_btn_t1.pack(side='right')

        log_frame = tk.Frame(right, bg=self.BG_CARD)
        log_frame.pack(fill='both', expand=True)

        self.log_text_t1 = tk.Text(
            log_frame, font=('Consolas', 9), bg=self.BG_INPUT, fg=self.TEXT,
            relief='flat', bd=0, insertbackground=self.TEXT, highlightthickness=0, state='disabled'
        )
        self.log_text_t1.pack(side='left', fill='both', expand=True, padx=1, pady=1)

        scrollbar = ttk.Scrollbar(log_frame, orient='vertical', command=self.log_text_t1.yview)
        scrollbar.pack(side='right', fill='y')
        self.log_text_t1.configure(yscrollcommand=scrollbar.set)

        self.log_text_t1.tag_config('info', foreground=self.TEXT)
        self.log_text_t1.tag_config('success', foreground=self.SUCCESS)
        self.log_text_t1.tag_config('warning', foreground=self.WARNING)
        self.log_text_t1.tag_config('error', foreground=self.ERROR)
        self.log_text_t1.tag_config('highlight', foreground=self.ACCENT, font=('Consolas', 9, 'bold'))

    def _build_tab2_ui(self):
        content = tk.Frame(self.tab2, bg=self.BG_DARK)
        content.pack(fill='both', expand=True, pady=10)

        left = tk.Frame(content, bg=self.BG_DARK, width=380)
        left.pack(side='left', fill='y', padx=(0, 10))
        left.pack_propagate(False)

        right = tk.Frame(content, bg=self.BG_DARK)
        right.pack(side='right', fill='both', expand=True)

        c1 = self._card(left, "1. File dữ liệu chi tiết đã xử lý *(done).xlsx")
        self.file_listbox_t2 = tk.Listbox(
            c1, font=('Segoe UI', 9), bg=self.BG_INPUT, fg=self.TEXT,
            selectbackground=self.ACCENT, selectforeground='white',
            height=8, relief='flat', bd=0, highlightthickness=1, 
            highlightcolor=self.BORDER, highlightbackground=self.BORDER
        )
        self.file_listbox_t2.pack(fill='x', pady=(0, 6))

        btn_frame = tk.Frame(c1, bg=self.BG_CARD)
        btn_frame.pack(fill='x')
        self._btn(btn_frame, "📂 Thêm file", self._add_files_t2).pack(side='left')
        self._btn(btn_frame, "✖ Xóa", self._remove_selected_t2, color='#e53935').pack(side='left', padx=4)
        self._btn(btn_frame, "🗑 Xóa hết", self._clear_files_t2, color='#555').pack(side='left')

        self.file_count_lbl_t2 = tk.Label(
            c1, text="Chưa chọn file nào.",
            font=('Segoe UI', 8), bg=self.BG_CARD, fg=self.TEXT_DIM, anchor='w'
        )
        self.file_count_lbl_t2.pack(fill='x', pady=(4, 0))

        c2 = self._card(left, "2. Cấu hình tiêu đề & ngày cập nhật")
        
        tk.Label(c2, text="Chuỗi báo cáo hiển thị:", font=('Segoe UI', 9), bg=self.BG_CARD, fg=self.TEXT, anchor='w').pack(fill='x', pady=(0, 2))
        self.chain_combo = ttk.Combobox(
            c2, textvariable=self.chain_title_var, 
            values=["Tổng quan chuỗi WIN", "Tổng quan chuỗi Rural", "Tổng quan chuỗi Urban", "Tổng quan chuỗi winmart"],
            state="readonly", font=('Segoe UI', 9)
        )
        self.chain_combo.pack(fill='x', pady=(0, 10))
        self.chain_combo.current(0)
        
        self._field(c2, "Ngày cập nhật hiển thị:", self.target_date_str, tip="Định dạng: DD/MM/YYYY")

        c3 = self._card(left, "💡 Lưu ý")
        tk.Label(
            c3, text="• Đầu vào bắt buộc là các file có đuôi *(done).xlsx\n• Tool sẽ tự động gộp dữ liệu để tối ưu kích thước\n• File HTML xuất ra không chứa bảng chi tiết sản phẩm\n• Giao diện Sáng mặc định và hiển thị breakdown đầy đủ\n  trên tooltip biểu đồ.",
            font=('Segoe UI', 8), bg=self.BG_CARD, fg=self.TEXT_DIM, justify='left', anchor='w'
        ).pack(fill='x')

        log_title_frame = tk.Frame(right, bg=self.BG_DARK)
        log_title_frame.pack(fill='x', pady=(0, 4))
        tk.Label(log_title_frame, text="Nhật ký tổng hợp chuỗi", font=('Segoe UI', 11, 'bold'), bg=self.BG_DARK, fg=self.TEXT).pack(side='left')
        
        self.run_btn_t2 = tk.Button(
            log_title_frame, text="📊  BẮT ĐẦU TỔNG HỢP CHUỖI", font=('Segoe UI', 10, 'bold'),
            bg=self.ACCENT, fg='white', activebackground=self.ACCENT_HOVER, activeforeground='white',
            cursor='hand2', relief='flat', padx=15, pady=4, command=self._on_run_t2
        )
        self.run_btn_t2.pack(side='right')

        log_frame = tk.Frame(right, bg=self.BG_CARD)
        log_frame.pack(fill='both', expand=True)

        self.log_text_t2 = tk.Text(
            log_frame, font=('Consolas', 9), bg=self.BG_INPUT, fg=self.TEXT,
            relief='flat', bd=0, insertbackground=self.TEXT, highlightthickness=0, state='disabled'
        )
        self.log_text_t2.pack(side='left', fill='both', expand=True, padx=1, pady=1)

        scrollbar = ttk.Scrollbar(log_frame, orient='vertical', command=self.log_text_t2.yview)
        scrollbar.pack(side='right', fill='y')
        self.log_text_t2.configure(yscrollcommand=scrollbar.set)

        self.log_text_t2.tag_config('info', foreground=self.TEXT)
        self.log_text_t2.tag_config('success', foreground=self.SUCCESS)
        self.log_text_t2.tag_config('warning', foreground=self.WARNING)
        self.log_text_t2.tag_config('error', foreground=self.ERROR)
        self.log_text_t2.tag_config('highlight', foreground=self.ACCENT, font=('Consolas', 9, 'bold'))

    def _card(self, parent, title: str) -> tk.Frame:
        outer = tk.Frame(parent, bg=self.BG_CARD)
        outer.pack(fill='x', pady=5)
        tk.Label(
            outer, text=title, font=('Segoe UI', 9, 'bold'),
            bg=self.BG_CARD, fg=self.ACCENT, anchor='w', padx=12, pady=6
        ).pack(fill='x')
        tk.Frame(outer, bg=self.BORDER, height=1).pack(fill='x', padx=12)
        inner = tk.Frame(outer, bg=self.BG_CARD, padx=12, pady=10)
        inner.pack(fill='x')
        return inner

    def _field(self, parent, label: str, var: tk.StringVar, tip: str = ''):
        frame = tk.Frame(parent, bg=self.BG_CARD)
        frame.pack(fill='x', pady=3)
        tk.Label(
            frame, text=label, font=('Segoe UI', 9),
            bg=self.BG_CARD, fg=self.TEXT, anchor='w', width=24
        ).pack(side='left')
        entry = tk.Entry(
            frame, textvariable=var, font=('Segoe UI', 9),
            bg=self.BG_INPUT, fg=self.TEXT, relief='flat', insertbackground=self.TEXT,
            highlightthickness=1, highlightcolor=self.ACCENT, highlightbackground=self.BORDER
        )
        entry.pack(side='left', fill='x', expand=True)
        if tip:
            tk.Label(
                parent, text=tip, font=('Segoe UI', 7),
                bg=self.BG_CARD, fg=self.TEXT_DIM, anchor='e'
            ).pack(fill='x')
        return entry

    def _btn(self, parent, text: str, cmd, color=None) -> tk.Button:
        bg_color = color if color else self.ACCENT
        active_bg = self.ACCENT_HOVER if not color else color
        btn = tk.Button(
            parent, text=text, font=('Segoe UI', 8, 'bold'),
            bg=bg_color, fg='white', activebackground=active_bg,
            activeforeground='white', relief='flat', cursor='hand2',
            padx=8, pady=3, command=cmd
        )
        return btn

    def _format_ref_amount(self, event=None):
        try:
            cursor_pos = self.ref_amount_entry.index(tk.INSERT)
        except Exception:
            cursor_pos = 0
        raw = self.ref_amount_str.get()
        commas_before = raw[:cursor_pos].count(',')
        clean = "".join([c for c in raw if c.isdigit()])
        if clean:
            try:
                formatted = f"{int(clean):,}"
                self.ref_amount_str.set(formatted)
                digits_before = cursor_pos - commas_before
                new_pos = 0
                digit_count = 0
                for char in formatted:
                    if digit_count == digits_before:
                        break
                    if char.isdigit():
                        digit_count += 1
                    new_pos += 1
                self.ref_amount_entry.icursor(new_pos)
            except (ValueError, TypeError):
                pass
        else:
            self.ref_amount_str.set("")

    def log_t1(self, message: str, tag: str = 'info'):
        self.log_text_t1.config(state='normal')
        self.log_text_t1.insert(tk.END, message + "\n", tag)
        self.log_text_t1.see(tk.END)
        self.log_text_t1.config(state='disabled')
        self.update_idletasks()

    def _add_files_t1(self):
        files = filedialog.askopenfilenames(
            title="Chọn file báo cáo tồn kho",
            filetypes=[("Excel/CSV Files", "*.xlsx;*.csv"), ("Excel Files", "*.xlsx"), ("CSV Files", "*.csv")]
        )
        for f in files:
            norm_f = os.path.abspath(f)
            if norm_f not in self.selected_files_t1:
                self.selected_files_t1.append(norm_f)
                self.file_listbox_t1.insert(tk.END, os.path.basename(norm_f))
        self._update_file_count_t1()

    def _remove_selected_t1(self):
        selected_indices = list(self.file_listbox_t1.curselection())
        for idx in reversed(selected_indices):
            self.selected_files_t1.pop(idx)
            self.file_listbox_t1.delete(idx)
        self._update_file_count_t1()

    def _clear_files_t1(self):
        self.selected_files_t1.clear()
        self.file_listbox_t1.delete(0, tk.END)
        self._update_file_count_t1()

    def _update_file_count_t1(self):
        count = len(self.selected_files_t1)
        if count == 0:
            self.file_count_lbl_t1.config(text="Chưa chọn file nào.", fg=self.TEXT_DIM)
        else:
            self.file_count_lbl_t1.config(text=f"Đã chọn {count} file.", fg=self.SUCCESS)

    def _browse_master(self):
        f = filedialog.askopenfilename(
            title="Chọn file danh mục sản phẩm (Master Article)",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if f:
            self.master_path.set(os.path.abspath(f))

    def _browse_store(self):
        f = filedialog.askopenfilename(
            title="Chọn file danh sách siêu thị (Store List)",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if f:
            self.store_path.set(os.path.abspath(f))

    def log_t2(self, message: str, tag: str = 'info'):
        self.log_text_t2.config(state='normal')
        self.log_text_t2.insert(tk.END, message + "\n", tag)
        self.log_text_t2.see(tk.END)
        self.log_text_t2.config(state='disabled')
        self.update_idletasks()

    def _add_files_t2(self):
        files = filedialog.askopenfilenames(
            title="Chọn file dữ liệu chi tiết *(done).xlsx",
            filetypes=[("Excel Done Files", "*(done).xlsx"), ("Excel Files", "*.xlsx")]
        )
        for f in files:
            norm_f = os.path.abspath(f)
            if norm_f not in self.selected_files_t2:
                self.selected_files_t2.append(norm_f)
                self.file_listbox_t2.insert(tk.END, os.path.basename(norm_f))
        self._update_file_count_t2()

    def _remove_selected_t2(self):
        selected_indices = list(self.file_listbox_t2.curselection())
        for idx in reversed(selected_indices):
            self.selected_files_t2.pop(idx)
            self.file_listbox_t2.delete(idx)
        self._update_file_count_t2()

    def _clear_files_t2(self):
        self.selected_files_t2.clear()
        self.file_listbox_t2.delete(0, tk.END)
        self._update_file_count_t2()

    def _update_file_count_t2(self):
        count = len(self.selected_files_t2)
        if count == 0:
            self.file_count_lbl_t2.config(text="Chưa chọn file nào.", fg=self.TEXT_DIM)
        else:
            self.file_count_lbl_t2.config(text=f"Đã chọn {count} file.", fg=self.SUCCESS)

    def _on_run_t1(self):
        if self.is_running_t1:
            return
        if not self.selected_files_t1:
            messagebox.showerror("Lỗi", "Vui lòng chọn ít nhất 1 file báo cáo đầu vào!")
            return
        master = self.master_path.get().strip()
        if not os.path.exists(master):
            messagebox.showerror("Lỗi", f"Không tìm thấy file danh mục sản phẩm tại:\n{master}")
            return
        store = self.store_path.get().strip()
        if not os.path.exists(store):
            messagebox.showerror("Lỗi", f"Không tìm thấy file danh sách siêu thị tại:\n{store}")
            return
        date_str = self.milestone_date_str.get().strip()
        try:
            milestone_date = datetime.strptime(date_str, '%d/%m/%Y')
        except ValueError:
            messagebox.showerror("Lỗi", "Ngày mốc không đúng định dạng DD/MM/YYYY!")
            return
        ref_str = self.ref_amount_str.get().strip().replace(',', '').replace('.', '')
        try:
            ref_amount = float(ref_str)
        except ValueError:
            messagebox.showerror("Lỗi", "Số tiền tham chiếu phải là một số hợp lệ!")
            return

        self.is_running_t1 = True
        self.run_btn_t1.config(state='disabled', text="⏳  ĐANG XỬ LÝ...")
        self.log_text_t1.config(state='normal')
        self.log_text_t1.delete('1.0', tk.END)
        self.log_text_t1.config(state='disabled')

        def worker():
            t_start = time.time()
            self.log_t1(f"=== BẮT ĐẦU XỬ LÝ CHI TIẾT ===", 'highlight')
            self.log_t1(f"Ngày mốc: {date_str} | Số tiền tham chiếu: {ref_amount:,.0f} VNĐ\n")
            
            store_map = load_store_details(store, lambda msg, t='info': self.log_t1(msg, t))
            product_map = load_product_details(master, lambda msg, t='info': self.log_t1(msg, t))
            shelf_life_map = {k: v[6] for k, v in product_map.items()}

            success_count = 0
            for idx, file in enumerate(self.selected_files_t1):
                filename = os.path.basename(file)
                self.log_t1(f"\n[{idx+1}/{len(self.selected_files_t1)}] Đang xử lý: {filename}...", 'highlight')
                chain_title_t1 = self.chain_title_var_t1.get().strip()
                try:
                    out = process_excel_file(
                        file, store_map, product_map, milestone_date, ref_amount,
                        log_func=lambda msg, t='info': self.log_t1(msg, t),
                        chain_title=chain_title_t1,
                        shelf_life_map=shelf_life_map
                    )
                    if out:
                        success_count += 1
                except Exception as ex:
                    self.log_t1(f"  ✗ Lỗi không xác định: {ex}", 'error')

            elapsed = time.time() - t_start
            self.log_t1(f"\n=== TIẾN TRÌNH HOÀN TẤT ===", 'highlight')
            self.log_t1(f"Đã xử lý thành công {success_count}/{len(self.selected_files_t1)} file.", 'success')
            self.log_t1(f"Tổng thời gian thực hiện: {elapsed:.1f} giây.\n", 'success')
            
            if success_count > 0 and self.selected_files_t1:
                try:
                    os.startfile(os.path.dirname(self.selected_files_t1[0]))
                except Exception:
                    pass

            self.is_running_t1 = False
            self.run_btn_t1.config(state='normal', text="▶  BẮT ĐẦU XỬ LÝ CHI TIẾT")

        Thread(target=worker, daemon=True).start()

    def _on_run_t2(self):
        if self.is_running_t2:
            return
        if not self.selected_files_t2:
            messagebox.showerror("Lỗi", "Vui lòng chọn ít nhất 1 file dữ liệu chi tiết *(done).xlsx!")
            return
        date_str = self.target_date_str.get().strip()
        if date_str:
            try:
                datetime.strptime(date_str, '%d/%m/%Y')
            except ValueError:
                messagebox.showerror("Lỗi", "Ngày cập nhật phải đúng định dạng DD/MM/YYYY!")
                return
        
        chain_title = self.chain_title_var.get().strip()

        self.is_running_t2 = True
        self.run_btn_t2.config(state='disabled', text="⏳  ĐANG TỔNG HỢP...")
        self.log_text_t2.config(state='normal')
        self.log_text_t2.delete('1.0', tk.END)
        self.log_text_t2.config(state='disabled')

        def worker():
            t_start = time.time()
            self.log_t2(f"=== BẮT ĐẦU TỔNG HỢP TỔNG QUAN CHUỖI ===", 'highlight')
            self.log_t2(f"Báo cáo: {chain_title}")
            self.log_t2(f"Ngày hiển thị: {date_str if date_str else 'Hôm nay'}\n")
            
            aggregated_rows, milestone_date, has_wm = read_and_aggregate_files(
                self.selected_files_t2,
                log_func=lambda msg, t='info': self.log_t2(msg, t)
            )
            
            if aggregated_rows:
                out_dir = os.path.dirname(self.selected_files_t2[0])
                file_slug = chain_title.lower().replace(" ", "_").replace("ố", "o").replace("ổ", "o").replace("ô", "o").replace("u", "u").replace("â", "a")
                out_html_path = os.path.join(out_dir, f"dashboard_{file_slug}.html")
                
                success = generate_overview_html(
                    out_html_path, aggregated_rows, has_wm,
                    log_func=lambda msg, t='info': self.log_t2(msg, t),
                    target_date=date_str,
                    chain_title=chain_title
                )
                
                if success:
                    self.log_t2(f"\n✓ Đã tạo báo cáo thành công tại: {out_html_path}", 'success')
                    try:
                        os.startfile(out_html_path)
                    except Exception:
                        pass
            else:
                self.log_t2("\n✗ Không có dữ liệu để tổng hợp.", 'error')

            elapsed = time.time() - t_start
            self.log_t2(f"\n=== TIẾN TRÌNH HOÀN TẤT ===", 'highlight')
            self.log_t2(f"Tổng thời gian thực hiện: {elapsed:.1f} giây.\n", 'success')

            self.is_running_t2 = False
            self.run_btn_t2.config(state='normal', text="📊  BẮT ĐẦU TỔNG HỢP CHUỖI")

        Thread(target=worker, daemon=True).start()

if __name__ == '__main__':
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass

    app = App()
    app.mainloop()
