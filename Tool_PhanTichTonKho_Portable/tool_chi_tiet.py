# -*- coding: utf-8 -*-
"""
TOOL ĐỔ DỮ LIỆU PHÂN TÍCH TỒN KHO
Phiên bản: 1.1.0
Mô tả: Tự động tính toán các chỉ số tồn kho, tra cứu hạn sử dụng tham khảo
       và phân loại sản phẩm cần kiểm tra theo đúng công thức nghiệp vụ.
       Tối ưu hóa hiệu năng xử lý tốc độ cao (One-pass stream).
"""

import os
import sys
import time
import zipfile
import xml.etree.ElementTree as ET
import ctypes
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, date
from threading import Thread
import tempfile
import shutil

# Force UTF-8 output
if sys.stdout is not None:
    sys.stdout.reconfigure(encoding='utf-8')


# ===================================================================
# CẤU HÌNH MẶC ĐỊNH
# ===================================================================
# Hỗ trợ chạy cả dạng script thường và dạng đóng gói (.exe)
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def find_default_file(filename):
    # Candidate 1: Same directory as script (BASE_DIR)
    p = os.path.abspath(os.path.join(BASE_DIR, filename))
    if os.path.exists(p):
        return p
    # Candidate 2: Parent directory
    p = os.path.abspath(os.path.join(BASE_DIR, "..", filename))
    if os.path.exists(p):
        return p
    # Candidate 3: Sibling directory 'Last GR.SALE'
    p = os.path.abspath(os.path.join(BASE_DIR, "..", "Last GR.SALE", filename))
    if os.path.exists(p):
        return p
    # Candidate 4: Sibling's parent directory 'Last GR.SALE' (when run inside Tool_PhanTichTonKho_Portable)
    p = os.path.abspath(os.path.join(BASE_DIR, "..", "..", "Last GR.SALE", filename))
    if os.path.exists(p):
        return p
    # Fallback default
    return os.path.abspath(os.path.join(BASE_DIR, "..", "Last GR.SALE", filename))

DEFAULT_MASTER_PATH = find_default_file("Master article 7.7.xlsx")
DEFAULT_STORE_PATH = find_default_file("Store List.xlsx")

APP_VERSION = "1.4.5"
APP_TITLE = "Tool Đổ Dữ Liệu Phân Tích Tồn Kho"

# Thư mục chứa file tạm
LAST_GR_SALE_DIR = os.path.dirname(DEFAULT_MASTER_PATH)

if os.path.exists(LAST_GR_SALE_DIR):
    SCRATCH_DIR = os.path.join(LAST_GR_SALE_DIR, "temp")
    os.makedirs(SCRATCH_DIR, exist_ok=True)
else:
    SCRATCH_DIR = tempfile.mkdtemp(prefix="tool_do_du_lieu_")




# ===================================================================
# HÀM TRỢ GIÚP
# ===================================================================
def col_to_idx(col_letter: str) -> int:
    """Chuyển đổi tên cột Excel (A, B, ..., AA, AB) sang index 0-based."""
    val = 0
    for char in col_letter.upper():
        val = val * 26 + (ord(char) - 64)
    return val - 1


def get_column_letter(col_idx: int) -> str:
    """Chuyển đổi index 1-based sang chữ cái cột Excel (ví dụ: 1->A, 27->AA)."""
    result = []
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result.append(chr(65 + remainder))
    return ''.join(reversed(result))


def get_actual_col(col_letter: str, offset: int) -> str:
    """Trả về ký tự cột thực tế dựa trên offset của cột RSM.
    
    col_letter: Ký tự cột chuẩn của file gốc (ví dụ: 'K', 'W', 'X')
    offset: 1 nếu RSM ở B (file chuẩn), 0 nếu RSM ở A (file khuyết cột A)
    """
    if offset == 1:
        return col_letter
    idx = col_to_idx(col_letter)
    actual_idx = idx - (1 - offset)
    if actual_idx < 0:
        return ''
    return get_column_letter(actual_idx + 1)


def safe_float(val) -> float | None:
    """Chuyển đổi giá trị sang float an toàn."""
    if val is None or val == '':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def excel_date_to_datetime(val) -> datetime | None:
    """Chuyển đổi số serial Excel hoặc chuỗi ngày sang đối tượng datetime."""
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        # Excel bug: 1900 được coi là năm nhuận, nên mốc bắt đầu thực tế là 1899-12-30
        try:
            return datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(val))
        except (ValueError, OverflowError):
            return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    
    # Thử parse định dạng chuỗi
    val_str = str(val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d.%m.%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(val_str, fmt)
        except ValueError:
            continue
    return None


# ===================================================================
# THAO TÁC FILE XLSB (TRA CỨU HẠN SỬ DỤNG)
# ===================================================================
def load_store_details(store_path: str, log_func=None) -> dict:
    """
    Đọc danh sách siêu thị từ Store List.xlsx.
    Trả về dict: {store_id: (rsm_name, asm_name, store_name, sheet_type)}
    """
    store_map = {}
    if not os.path.exists(store_path):
        if log_func:
            log_func(f"  [CẢNH BÁO] Không tìm thấy file Store List tại: {store_path}")
        return store_map

    if log_func:
        log_func(f"  Đang đọc danh sách siêu thị từ: {os.path.basename(store_path)}")

    t0 = time.time()
    try:
        import openpyxl
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, f"temp_{os.path.basename(store_path)}")
        res = ctypes.windll.kernel32.CopyFileW(store_path, temp_path, False)
        if not res:
            if log_func:
                log_func("  ✗ Không thể copy file Store List (đang mở?)")
            return store_map

        wb = openpyxl.load_workbook(temp_path, read_only=True)
        
        def parse_wm_sheet(ws_obj):
            headers = next(ws_obj.iter_rows(values_only=True))
            idx_code = 0
            idx_name = 1
            idx_gdv = 2
            if headers:
                for idx, val in enumerate(headers):
                    if val:
                        val_str = str(val).strip().lower()
                        if any(x in val_str for x in ('code', 'store', 'mã', 'ma ch')) and not any(x in val_str for x in ('name', 'tên')):
                            idx_code = idx
                        elif any(x in val_str for x in ('name', 'tên', 'ten ch')):
                            idx_name = idx
                        elif any(x in val_str for x in ('gđv', 'gđm', 'gdv', 'gdm', 'giám sát', 'giam sat', 'rsm')):
                            idx_gdv = idx
            return idx_code, idx_name, idx_gdv

        # Duyệt qua tất cả các sheets của file Store List
        sheet_count = 0
        for sheetname in wb.sheetnames:
            if sheetname == '_com.sap.ip.bi.xl.hiddensheet':
                continue
            ws = wb[sheetname]
            sheet_count += 1
            
            # Phân loại sheet: Nếu tên sheet chứa 'wm' và không có '+' thì xem như sheet WM
            # Nếu tên sheet là 'WM+' hoặc 'Sheet1' (hoặc sheet mặc định) thì xem như sheet WM+
            sheet_type = 'WM+'
            if 'wm' in sheetname.lower() and '+' not in sheetname:
                sheet_type = 'WM'
                
            if sheet_type == 'WM':
                idx_code, idx_name, idx_gdv = parse_wm_sheet(ws)
                is_first = True
                for row in ws.iter_rows(values_only=True):
                    if is_first:
                        is_first = False
                        continue
                    if len(row) > max(idx_code, idx_name, idx_gdv):
                        store_id = str(row[idx_code]).strip() if row[idx_code] is not None else ''
                        store_name = str(row[idx_name]).strip() if row[idx_name] is not None else ''
                        gdv = str(row[idx_gdv]).strip() if row[idx_gdv] is not None else ''
                        if store_id:
                            if store_id.endswith('.0'):
                                store_id = store_id[:-2]
                            # Lưu vào map: (GĐV/GĐM, ASM=StoreName, StoreName, sheet_type)
                            store_map[store_id] = (gdv, store_name, store_name, 'WM')
            else:
                is_first = True
                for row in ws.iter_rows(values_only=True):
                    if is_first:
                        is_first = False
                        continue
                    if len(row) >= 4:
                        rsm = str(row[0]).strip() if row[0] is not None else ''
                        asm = str(row[1]).strip() if row[1] is not None else ''
                        store_id = str(row[2]).strip() if row[2] is not None else ''
                        store_name = str(row[3]).strip() if row[3] is not None else ''
                        if store_id:
                            if store_id.endswith('.0'):
                                store_id = store_id[:-2]
                            # Lưu vào map: (RSM, ASM, StoreName, sheet_type)
                            store_map[store_id] = (rsm, asm, store_name, 'WM+')
                            
        wb.close()
        try:
            os.remove(temp_path)
        except Exception:
            pass
        t1 = time.time()
        if log_func:
            log_func(f"  ✓ Đọc xong danh sách siêu thị từ {sheet_count} sheet: {len(store_map):,} cửa hàng | Thời gian: {t1 - t0:.1f}s")
    except Exception as e:
        if log_func:
            log_func(f"  ✗ Lỗi nạp Store List: {e}")
    return store_map



def load_product_details(master_path: str, log_func=None) -> dict:
    """
    Đọc danh mục sản phẩm từ Master article.xlsx.
    Trả về dict: {product_id: (product_name, mch2_id, mch2_desc, mch3_id, mch3_desc, mch4_id, total_shelf_life)}
    """
    product_map = {}
    if not os.path.exists(master_path):
        if log_func:
            log_func(f"  [CẢNH BÁO] Không tìm thấy file danh mục sản phẩm tại: {master_path}")
        return product_map

    if log_func:
        log_func(f"  Đang đọc danh mục sản phẩm từ: {os.path.basename(master_path)}")

    t0 = time.time()
    try:
        import openpyxl
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, f"temp_{os.path.basename(master_path)}")
        res = ctypes.windll.kernel32.CopyFileW(master_path, temp_path, False)
        if not res:
            if log_func:
                log_func("  ✗ Không thể copy file danh mục sản phẩm (đang mở?)")
            return product_map

        wb = openpyxl.load_workbook(temp_path, read_only=True)
        # Sử dụng sheet Export
        ws = wb['Export'] if 'Export' in wb.sheetnames else wb.active
        
        is_first = True
        for row in ws.iter_rows(values_only=True):
            if is_first:
                is_first = False
                continue
            if len(row) >= 26:
                prod_id = str(row[0]).strip() if row[0] is not None else ''
                prod_name = str(row[1]).strip() if row[1] is not None else ''
                mch2_id = str(row[15]).strip() if row[15] is not None else ''
                mch2_desc = str(row[16]).strip() if row[16] is not None else ''
                mch3_id = str(row[17]).strip() if row[17] is not None else ''
                mch3_desc = str(row[18]).strip() if row[18] is not None else ''
                mch4_id = str(row[19]).strip() if row[19] is not None else ''
                
                shelf_life = None
                if row[25] is not None:
                    try:
                        shelf_life = int(float(row[25]))
                    except (ValueError, TypeError):
                        pass
                
                mch5_desc = str(row[22]).strip() if len(row) > 22 and row[22] is not None else 'Khác'
                
                if prod_id:
                    if prod_id.endswith('.0'):
                        prod_id = prod_id[:-2]
                    product_map[prod_id] = (prod_name, mch2_id, mch2_desc, mch3_id, mch3_desc, mch4_id, shelf_life, mch5_desc)
        wb.close()
        try:
            os.remove(temp_path)
        except Exception:
            pass
        t1 = time.time()
        if log_func:
            log_func(f"  ✓ Đọc xong danh mục sản phẩm: {len(product_map):,} sản phẩm | Thời gian: {t1 - t0:.1f}s")
    except Exception as e:
        if log_func:
            log_func(f"  ✗ Lỗi nạp danh mục sản phẩm: {e}")
    return product_map



# ===================================================================
# LOGIC TÍNH TOÁN NGHIỆP VỤ (FORMULAS)
# ===================================================================
def compute_row(row_dict: dict, milestone_date: datetime, ref_amount: float,
                shelf_life_map: dict) -> dict:
    """
    Tính toán các chỉ số cho một dòng dữ liệu (Row 3+).
    Tham số truyền vào row_dict đã được chuẩn hóa về các cột chuẩn (B->X).
    """
    result = {}
    
    art_code_raw = row_dict.get('K')
    if art_code_raw is None:
        art_code_raw = ''
    art_code_str = str(art_code_raw).strip()
    if art_code_str.endswith('.0'):
        art_code_str = art_code_str[:-2]

    # Trích xuất các giá trị số
    ton_dau_ky = safe_float(row_dict.get('N')) or 0
    nhap_trong_ky = safe_float(row_dict.get('O')) or 0
    xuat_trong_ky = safe_float(row_dict.get('P')) or 0
    ton_cuoi_ky = safe_float(row_dict.get('Q')) or 0
    gia_tri_ton = safe_float(row_dict.get('R')) or 0
    doanh_thu = safe_float(row_dict.get('S')) or 0
    gia_von = safe_float(row_dict.get('T')) or 0
    ton_d15 = safe_float(row_dict.get('U')) or 0
    gia_ton_d15 = safe_float(row_dict.get('V')) or 0

    # ================================================================
    # Đọc Last GR (W) và Last Sale (X) - có thể là số serial hoặc ">90"
    # Khi giá trị > 90 ngày trước ngày mốc, SAP xuất ">90" dạng text
    # ================================================================
    last_gr_raw = row_dict.get('W')
    last_sale_raw = row_dict.get('X')

    # Xác định xem Last GR và Last Sale có phải ">90" hoặc bị trống không
    last_gr_is_empty = (last_gr_raw is None or str(last_gr_raw).strip() == '' or str(last_gr_raw).strip().lower() == 'none')
    last_sale_is_empty = (last_sale_raw is None or str(last_sale_raw).strip() == '' or str(last_sale_raw).strip().lower() == 'none')

    last_gr_gt90 = (str(last_gr_raw).strip() == '>90' or last_gr_is_empty)
    last_sale_gt90 = (str(last_sale_raw).strip() == '>90' or last_sale_is_empty)

    last_gr = None if last_gr_gt90 else excel_date_to_datetime(last_gr_raw)
    last_sale = None if last_sale_gt90 else excel_date_to_datetime(last_sale_raw)

    # Lưu lại giá trị W và X sau khi xử lý (điền >90 nếu trống)
    result['W'] = '>90' if last_gr_gt90 else last_gr_raw
    result['X'] = '>90' if last_sale_gt90 else last_sale_raw


    # --- Col Y: Số ngày không nhập ---
    # Chỉnh sửa: số ngày không nhập trừ quá 90 hoặc không có giá trị số thì điền ">90"
    if last_gr_gt90 or last_gr is None or milestone_date is None:
        result['Y'] = '>90'
        y_days = 91               # Dùng 91 để so sánh logic nội bộ (>90)
    else:
        delta = milestone_date - last_gr
        diff_days = delta.days
        if diff_days > 90:
            result['Y'] = '>90'
            y_days = 91
        else:
            result['Y'] = diff_days
            y_days = diff_days

    # --- Col Z: Date tham khảo = tra cứu từ XLSB theo mã SP ---
    z_shelf_life = shelf_life_map.get(art_code_str)
    if z_shelf_life is None:
        orig_z = row_dict.get('Z')
        if orig_z == '#N/A' or orig_z is None:
            z_shelf_life = '#N/A'
        else:
            z_shelf_life = safe_float(orig_z)
    result['Z'] = z_shelf_life

    # --- Col AA: DIO (D-15) ---
    # Công thức mới: số lượng tồn U / số lượng bán P * 90 ngày
    if xuat_trong_ky == 0:
        aa_val = 9999  # Giống công thức Excel
    else:
        aa_val = (ton_d15 / xuat_trong_ky) * 90
    result['AA'] = aa_val


    # Nếu Z là '#N/A', propagate '#N/A' cho AB, AD, AE, AF
    if z_shelf_life == '#N/A':
        result['AB'] = '#N/A'
        result['AD'] = '#N/A'
        result['AE'] = '#N/A'
        result['AF'] = '#N/A'
        # Tính AC độc lập
        ac_note = ''
        if last_sale_gt90:
            if last_gr_gt90:
                ac_note = 'không giao dịch >90 ngày'
            else:
                ac_note = 'Không sale 90 ngày'
        result['AC'] = ac_note if ac_note else None
        return result

    # --- Col AB: DIO/Date = AA / Z ---
    if aa_val is not None and z_shelf_life is not None and z_shelf_life > 0:
        result['AB'] = aa_val / z_shelf_life
    else:
        result['AB'] = None
    ab_val = result['AB']

    # --- Col AC: Note ---
    # Công thức: IF(X3=">90", IF(W3=">90","không giao dịch >90 ngày","Không sale 90 ngày"), "")
    # → Last Sale ">90": không có bán trong 90 ngày
    #   - Nếu Last GR cũng ">90": "không giao dịch >90 ngày" (Nghi vấn tồn ảo)
    #   - Nếu Last GR không phải ">90": "Không sale 90 ngày" (Non-moving)
    # → Last Sale không phải ">90": AC = rỗng (có bán gần đây)
    la_nghi_van = False
    la_non_moving = False
    ac_note = ''

    if last_sale_gt90:
        if last_gr_gt90:
            ac_note = 'không giao dịch >90 ngày'
            la_nghi_van = True
        else:
            ac_note = 'Không sale 90 ngày'
            la_non_moving = True
    result['AC'] = ac_note if ac_note else None

    # --- Col AD: Note check slow ---
    # Công thức: IF(AND(AB3>5,R3>$AF$1),"Check","")
    if ab_val is not None and ab_val > 5 and gia_tri_ton > ref_amount:
        result['AD'] = 'Check'
    else:
        result['AD'] = None

    # --- Col AE: note hết hạn ---
    # Công thức: IF(Y3=">90", IF(Z3<90,"Hết hạn",""), IF(Y3>Z3,"Hết hạn",""))
    # → Nếu Y = ">90" (Last GR không rõ/ảo): hết hạn nếu Date tham khảo < 90 ngày
    # → Nếu Y là số ngày: hết hạn nếu Số ngày ko nhập > Date tham khảo
    ae_note = ''
    if z_shelf_life is not None:
        if last_gr_gt90 or y_days == 91:  # Y = ">90"
            if z_shelf_life < 90:
                ae_note = 'Hết hạn'
        elif y_days is not None:
            if y_days > z_shelf_life:
                ae_note = 'Hết hạn'
    result['AE'] = ae_note if ae_note else None

    # --- Col AF: Phân loại kiểm tra ---
    # Công thức: IF(AE3<>"","Hết hạn",
    #               IF(AC3="Không giao dịch >90 ngày","Nghi vấn tồn ảo",
    #                  IF(AC3="Không sale 90 ngày","Non-moving",
    #                     IF(AD3<>"","Slow moving",""))))
    phan_loai = None

    if ae_note:
        phan_loai = 'Hết hạn'
    elif la_nghi_van:
        phan_loai = 'Nghi vấn tồn ảo'
    elif la_non_moving:
        phan_loai = 'Non-moving'
    elif result['AD'] == 'Check':
        phan_loai = 'Slow moving'

    result['AF'] = phan_loai
    return result


# ===================================================================
# XỬ LÝ CHÍNH EXCEL XML STREAM (ONE-PASS OPTIMIZED)
def process_csv_file(src_path: str, store_map: dict, product_map: dict, milestone_date: datetime,
                     ref_amount: float, log_func=None) -> str | None:
    """
    Xử lý file dữ liệu CSV mẫu mới, nạp thông tin từ store_map và product_map.
    Lọc và giữ lại các hàng có Phân loại kiểm tra.
    Trả về đường dẫn file output XLSX hoặc None nếu lỗi.
    """
    import csv
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.cell import WriteOnlyCell

    filename = os.path.basename(src_path)
    base, ext = os.path.splitext(filename)
    out_filename = f"{base} (done).xlsx"
    out_path = os.path.join(os.path.dirname(src_path), out_filename)

    # Kiểm tra xem file kết quả có bị khóa (đang mở trong Excel) không trước khi xử lý
    if os.path.exists(out_path):
        try:
            with open(out_path, 'a'):
                pass
        except PermissionError:
            if log_func:
                log_func(f"  ✗ [LỖI KHÓA FILE] File kết quả '{out_filename}' đang mở trong Excel. Vui lòng đóng file này và chạy lại!", 'error')
            return None

    if log_func:
        log_func(f"  Bắt đầu xử lý (CSV)...")

    t0 = time.time()
    temp_path = os.path.join(SCRATCH_DIR, f"temp_{filename}")


    try:
        # Sao chép file để tránh lock
        res = ctypes.windll.kernel32.CopyFileW(src_path, temp_path, False)
        if not res:
            if log_func:
                log_func(f"  ✗ Không thể copy file (đang mở?): {filename}")
            return None

        # Định nghĩa styles
        font_white_bold = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        font_normal = Font(name='Arial', size=9)
        fill_header = PatternFill(fill_type='solid', start_color='1F4E78', end_color='1F4E78')
        
        fill_het_han = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
        fill_nghi_van = PatternFill(fill_type='solid', start_color='FF6600', end_color='FF6600')
        fill_non_moving = PatternFill(fill_type='solid', start_color='FFA500', end_color='FFA500')
        fill_slow_moving = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
        font_slow = Font(name='Arial', size=9, color='000000')
        
        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')

        # Màu cho từng phân loại
        phan_loai_style = {
            'Hết hạn': (fill_het_han, Font(name='Arial', size=9, bold=True, color='FFFFFF')),
            'Nghi vấn tồn ảo': (fill_nghi_van, Font(name='Arial', size=9, bold=True, color='FFFFFF')),
            'Non-moving': (fill_non_moving, Font(name='Arial', size=9, bold=True, color='000000')),
            'Slow moving': (fill_slow_moving, font_slow),
        }

        wb_out = openpyxl.Workbook(write_only=True)
        ws_out = wb_out.create_sheet(title='Dữ liệu kiểm tra')

        # 27 cột kết quả
        headers = [
            'RSM', 'ASM', 'Store Code', 'Store Name', 
            'MCH2 ID', 'MCH2 Desc', 'MCH3 ID', 'MCH3 Desc', 'MCH4 ID', 
            'Article Code', 'Article Name', 
            'Tồn đầu kỳ (D-90)', 'Nhập trong kỳ', 'Xuất trong kỳ', 
            'Tồn cuối kỳ (D)', 'Giá trị tồn', 'Tồn D-15', 
            'Last GR', 'Last Sale', 'Số ngày ko nhập', 'Date tham khảo', 
            'DIO (D-15)', 'DIO/Date', 'Note', 'Note check slow', 'note hết hạn', 
            'Phân loại kiểm tra'
        ]

        # Trích xuất date_tham_khao map cho compute_row
        shelf_life_map = {k: v[6] for k, v in product_map.items()}

        rows_written = 0
        rows_filtered = 0
        total_rows_read = 0
        written_rows = []

        # Phát hiện dấu phân cách trong CSV
        with open(temp_path, 'r', encoding='utf-8', errors='ignore') as f:
            first_line = f.readline()
            delimiter = ';' if ';' in first_line else ','
        
        with open(temp_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            
            for row in reader:
                total_rows_read += 1
                
                store_id = str(row.get('STORE_ID', '')).strip()
                if store_id.endswith('.0'):
                    store_id = store_id[:-2]
                    
                prod_id = str(row.get('PRODUCT_ID', '')).strip()
                if prod_id.endswith('.0'):
                    prod_id = prod_id[:-2]

                # Bỏ qua hàng tổng kết (Result)
                store_name_raw = str(row.get('STORE_NAME', '')).strip()
                if 'result' in store_id.lower() or 'result' in store_name_raw.lower():
                    rows_filtered += 1
                    continue

                # Tra cứu thông tin để sửa lỗi tiếng Việt và lấy các cột thiếu
                store_info = store_map.get(store_id)
                if store_info:
                    rsm_name, asm_name, store_name, sheet_type = store_info[:4]
                else:
                    store_name = store_name_raw
                    # Tự động nhận diện siêu thị (WM) qua định dạng Store ID (số)
                    if store_id.isdigit():
                        rsm_name = row.get('GDV', '') or row.get('RSM', '') or ''
                        asm_name = store_name_raw
                        sheet_type = 'WM'
                    else:
                        rsm_name = row.get('GDV', '') or row.get('RSM', '') or ''
                        asm_name = row.get('ASM', '') or ''
                        sheet_type = 'WM+'
                
                prod_info = product_map.get(prod_id, (row.get('PRODUCT_NAME', ''), '', '', '', '', '', None, 'Khác'))
                prod_name, mch2_id, mch2_desc, mch3_id, mch3_desc, mch4_id, *rest = prod_info

                # Bỏ qua các sản phẩm thuộc MCH2 là "101" hoặc thiếu thông tin MCH2/MCH3
                if mch2_id == '101' or not mch2_id or not mch3_id or str(mch2_id).strip() == '' or str(mch3_id).strip() == '':
                    rows_filtered += 1
                    continue

                norm_row = {
                    'K': prod_id,
                    'N': row.get('CLOSING_STOCK_QUANTITY_D90'),
                    'O': row.get('GR_QTY_LAST90_D'),
                    'P': row.get('GI_QTY_LAST90_D'),
                    'Q': row.get('CLOSING_STOCK_QUANTITY_LASTDAY'),
                    'R': row.get('CLOSING_STOCK_VALUE_LASTDAY'),
                    'S': 0,
                    'T': 0,
                    'U': row.get('CLOSING_STOCK_QUANTITY_D15'),
                    'V': 0,
                    'W': row.get('MAX_GR_DATE'),
                    'X': row.get('MAX_SALE_DATE')
                }

                computed = compute_row(norm_row, milestone_date, ref_amount, shelf_life_map)
                phan_loai = computed.get('AF')

                if phan_loai is None or str(phan_loai).strip() == '':
                    rows_filtered += 1
                    continue

                # Xây dựng row_list 27 phần tử theo cấu trúc mới
                row_list = [None] * 27
                row_list[0] = rsm_name
                row_list[1] = asm_name
                row_list[2] = store_id
                row_list[3] = store_name
                row_list[4] = mch2_id
                row_list[5] = mch2_desc
                row_list[6] = mch3_id
                row_list[7] = mch3_desc
                row_list[8] = mch4_id
                row_list[9] = prod_id
                row_list[10] = prod_name
                row_list[11] = norm_row['N']
                row_list[12] = norm_row['O']
                row_list[13] = norm_row['P']
                row_list[14] = norm_row['Q']
                row_list[15] = norm_row['R']
                row_list[16] = norm_row['U']
                row_list[17] = computed['W']
                row_list[18] = computed['X']
                row_list[19] = computed['Y']
                row_list[20] = computed['Z']
                row_list[21] = computed['AA']
                row_list[22] = computed['AB']
                row_list[23] = computed['AC']
                row_list[24] = computed['AD']
                row_list[25] = computed['AE']
                row_list[26] = computed['AF']

                written_rows.append((row_list, phan_loai, sheet_type))

        # Sau khi lặp xong và collect all rows in written_rows:
        has_wm = any(item[2] == 'WM' for item in written_rows)
        has_wm_plus = any(item[2] == 'WM+' for item in written_rows)
        use_asm = True
        
        # 1. Ghi Row 1: Headers (Đã loại bỏ Row 1 cũ là Nhãn tham chiếu)
        if use_asm:
            headers = [
                'RSM' if not has_wm else 'RSM / GĐV/GĐM', 'ASM', 'Store Code', 'Store Name', 
                'MCH2 ID', 'MCH2 Desc', 'MCH3 ID', 'MCH3 Desc', 'MCH4 ID', 
                'Article Code', 'Article Name', 
                'Tồn đầu kỳ (D-90)', 'Nhập trong kỳ', 'Xuất trong kỳ', 
                'Tồn cuối kỳ (D)', 'Giá trị tồn', 'Tồn D-15', 
                'Last GR', 'Last Sale', 'Số ngày ko nhập', 'Date tham khảo', 
                'DIO (D-15)', 'DIO/Date', 'Note', 'Note check slow', 'note hết hạn', 
                'Phân loại kiểm tra'
            ]
        else:
            headers = [
                'GĐV/GĐM', 'Store Code', 'Store Name', 
                'MCH2 ID', 'MCH2 Desc', 'MCH3 ID', 'MCH3 Desc', 'MCH4 ID', 
                'Article Code', 'Article Name', 
                'Tồn đầu kỳ (D-90)', 'Nhập trong kỳ', 'Xuất trong kỳ', 
                'Tồn cuối kỳ (D)', 'Giá trị tồn', 'Tồn D-15', 
                'Last GR', 'Last Sale', 'Số ngày ko nhập', 'Date tham khảo', 
                'DIO (D-15)', 'DIO/Date', 'Note', 'Note check slow', 'note hết hạn', 
                'Phân loại kiểm tra'
            ]
            
        cells_out2 = []
        for val in headers:
            cell = WriteOnlyCell(ws_out, value=val)
            cell.font = font_white_bold
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin
            cells_out2.append(cell)
        ws_out.append(cells_out2)
        
        # 2. Ghi các dòng dữ liệu
        dashboard_rows = []
        for row_list, phan_loai, sheet_type in written_rows:
            cells_temp = []
            for col_idx in range(27):
                val = row_list[col_idx]
                if val is None or val == '':
                    cells_temp.append(None)
                    continue

                if col_idx in (17, 18):
                    dt = excel_date_to_datetime(val) if not (val == '>90') else None
                    if dt:
                        cell = WriteOnlyCell(ws_out, value=dt)
                        cell.number_format = 'dd/mm/yyyy'
                        cells_temp.append(cell)
                    else:
                        cells_temp.append(val)
                elif col_idx in (11, 12, 13, 14, 15, 16, 19, 20):
                    if col_idx == 19 and str(val).strip() == '>90':
                        cells_temp.append(val)
                    else:
                        try:
                            num_val = int(float(val))
                            cell = WriteOnlyCell(ws_out, value=num_val)
                            cell.number_format = '#,##0'
                            cells_temp.append(cell)
                        except (ValueError, TypeError):
                            cells_temp.append(val)
                elif col_idx in (21, 22):
                    try:
                        num_val = float(val)
                        cell = WriteOnlyCell(ws_out, value=num_val)
                        cell.number_format = '0.00'
                        cells_temp.append(cell)
                    except (ValueError, TypeError):
                        cells_temp.append(val)
                elif col_idx == 26:
                    fill_af, font_af = phan_loai_style.get(phan_loai, (None, None))
                    cell = WriteOnlyCell(ws_out, value=val)
                    if font_af:
                        cell.font = font_af
                    if fill_af:
                        cell.fill = fill_af
                    cells_temp.append(cell)
                else:
                    cells_temp.append(val)

            # Lọc cột ASM nếu không use_asm
            if use_asm:
                cells_out = cells_temp
                final_row_list = row_list
            else:
                cells_out = [cells_temp[0]] + cells_temp[2:]
                final_row_list = [row_list[0]] + row_list[2:]

            ws_out.append(cells_out)
            rows_written += 1
            dashboard_rows.append(final_row_list)

        wb_out.save(out_path)
        # Sinh file HTML Dashboard
        if rows_written > 0:
            generate_html_dashboard(out_path, dashboard_rows, product_map, log_func, milestone_date)

        elapsed = time.time() - t0
        if log_func:
            log_func(f"  ✓ Hoàn thành: {rows_written:,} hàng xuất, {rows_filtered:,} hàng lọc bỏ")
            log_func(f"  → Thời gian: {elapsed:.1f}s | File: {out_filename}")
        return out_path

    except Exception as e:
        if log_func:
            if isinstance(e, PermissionError):
                log_func(f"  ✗ [LỖI KHÓA FILE] Không thể ghi đè lên file kết quả '{out_filename}'. Có thể file này đang mở trong Excel. Vui lòng đóng file và chạy lại!", 'error')
            else:
                log_func(f"  ✗ Lỗi xử lý {filename}: {e}")
                import traceback
                log_func(traceback.format_exc())
        return None
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass


def generate_html_dashboard(out_path: str, rows_data: list, product_map: dict, log_func=None, target_date=None) -> bool:
    """
    Tạo báo cáo Dashboard HTML từ dữ liệu đã lọc và ghi.
    """
    import json
    try:
        t0 = time.time()
        # Xác định đường dẫn file template
        current_dir = os.path.dirname(out_path)
        template_path = os.path.join(current_dir, "dashboard_chi_tiet_template.html")
        if not os.path.exists(template_path):
            script_dir = os.path.dirname(os.path.abspath(__file__))
            template_path = os.path.join(script_dir, "dashboard_chi_tiet_template.html")
            
        if not os.path.exists(template_path):
            template_path = r"D:\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu\dashboard_chi_tiet_template.html"

        if not os.path.exists(template_path):
            if log_func:
                log_func("  ✗ Không tìm thấy file template 'dashboard_chi_tiet_template.html'. Bỏ qua bước sinh Dashboard!")
            return False

        # Xác định vị trí các cột động dựa trên độ dài của dòng đầu tiên
        sample_len = len(rows_data[0]) if rows_data else 27
        is_excel = sample_len in (30, 31)
        has_asm = sample_len in (27, 31)
        
        idx_rsm = 0
        if has_asm:
            idx_asm = 1
            idx_store_code = 2
            idx_store_name = 3
            idx_mch2_desc = 5
            idx_mch3_desc = 7
            idx_mch4_id = 8
            idx_art_code = 9
            idx_art_name = 10
            
            if is_excel:
                idx_qty = 15
                idx_val = 16
                idx_class = 30
            else:
                idx_qty = 14
                idx_val = 15
                idx_class = 26
        else:
            idx_asm = -1
            idx_store_code = 1
            idx_store_name = 2
            idx_mch2_desc = 4
            idx_mch3_desc = 6
            idx_mch4_id = 7
            idx_art_code = 8
            idx_art_name = 9
            
            if is_excel:
                idx_qty = 14
                idx_val = 15
                idx_class = 29
            else:
                idx_qty = 13
                idx_val = 14
                idx_class = 25

        # Thu thập dữ liệu duy nhất để tạo index mapping
        rsms = set()
        asms = set()
        stores = {} # store_code: store_name
        mch2s = set()
        mch3s = set()
        mch4s = set()
        mch5s = set()
        classes = set()
        articles = {} # (art_code, art_name): idx
        
        for r in rows_data:
            rsm = r[idx_rsm] or ''
            asm = r[idx_asm] or '' if has_asm else ''
            store_code = r[idx_store_code] or ''
            store_name = r[idx_store_name] or ''
            mch2_desc = r[idx_mch2_desc] or ''
            mch3_desc = r[idx_mch3_desc] or ''
            mch4_id = r[idx_mch4_id] or ''
            art_code = r[idx_art_code] or ''
            art_name = r[idx_art_name] or ''
            check_class = r[idx_class] or ''
            
            rsms.add(rsm)
            if has_asm and asm:
                asms.add(asm)
            stores[store_code] = store_name
            mch2s.add(mch2_desc)
            mch3s.add(mch3_desc)
            mch4s.add(mch4_id)
            classes.add(check_class)
            
            prod_info = product_map.get(art_code)
            mch5 = 'Khác'
            if prod_info and len(prod_info) > 7:
                mch5 = prod_info[7] or 'Khác'
            mch5s.add(mch5)
            
            art_key = (art_code, art_name)
            if art_key not in articles:
                articles[art_key] = len(articles)
                
        rsm_list = sorted(list(rsms))
        asm_list = sorted(list(asms)) if has_asm else []
        store_list = [[code, name] for code, name in sorted(stores.items())]
        mch2_list = sorted(list(mch2s))
        mch3_list = sorted(list(mch3s))
        mch4_list = sorted(list(mch4s))
        mch5_list = sorted(list(mch5s))
        class_list = sorted(list(classes))
        
        article_list = [None] * len(articles)
        for (code, name), idx in articles.items():
            article_list[idx] = [code, name]
            
        rsm_map = {name: idx for idx, name in enumerate(rsm_list)}
        asm_map = {name: idx for idx, name in enumerate(asm_list)}
        store_map = {item[0]: idx for idx, item in enumerate(store_list)}
        mch2_map = {name: idx for idx, name in enumerate(mch2_list)}
        mch3_map = {name: idx for idx, name in enumerate(mch3_list)}
        mch4_map = {name: idx for idx, name in enumerate(mch4_list)}
        mch5_map = {name: idx for idx, name in enumerate(mch5_list)}
        class_map = {name: idx for idx, name in enumerate(class_list)}
        
        compressed_rows = []
        for r in rows_data:
            rsm = r[idx_rsm] or ''
            asm = r[idx_asm] or '' if has_asm else ''
            store_code = r[idx_store_code] or ''
            mch2_desc = r[idx_mch2_desc] or ''
            mch3_desc = r[idx_mch3_desc] or ''
            mch4_id = r[idx_mch4_id] or ''
            art_code = r[idx_art_code] or ''
            art_name = r[idx_art_name] or ''
            qty_val = safe_float(r[idx_qty]) or 0.0
            val_val = safe_float(r[idx_val]) or 0.0
            check_class = r[idx_class] or ''
            
            prod_info = product_map.get(art_code)
            mch5 = 'Khác'
            if prod_info and len(prod_info) > 7:
                mch5 = prod_info[7] or 'Khác'
                
            art_key = (art_code, art_name)
            
            comp_row = [
                store_map[store_code],
                rsm_map[rsm],
                asm_map.get(asm, -1),
                mch2_map[mch2_desc],
                mch3_map[mch3_desc],
                mch4_map[mch4_id],
                mch5_map[mch5],
                articles[art_key],
                class_map[check_class],
                round(qty_val),
                int(val_val)
            ]
            compressed_rows.append(comp_row)
            
        has_wm = any(r[1] == r[3] for r in rows_data if len(r) >= 4)
        out_dict = {
            "rsm_list": rsm_list,
            "asm_list": asm_list,
            "store_list": store_list,
            "mch2_list": mch2_list,
            "mch3_list": mch3_list,
            "mch4_list": mch4_list,
            "mch5_list": mch5_list,
            "class_list": class_list,
            "article_list": article_list,
            "has_wm": has_wm,
            "rows": compressed_rows
        }
        
        # Đọc template HTML
        with open(template_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
            
        filename = os.path.basename(out_path)
        title = "WIN"
        if "rural" in filename.lower():
            title = "Rural"
        elif "urban" in filename.lower():
            title = "Urban"
            
        html_content = html_content.replace(
            "<title>Báo Cáo Đối Soát Dữ Liệu Kiểm Tra - Chuỗi WIN</title>",
            f"<title>Báo Cáo Đối Soát Dữ Liệu Kiểm Tra - Chuỗi {title}</title>"
        )
        html_content = html_content.replace(
            '<span class="px-2.5 py-0.5 bg-indigo-500/10 text-indigo-400 border border-indigo-500/20 text-xs font-bold rounded-full">Chuỗi WIN</span>',
            f'<span class="px-2.5 py-0.5 bg-indigo-500/10 text-indigo-400 border border-indigo-500/20 text-xs font-bold rounded-full">Chuỗi {title}</span>'
        )
        
        if target_date:
            if isinstance(target_date, str):
                today_str = target_date
            else:
                today_str = target_date.strftime("%d/%m/%Y")
        else:
            today_str = datetime.now().strftime("%d/%m/%Y")
        html_content = html_content.replace(
            "Cập nhật: 01/07/2026",
            f"Cập nhật: {today_str}"
        )
        
        js_injection = f"const dbData = {json.dumps(out_dict, ensure_ascii=False, separators=(',', ':'))};\n"
        js_injection += "const rawData = dbData.rows.map(r => ({\n"
        js_injection += "    store_idx: r[0],\n"
        js_injection += "    store_code: dbData.store_list[r[0]][0],\n"
        js_injection += "    store_name: dbData.store_list[r[0]][1],\n"
        js_injection += "    rsm: dbData.rsm_list[r[1]],\n"
        js_injection += "    rsm_idx: r[1],\n"
        js_injection += "    asm: r[2] !== -1 ? dbData.asm_list[r[2]] : '',\n"
        js_injection += "    asm_idx: r[2],\n"
        js_injection += "    mch2: dbData.mch2_list[r[3]],\n"
        js_injection += "    mch2_idx: r[3],\n"
        js_injection += "    mch3: dbData.mch3_list[r[4]],\n"
        js_injection += "    mch3_idx: r[4],\n"
        js_injection += "    mch4: dbData.mch4_list[r[5]],\n"
        js_injection += "    mch4_idx: r[5],\n"
        js_injection += "    mch5: dbData.mch5_list[r[6]],\n"
        js_injection += "    mch5_idx: r[6],\n"
        js_injection += "    article_code: dbData.article_list[r[7]][0],\n"
        js_injection += "    article_name: dbData.article_list[r[7]][1],\n"
        js_injection += "    check_class: dbData.class_list[r[8]],\n"
        js_injection += "    class_idx: r[8],\n"
        js_injection += "    qty: r[9],\n"
        js_injection += "    value: r[10]\n"
        js_injection += "}));"
        
        html_content = html_content.replace('// DATA_PLACEHOLDER', js_injection)
        
        out_html_path = out_path.replace(".xlsx", "_dashboard.html")
        with open(out_html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
            
        elapsed = time.time() - t0
        if log_func:
            log_func(f"  ✓ Đã sinh Dashboard: {os.path.basename(out_html_path)} | Thời gian: {elapsed:.1f}s")
        return True
    except Exception as e:
        if log_func:
            log_func(f"  ✗ Lỗi sinh Dashboard: {e}")
        return False


def process_excel_file(src_path: str, store_map: dict, product_map: dict, milestone_date: datetime,
                       ref_amount: float, log_func=None) -> str | None:

    """
    Xử lý file dữ liệu Excel bằng iterparse (one-pass) để tối ưu hóa tốc độ.
    Lọc và giữ lại các hàng có Phân loại kiểm tra.
    Trả về đường dẫn file output hoặc None nếu lỗi.
    """
    filename = os.path.basename(src_path)
    base, ext = os.path.splitext(filename)
    if ext.lower() == '.csv':
        return process_csv_file(src_path, store_map, product_map, milestone_date, ref_amount, log_func)

    shelf_life_map = {k: v[6] for k, v in product_map.items()}

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.cell import WriteOnlyCell

    out_filename = f"{base} (done){ext}"
    out_path = os.path.join(os.path.dirname(src_path), out_filename)

    if log_func:
        log_func(f"\n[{filename}] Bắt đầu xử lý...")

    t0 = time.time()
    temp_path = os.path.join(SCRATCH_DIR, f"temp_{filename}")


    try:
        # Sao chép file để tránh lock
        res = ctypes.windll.kernel32.CopyFileW(src_path, temp_path, False)
        if not res:
            if log_func:
                log_func(f"  ✗ Không thể copy file (đang mở?): {filename}")
            return None

        # Định nghĩa styles
        font_white_bold = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        font_normal = Font(name='Arial', size=9)
        fill_header = PatternFill(fill_type='solid', start_color='1F4E78', end_color='1F4E78')
        
        fill_het_han = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
        fill_nghi_van = PatternFill(fill_type='solid', start_color='FF6600', end_color='FF6600')
        fill_non_moving = PatternFill(fill_type='solid', start_color='FFA500', end_color='FFA500')
        fill_slow_moving = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
        font_slow = Font(name='Arial', size=9, color='000000')
        
        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')

        # Màu cho từng phân loại
        phan_loai_style = {
            'Hết hạn': (fill_het_han, Font(name='Arial', size=9, bold=True, color='FFFFFF')),
            'Nghi vấn tồn ảo': (fill_nghi_van, Font(name='Arial', size=9, bold=True, color='FFFFFF')),
            'Non-moving': (fill_non_moving, Font(name='Arial', size=9, bold=True, color='000000')),
            'Slow moving': (fill_slow_moving, font_slow),
        }

        wb_out = openpyxl.Workbook(write_only=True)
        ws_out = wb_out.create_sheet(title='Dữ liệu kiểm tra')
        written_rows = []

        rows_written = 0
        rows_filtered = 0
        total_rows_read = 0

        with zipfile.ZipFile(temp_path, 'r') as z:
            # Tìm sheet dữ liệu chính (bỏ qua hidden và RC)
            wb_data = z.read('xl/workbook.xml')
            root_wb = ET.fromstring(wb_data)
            ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

            sheets_info = []
            for sheet_elem in root_wb.findall('.//ns:sheet', ns):
                sheets_info.append((
                    sheet_elem.attrib.get('name'),
                    sheet_elem.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                ))

            data_sheets = [s for s in sheets_info
                           if s[0] not in ['_com.sap.ip.bi.xl.hiddensheet', 'RC']]
            if not data_sheets:
                if log_func:
                    log_func(f"  ✗ Không tìm thấy sheet dữ liệu trong: {filename}")
                return None

            _, target_rid = data_sheets[0]

            # Tìm đường dẫn file worksheet
            rels_data = z.read('xl/_rels/workbook.xml.rels')
            root_rels = ET.fromstring(rels_data)
            r_ns = {'rels': 'http://schemas.openxmlformats.org/package/2006/relationships'}
            sheet_path_in_zip = None
            for rel in root_rels.findall('.//rels:Relationship', r_ns):
                if rel.attrib.get('Id') == target_rid:
                    sheet_path_in_zip = f"xl/{rel.attrib.get('Target')}"
                    break

            if not sheet_path_in_zip:
                if log_func:
                    log_func(f"  ✗ Không tìm thấy path worksheet: {filename}")
                return None

            # Load shared strings
            shared_strings = []
            if 'xl/sharedStrings.xml' in z.namelist():
                sst_file = z.open('xl/sharedStrings.xml')
                context_sst = ET.iterparse(sst_file, events=('start', 'end'))
                for event, elem in context_sst:
                    if event == 'end' and elem.tag.endswith('}si'):
                        text = "".join([
                            t.text for t in elem.findall(
                                './/{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')
                            if t.text is not None
                        ])
                        shared_strings.append(text)
                        elem.clear()

            # Tối ưu hóa hàm convert số thực
            def get_cell_value(cell_elem):
                t_type = cell_elem.attrib.get('t')
                v_elem = cell_elem.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                if v_elem is None:
                    return None
                val_text = v_elem.text
                if val_text is None:
                    return None
                if t_type == 's':
                    idx = int(val_text)
                    return shared_strings[idx] if idx < len(shared_strings) else ''
                try:
                    if '.' in val_text:
                        return float(val_text)
                    return int(val_text)
                except ValueError:
                    return val_text

            # Stream parse worksheet XML
            sheet_file = z.open(sheet_path_in_zip)
            context = ET.iterparse(sheet_file, events=('start', 'end'))

            current_row_data = {}
            current_row_idx = None
            offset = 1  # Mặc định lệch 1 cột (có cột map)

            for event, elem in context:
                if event == 'start' and elem.tag.endswith('}row'):
                    current_row_data = {}
                    r_attr = elem.attrib.get('r')
                    current_row_idx = int(r_attr) if r_attr and r_attr.isdigit() else None

                elif event == 'end' and elem.tag.endswith('}c'):
                    r_ref = elem.attrib.get('r')
                    if r_ref:
                        col_end = 0
                        for ch in r_ref:
                            if ch.isdigit():
                                break
                            col_end += 1
                        current_row_data[r_ref[:col_end]] = get_cell_value(elem)

                elif event == 'end' and elem.tag.endswith('}row'):
                    if current_row_idx is None:
                        elem.clear()
                        continue

                    total_rows_read += 1

                    if current_row_idx == 1:
                        # Bỏ qua Row 1 (Nhãn tham chiếu) hoàn toàn
                        elem.clear()
                        continue

                    elif current_row_idx == 2:
                        # Row 2: Header + Xác định offset
                        offset = 1
                        for col_l, val in current_row_data.items():
                            if str(val).strip() == 'RSM':
                                offset = col_to_idx(col_l)
                                break
                        if log_func:
                            log_func(f"  → Dò tìm cấu trúc cột: offset = {offset} (cột {'B' if offset == 1 else 'A'})")

                    else:
                        # Row 3+: Dữ liệu
                        actual_d = get_actual_col('D', offset)
                        actual_e = get_actual_col('E', offset)
                        store_code = str(current_row_data.get(actual_d, '')).strip()
                        store_name = str(current_row_data.get(actual_e, '')).strip()

                        if 'result' in store_code.lower() or 'result' in store_name.lower():
                            rows_filtered += 1
                            elem.clear()
                            continue

                        # Chuẩn hóa (nạp thêm Z cũ từ dữ liệu gốc)
                        norm_row = {}
                        cols_to_norm = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                                        'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Z']
                        for col_std in cols_to_norm:
                            col_act = get_actual_col(col_std, offset)
                            norm_row[col_std] = current_row_data.get(col_act)

                        # Bỏ qua các sản phẩm thuộc MCH2 là "101" hoặc thiếu thông tin MCH2/MCH3
                        mch2_id = str(norm_row.get('F') or '').strip()
                        mch3_id = str(norm_row.get('H') or '').strip()
                        if mch2_id == '101' or not mch2_id or not mch3_id or mch2_id == '' or mch3_id == '':
                            rows_filtered += 1
                            elem.clear()
                            continue

                        computed = compute_row(norm_row, milestone_date, ref_amount, shelf_life_map)
                        phan_loai = computed.get('AF')

                        if phan_loai is None or str(phan_loai).strip() == '':
                            rows_filtered += 1
                            elem.clear()
                            continue

                        # Build row_list
                        row_list = [None] * 31
                        # Sao chép các cột từ B đến V
                        for col_l in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                                      'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V']:
                            row_list[col_to_idx(col_l) - 1] = norm_row.get(col_l)

                        # Tra cứu thông tin cửa hàng để sửa lỗi tiếng Việt và lấy phân loại sheet
                        store_info = store_map.get(store_code)
                        if store_info:
                            rsm_name, asm_name, store_name, sheet_type = store_info[:4]
                        else:
                            store_name = store_name
                            if store_code.isdigit():
                                rsm_name = row_list[0] or ''
                                asm_name = store_name
                                sheet_type = 'WM'
                            else:
                                rsm_name = row_list[0] or ''
                                asm_name = row_list[1] or ''
                                sheet_type = 'WM+'

                        row_list[0] = rsm_name
                        row_list[1] = asm_name
                        row_list[3] = store_name

                        for col_l, key in [('W', 'W'), ('X', 'X'), ('Y', 'Y'), ('Z', 'Z'), ('AA', 'AA'),
                                            ('AB', 'AB'), ('AC', 'AC'), ('AD', 'AD'),
                                            ('AE', 'AE'), ('AF', 'AF')]:
                            row_list[col_to_idx(col_l) - 1] = computed.get(key)

                        written_rows.append((row_list, phan_loai, sheet_type))


        # Sau khi lặp xong và collect all rows in written_rows:
        has_wm = any(item[2] == 'WM' for item in written_rows)
        has_wm_plus = any(item[2] == 'WM+' for item in written_rows)
        use_asm = True
        
        # 1. Ghi Row 1: Headers (Đã loại bỏ Row 1 cũ là Nhãn tham chiếu)
        if use_asm:
            headers = {
                'B': 'RSM' if not has_wm else 'RSM / GĐV/GĐM', 'C': 'ASM', 'D': 'Store Code', 'E': 'Store Name',
                'F': 'MCH2 ID', 'G': 'MCH2 Desc', 'H': 'MCH3 ID', 'I': 'MCH3 Desc',
                'J': 'MCH4 ID', 'K': 'Article Code', 'L': 'Article Name',
                'M': 'Created on', 'N': 'Tồn đầu kỳ (D-90)',
                'O': 'Nhập trong kỳ', 'P': 'Xuất trong kỳ',
                'Q': 'Tồn cuối kỳ (D)', 'R': 'Giá trị tồn',
                'S': 'Doanh thu', 'T': 'Giá vốn', 'U': 'Tồn D-15',
                'V': 'Giá tồn D-15', 'W': 'Last GR', 'X': 'Last Sale',
                'Y': 'Số ngày ko nhập', 'Z': 'Date tham khảo',
                'AA': 'DIO (D-15)', 'AB': 'DIO/Date', 'AC': 'Note',
                'AD': 'Note check slow', 'AE': 'note hết hạn',
                'AF': 'Phân loại kiểm tra'
            }
        else:
            headers = {
                'B': 'GĐV/GĐM', 'C': 'Store Code', 'D': 'Store Name',
                'E': 'MCH2 ID', 'F': 'MCH2 Desc', 'G': 'MCH3 ID', 'H': 'MCH3 Desc',
                'I': 'MCH4 ID', 'J': 'Article Code', 'K': 'Article Name',
                'L': 'Created on', 'M': 'Tồn đầu kỳ (D-90)',
                'N': 'Nhập trong kỳ', 'O': 'Xuất trong kỳ',
                'P': 'Tồn cuối kỳ (D)', 'Q': 'Giá trị tồn',
                'R': 'Doanh thu', 'S': 'Giá vốn', 'T': 'Tồn D-15',
                'U': 'Giá tồn D-15', 'V': 'Last GR', 'W': 'Last Sale',
                'X': 'Số ngày ko nhập', 'Y': 'Date tham khảo',
                'Z': 'DIO (D-15)', 'AA': 'DIO/Date', 'AB': 'Note',
                'AC': 'Note check slow', 'AD': 'note hết hạn',
                'AE': 'Phân loại kiểm tra'
            }
            
        cells_out2 = []
        for col_l, label in sorted(headers.items(), key=lambda x: col_to_idx(x[0])):
            cell = WriteOnlyCell(ws_out, value=label)
            cell.font = font_white_bold
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin
            cells_out2.append(cell)
        ws_out.append(cells_out2)
        
        # 2. Ghi các dòng dữ liệu
        dashboard_rows = []
        for row_list, phan_loai, sheet_type in written_rows:
            cells_temp = []
            for col_out_idx in range(31):
                col_letter = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
                              'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                              'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB',
                              'AC', 'AD', 'AE', 'AF'][col_out_idx]
                val = row_list[col_out_idx]

                if val is None or val == '':
                    cells_temp.append(None)
                    continue

                if col_letter in ('M', 'W', 'X'):
                    dt = excel_date_to_datetime(val) if not (col_letter in ('W', 'X') and str(val).strip() == '>90') else None
                    if dt:
                        cell = WriteOnlyCell(ws_out, value=dt)
                        cell.number_format = 'dd/mm/yyyy'
                        cell.font = font_normal
                        cell.alignment = align_center
                    else:
                        cell = WriteOnlyCell(ws_out, value=val)
                        cell.font = font_normal
                        cell.alignment = align_center
                    cells_temp.append(cell)

                elif col_letter in ('N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'Y', 'Z'):
                    try:
                        if col_letter == 'Y' and str(val).strip() == '>90':
                            cell = WriteOnlyCell(ws_out, value=val)
                            cell.font = font_normal
                            cell.alignment = align_right
                        else:
                            num_val = int(float(val))
                            cell = WriteOnlyCell(ws_out, value=num_val)
                            cell.number_format = '#,##0'
                            cell.font = font_normal
                            cell.alignment = align_right
                    except (ValueError, TypeError):
                        cell = WriteOnlyCell(ws_out, value=val)
                        cell.font = font_normal
                        cell.alignment = align_right
                    cells_temp.append(cell)

                elif col_letter in ('AA', 'AB'):
                    try:
                        num_val = float(val)
                        cell = WriteOnlyCell(ws_out, value=num_val)
                        cell.number_format = '0.00'
                        cell.font = font_normal
                        cell.alignment = align_right
                    except (ValueError, TypeError):
                        cell = WriteOnlyCell(ws_out, value=val)
                        cell.font = font_normal
                        cell.alignment = align_right
                    cells_temp.append(cell)

                elif col_letter == 'AF':
                    fill_af, font_af = phan_loai_style.get(phan_loai, (None, font_normal))
                    cell = WriteOnlyCell(ws_out, value=val)
                    cell.font = font_af
                    if fill_af:
                        cell.fill = fill_af
                    cell.alignment = align_center
                    cells_temp.append(cell)

                else:
                    cells_temp.append(val)

            # Lọc cột ASM nếu không use_asm
            if use_asm:
                cells_out = cells_temp
                final_row_list = row_list
            else:
                cells_out = [cells_temp[0]] + cells_temp[2:]
                final_row_list = [row_list[0]] + row_list[2:]

            ws_out.append(cells_out)
            rows_written += 1
            dashboard_rows.append(final_row_list)

        wb_out.save(out_path)
        # Sinh file HTML Dashboard
        if rows_written > 0:
            generate_html_dashboard(out_path, dashboard_rows, product_map, log_func, milestone_date)
        elapsed = time.time() - t0
        if log_func:
            log_func(f"  ✓ Hoàn thành: {rows_written:,} hàng xuất, {rows_filtered:,} hàng lọc bỏ")
            log_func(f"  → Thời gian: {elapsed:.1f}s | File: {out_filename}")
        return out_path

    except Exception as e:
        if log_func:
            log_func(f"  ✗ Lỗi xử lý {filename}: {e}")
            import traceback
            log_func(traceback.format_exc())
        return None
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass


# ===================================================================
# GIAO DIỆN NGƯỜI DÙNG (GUI)
# ===================================================================
class App(tk.Tk):
    """
    Giao diện chính của Tool Đổ Dữ Liệu.
    Các chức năng:
      - Chọn nhiều file Excel đầu vào
      - Chọn file XLSB tra cứu
      - Nhập ngày mốc và số tiền tham chiếu
      - Chạy xử lý và hiển thị log
    """

    # Màu sắc giao diện (dark theme)
    BG_DARK = '#1a1d2e'
    BG_CARD = '#252840'
    BG_INPUT = '#1e2235'
    ACCENT = '#4f8ef7'
    ACCENT_HOVER = '#6ba3ff'
    SUCCESS = '#4caf50'
    WARNING = '#ff9800'
    ERROR = '#f44336'
    TEXT = '#e8eaed'
    TEXT_DIM = '#9aa0b8'
    BORDER = '#353a5e'

    def __init__(self):
        super().__init__()
        self.title(f"{APP_TITLE} v{APP_VERSION}")
        self.geometry("980x720")
        self.minsize(800, 600)
        self.configure(bg=self.BG_DARK)

        # Biến trạng thái
        self.selected_files: list[str] = []
        self.master_path = tk.StringVar(value=DEFAULT_MASTER_PATH)
        self.store_path = tk.StringVar(value=DEFAULT_STORE_PATH)
        self.milestone_date_str = tk.StringVar(value=datetime.now().strftime('%d/%m/%Y'))
        self.ref_amount_str = tk.StringVar(value='100,000')
        self.is_running = False


        self._build_ui()
        self._center_window()

    def _center_window(self):
        self.update_idletasks()
        w = self.winfo_width()
        h = self.winfo_height()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _build_ui(self):
        """Xây dựng toàn bộ giao diện."""
        # === Header ===
        header = tk.Frame(self, bg='#131629', pady=12)
        header.pack(fill='x')

        tk.Label(
            header, text="📊  Tool Đổ Dữ Liệu Phân Tích Tồn Kho",
            font=('Segoe UI', 16, 'bold'), bg='#131629', fg=self.TEXT
        ).pack(side='left', padx=20)

        tk.Label(
            header, text=f"v{APP_VERSION}", font=('Segoe UI', 9),
            bg='#131629', fg=self.TEXT_DIM
        ).pack(side='right', padx=20)

        # === Footer: nút chạy ===
        footer = tk.Frame(self, bg='#131629', pady=10)
        footer.pack(fill='x', side='bottom')

        self.run_btn = tk.Button(
            footer,
            text="▶  BẮT ĐẦU XỬ LÝ",
            font=('Segoe UI', 12, 'bold'),
            bg=self.ACCENT, fg='white',
            activebackground=self.ACCENT_HOVER,
            activeforeground='white',
            cursor='hand2',
            relief='flat',
            padx=30, pady=8,
            command=self._on_run
        )
        self.run_btn.pack(side='left', padx=20)

        self.status_lbl = tk.Label(
            footer, text="Sẵn sàng.",
            font=('Segoe UI', 10), bg='#131629', fg=self.TEXT_DIM
        )
        self.status_lbl.pack(side='left', padx=10)

        # === Main content ===
        content = tk.Frame(self, bg=self.BG_DARK)
        content.pack(fill='both', expand=True, padx=15, pady=10)

        # Left panel: cấu hình
        left = tk.Frame(content, bg=self.BG_DARK)
        left.pack(side='left', fill='y', padx=(0, 10))

        # Right panel: log
        right = tk.Frame(content, bg=self.BG_DARK)
        right.pack(side='right', fill='both', expand=True)

        self._build_config_panel(left)
        self._build_log_panel(right)

    def _card(self, parent, title: str) -> tk.Frame:
        """Tạo card với tiêu đề."""
        outer = tk.Frame(parent, bg=self.BG_CARD, pady=0)
        outer.pack(fill='x', pady=5)
        tk.Label(
            outer, text=title, font=('Segoe UI', 10, 'bold'),
            bg=self.BG_CARD, fg=self.ACCENT, anchor='w', padx=12, pady=6
        ).pack(fill='x')
        tk.Frame(outer, bg=self.BORDER, height=1).pack(fill='x', padx=12)
        inner = tk.Frame(outer, bg=self.BG_CARD, padx=12, pady=10)
        inner.pack(fill='x')
        return inner

    def _build_config_panel(self, parent):
        """Panel bên trái: cấu hình tham số."""
        parent.configure(width=360)

        # --- Card 1: File đầu vào ---
        c1 = self._card(parent, "1. File báo cáo đầu vào (.xlsx, .csv)")

        self.file_listbox = tk.Listbox(
            c1, font=('Segoe UI', 9), bg=self.BG_INPUT, fg=self.TEXT,
            selectbackground=self.ACCENT, selectforeground='white',
            height=6, relief='flat', bd=0,
            highlightthickness=1, highlightcolor=self.BORDER,
            highlightbackground=self.BORDER
        )
        self.file_listbox.pack(fill='x', pady=(0, 6))

        btn_frame = tk.Frame(c1, bg=self.BG_CARD)
        btn_frame.pack(fill='x')
        self._btn(btn_frame, "📂 Thêm file", self._add_files).pack(side='left')
        self._btn(btn_frame, "✖ Xóa chọn", self._remove_selected, color='#e53935').pack(side='left', padx=4)
        self._btn(btn_frame, "🗑 Xóa tất cả", self._clear_files, color='#555').pack(side='left')

        self.file_count_lbl = tk.Label(
            c1, text="Chưa chọn file nào.",
            font=('Segoe UI', 8), bg=self.BG_CARD, fg=self.TEXT_DIM, anchor='w'
        )
        self.file_count_lbl.pack(fill='x', pady=(4, 0))

        # --- Card 2: File tra cứu (Master & Store List) ---
        c2 = self._card(parent, "2. File tra cứu danh mục & siêu thị")

        # Hàng 1: Master Article
        tk.Label(c2, text="Danh mục sản phẩm (Master Article):", font=('Segoe UI', 8), bg=self.BG_CARD, fg=self.TEXT_DIM, anchor='w').pack(fill='x', pady=(0, 2))
        master_frame = tk.Frame(c2, bg=self.BG_CARD)
        master_frame.pack(fill='x', pady=(0, 6))

        self.master_entry = tk.Entry(
            master_frame, textvariable=self.master_path,
            font=('Segoe UI', 8), bg=self.BG_INPUT, fg=self.TEXT,
            relief='flat', bd=0, insertbackground=self.TEXT,
            highlightthickness=1, highlightcolor=self.BORDER,
            highlightbackground=self.BORDER
        )
        self.master_entry.pack(side='left', fill='x', expand=True, pady=2)
        self._btn(master_frame, "📂", self._browse_master).pack(side='right', padx=(4, 0))

        # Hàng 2: Store List
        tk.Label(c2, text="Danh sách siêu thị (Store List):", font=('Segoe UI', 8), bg=self.BG_CARD, fg=self.TEXT_DIM, anchor='w').pack(fill='x', pady=(0, 2))
        store_frame = tk.Frame(c2, bg=self.BG_CARD)
        store_frame.pack(fill='x')

        self.store_entry = tk.Entry(
            store_frame, textvariable=self.store_path,
            font=('Segoe UI', 8), bg=self.BG_INPUT, fg=self.TEXT,
            relief='flat', bd=0, insertbackground=self.TEXT,
            highlightthickness=1, highlightcolor=self.BORDER,
            highlightbackground=self.BORDER
        )
        self.store_entry.pack(side='left', fill='x', expand=True, pady=2)
        self._btn(store_frame, "📂", self._browse_store).pack(side='right', padx=(4, 0))


        # --- Card 3: Tham số ---
        c3 = self._card(parent, "3. Tham số tính toán")

        self._field(c3, "Ngày mốc (Y1):", self.milestone_date_str,
                    tip="Định dạng: DD/MM/YYYY")
        self.ref_amount_entry = self._field(
            c3, "Số tiền tham chiếu (AF1):", self.ref_amount_str,
            tip="Đơn vị: VNĐ (ví dụ: 100,000)"
        )
        self.ref_amount_entry.bind("<KeyRelease>", self._format_ref_amount)
        self.ref_amount_entry.bind("<FocusOut>", self._format_ref_amount)

        # --- Card 4: Ghi chú ---
        c4 = self._card(parent, "📋 Ghi chú phân loại")
        notes = [
            ("🔴", "Hết hạn", "Last GR + Date tham khảo < Ngày mốc"),
            ("🟠", "Nghi vấn tồn ảo", "Không nhập, không bán, có tồn"),
            ("🟡", "Non-moving", "Doanh thu = 0 (90 ngày)"),
            ("⚪", "Slow moving", "DIO/Date ≥ 5 và Giá trị tồn ≥ AF1"),
        ]
        for icon, label, desc in notes:
            row = tk.Frame(c4, bg=self.BG_CARD)
            row.pack(fill='x', pady=1)
            tk.Label(row, text=icon, bg=self.BG_CARD, font=('Segoe UI', 10)).pack(side='left')
            tk.Label(row, text=f" {label}:", font=('Segoe UI', 8, 'bold'),
                     bg=self.BG_CARD, fg=self.TEXT, width=18, anchor='w').pack(side='left')
            tk.Label(row, text=desc, font=('Segoe UI', 8),
                     bg=self.BG_CARD, fg=self.TEXT_DIM, anchor='w').pack(side='left')

    def _build_log_panel(self, parent):
        """Panel bên phải: log output."""
        tk.Label(
            parent, text="Nhật ký xử lý",
            font=('Segoe UI', 11, 'bold'), bg=self.BG_DARK, fg=self.TEXT, anchor='w'
        ).pack(fill='x', pady=(0, 4))

        log_frame = tk.Frame(parent, bg=self.BG_CARD, bd=0)
        log_frame.pack(fill='both', expand=True)

        self.log_text = tk.Text(
            log_frame, font=('Consolas', 9), bg=self.BG_INPUT, fg=self.TEXT,
            relief='flat', bd=0, insertbackground=self.TEXT,
            highlightthickness=0, state='disabled'
        )
        self.log_text.pack(side='left', fill='both', expand=True, padx=(1, 0), pady=1)

        scrollbar = ttk.Scrollbar(log_frame, orient='vertical', command=self.log_text.yview)
        scrollbar.pack(side='right', fill='y')
        self.log_text.configure(yscrollcommand=scrollbar.set)

        # Định nghĩa tags màu cho log
        self.log_text.tag_config('info', foreground=self.TEXT)
        self.log_text.tag_config('success', foreground=self.SUCCESS)
        self.log_text.tag_config('warning', foreground=self.WARNING)
        self.log_text.tag_config('error', foreground=self.ERROR)
        self.log_text.tag_config('highlight', foreground=self.ACCENT, font=('Consolas', 9, 'bold'))

    def _btn(self, parent, text: str, cmd, color=None) -> tk.Button:
        """Helper tạo button."""
        bg_color = color if color else self.ACCENT
        active_bg = self.ACCENT_HOVER if not color else color
        btn = tk.Button(
            parent, text=text, font=('Segoe UI', 8, 'bold'),
            bg=bg_color, fg='white', activebackground=active_bg,
            activeforeground='white', relief='flat', cursor='hand2',
            padx=10, pady=4, command=cmd
        )
        return btn

    def _field(self, parent, label: str, var: tk.StringVar, tip: str = ''):
        frame = tk.Frame(parent, bg=self.BG_CARD)
        frame.pack(fill='x', pady=3)
        tk.Label(
            frame, text=label, font=('Segoe UI', 9),
            bg=self.BG_CARD, fg=self.TEXT, anchor='w', width=25
        ).pack(side='left')
        entry = tk.Entry(
            frame, textvariable=var, font=('Segoe UI', 9),
            bg=self.BG_INPUT, fg=self.TEXT, relief='flat',
            insertbackground=self.TEXT,
            highlightthickness=1, highlightcolor=self.ACCENT,
            highlightbackground=self.BORDER
        )
        entry.pack(side='left', fill='x', expand=True)
        if tip:
            tk.Label(
                parent, text=tip, font=('Segoe UI', 7),
                bg=self.BG_CARD, fg=self.TEXT_DIM, anchor='e'
            ).pack(fill='x')
        return entry

    def _format_ref_amount(self, event=None):
        """Tự động thêm dấu phẩy phân tách hàng nghìn khi người dùng gõ phím."""
        try:
            cursor_pos = self.ref_amount_entry.index(tk.INSERT)
        except Exception:
            cursor_pos = 0

        raw = self.ref_amount_str.get()
        # Đếm số dấu phẩy trước con trỏ trước khi định dạng
        commas_before = raw[:cursor_pos].count(',')
        
        # Chỉ lấy chữ số
        clean = "".join([c for c in raw if c.isdigit()])
        if clean:
            try:
                formatted = f"{int(clean):,}"
                self.ref_amount_str.set(formatted)
                
                # Tính toán lại vị trí con trỏ để tránh bị nhảy về cuối ô nhập liệu
                digits_before = cursor_pos - commas_before
                new_pos = 0
                digit_count = 0
                for char in formatted:
                    if digit_count == digits_before:
                        break
                    if char.isdigit():
                        digit_count += 1
                    new_pos += 1
                
                self.ref_amount_entry.icursor(new_pos)
            except (ValueError, TypeError):
                pass
        else:
            self.ref_amount_str.set("")

    # --- Actions ---
    def _add_files(self):
        files = filedialog.askopenfilenames(
            title="Chọn file báo cáo tồn kho",
            filetypes=[("Excel/CSV Files", "*.xlsx;*.csv"), ("Excel Files", "*.xlsx"), ("CSV Files", "*.csv")]
        )
        for f in files:
            norm_f = os.path.abspath(f)
            if norm_f not in self.selected_files:
                self.selected_files.append(norm_f)
                self.file_listbox.insert(tk.END, os.path.basename(norm_f))
        self._update_file_count()

    def _remove_selected(self):
        selected_indices = list(self.file_listbox.curselection())
        for idx in reversed(selected_indices):
            self.selected_files.pop(idx)
            self.file_listbox.delete(idx)
        self._update_file_count()

    def _clear_files(self):
        self.selected_files.clear()
        self.file_listbox.delete(0, tk.END)
        self._update_file_count()

    def _update_file_count(self):
        count = len(self.selected_files)
        if count == 0:
            self.file_count_lbl.config(text="Chưa chọn file nào.", fg=self.TEXT_DIM)
        else:
            self.file_count_lbl.config(text=f"Đã chọn {count} file.", fg=self.SUCCESS)

    def _browse_master(self):
        f = filedialog.askopenfilename(
            title="Chọn file danh mục sản phẩm (Master Article)",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if f:
            self.master_path.set(os.path.abspath(f))

    def _browse_store(self):
        f = filedialog.askopenfilename(
            title="Chọn file danh sách siêu thị (Store List)",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if f:
            self.store_path.set(os.path.abspath(f))


    def log(self, message: str, tag: str = 'info'):
        """Ghi log vào khung text."""
        self.log_text.config(state='normal')
        self.log_text.insert(tk.END, message + "\n", tag)
        self.log_text.see(tk.END)
        self.log_text.config(state='disabled')
        self.update_idletasks()

    def _on_run(self):
        if self.is_running:
            return
        
        # Kiểm tra đầu vào
        if not self.selected_files:
            messagebox.showerror("Lỗi", "Vui lòng chọn ít nhất 1 file báo cáo đầu vào!")
            return

        master = self.master_path.get().strip()
        if not os.path.exists(master):
            messagebox.showerror("Lỗi", f"Không tìm thấy file danh mục sản phẩm tại:\n{master}")
            return

        store = self.store_path.get().strip()
        if not os.path.exists(store):
            messagebox.showerror("Lỗi", f"Không tìm thấy file danh sách siêu thị tại:\n{store}")
            return

        # Parse ngày mốc
        date_str = self.milestone_date_str.get().strip()
        try:
            milestone_date = datetime.strptime(date_str, '%d/%m/%Y')
        except ValueError:
            messagebox.showerror("Lỗi", "Ngày mốc không đúng định dạng DD/MM/YYYY!\nVí dụ: 01/07/2026")
            return

        # Parse số tiền tham chiếu (loại bỏ dấu phẩy/chấm phân tách hàng nghìn)
        ref_str = self.ref_amount_str.get().strip().replace(',', '').replace('.', '')
        try:
            ref_amount = float(ref_str)
        except ValueError:
            messagebox.showerror("Lỗi", "Số tiền tham chiếu phải là một số hợp lệ!")
            return

        # Vô hiệu hóa nút và bắt đầu chạy thread ngầm
        self.is_running = True
        self.run_btn.config(state='disabled', text="⏳  ĐANG XỬ LÝ...")
        self.status_lbl.config(text="Đang xử lý dữ liệu...", fg=self.WARNING)
        self.log_text.config(state='normal')
        self.log_text.delete('1.0', tk.END)
        self.log_text.config(state='disabled')

        def worker():
            t_start = time.time()
            self.log(f"=== BẮT ĐẦU TIẾN TRÌNH XỬ LÝ (v{APP_VERSION}) ===", 'highlight')
            self.log(f"Ngày mốc: {date_str} | Số tiền tham chiếu: {ref_amount:,.0f} VNĐ\n")
            
            # 1. Nạp danh mục sản phẩm và danh sách siêu thị
            store_map = load_store_details(store, lambda msg, t='info': self.log(msg, t))
            product_map = load_product_details(master, lambda msg, t='info': self.log(msg, t))
            
            # 2. Xử lý từng file
            success_count = 0
            for idx, file in enumerate(self.selected_files):
                filename = os.path.basename(file)
                self.log(f"\n[{idx+1}/{len(self.selected_files)}] Đang xử lý: {filename}...", 'highlight')
                
                try:
                    out = process_excel_file(
                        file, store_map, product_map, milestone_date, ref_amount,
                        log_func=lambda msg, t='info': self.log(msg, t)
                    )
                    if out:
                        success_count += 1
                except Exception as ex:
                    self.log(f"  ✗ Lỗi không xác định: {ex}", 'error')


            elapsed = time.time() - t_start
            self.log(f"\n=== TIẾN TRÌNH HOÀN TẤT ===", 'highlight')
            self.log(f"Đã xử lý thành công {success_count}/{len(self.selected_files)} file.", 'success')
            self.log(f"Tổng thời gian thực hiện: {elapsed:.1f} giây.\n", 'success')
            
            # Mở thư mục chứa file done đầu tiên nếu thành công
            if success_count > 0 and self.selected_files:
                first_dir = os.path.dirname(self.selected_files[0])
                try:
                    os.startfile(first_dir)
                except Exception:
                    pass

            # Khôi phục trạng thái giao diện
            self.is_running = False
            self.run_btn.config(state='normal', text="▶  BẮT ĐẦU XỬ LÝ")
            self.status_lbl.config(text="Đã hoàn tất xử lý.", fg=self.SUCCESS)

        # Chạy thread ngầm để không đơ GUI
        Thread(target=worker, daemon=True).start()


# ===================================================================
# KÍCH HOẠT ỨNG DỤNG
# ===================================================================
if __name__ == '__main__':
    # Đăng ký DPI để giao diện hiển thị sắc nét trên màn hình Windows độ phân giải cao
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass

    app = App()
    app.mainloop()
