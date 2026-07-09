# -*- coding: utf-8 -*-
"""
Tool Tổng Quan Chuỗi - Phân Tích Tồn Kho (Phiên bản v1.0.0)
Tác giả: Antigravity AI
Mục tiêu: Đọc các file dữ liệu chi tiết *(done).xlsx, tổng hợp dữ liệu (gộp) 
          và xuất file Dashboard HTML Tổng quan chuỗi tối ưu dung lượng.
"""

import os
import sys
import time
import json
import openpyxl
import ctypes
from datetime import datetime
from threading import Thread

# Import Tkinter for GUI
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

APP_TITLE = "Tool Tổng Quan Chuỗi - Phân Tích Tồn Kho"
APP_VERSION = "1.0.0"

# Thêm libs vào path nếu chạy trong môi trường Portable
current_dir = os.path.dirname(os.path.abspath(__file__))
libs_path = os.path.join(current_dir, "libs")
if os.path.exists(libs_path) and libs_path not in sys.path:
    sys.path.insert(0, libs_path)

# Mặc định tìm template trong cùng thư mục
DEFAULT_TEMPLATE_NAME = "dashboard_tong_quan_template.html"


def read_and_aggregate_files(file_paths, log_func=None) -> tuple[list[dict], datetime | None, bool]:
    """
    Đọc nhiều file Excel *(done).xlsx và cộng gộp số liệu.
    Trả về: (aggregated_rows_list, milestone_date, has_wm)
    """
    all_data = []
    milestone_date = None
    has_wm = False
    
    total_files = len(file_paths)
    for idx, path in enumerate(file_paths):
        filename = os.path.basename(path)
        if log_func:
            log_func(f"  [{idx+1}/{total_files}] Đang đọc dữ liệu từ: {filename}...")
            
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            # Ưu tiên sheet 'Dữ liệu kiểm tra'
            sheet_name = 'Dữ liệu kiểm tra'
            if sheet_name not in wb.sheetnames:
                sheet_name = wb.sheetnames[0]
            sheet = wb[sheet_name]
            
            # Đọc dòng tiêu đề (quét tối đa 5 dòng đầu)
            header_row = None
            for row in sheet.iter_rows(max_rows=5, values_only=True):
                if row and any(h in row for h in ['RSM', 'Store Code', 'Phân loại kiểm tra']):
                    header_row = row
                    break
            
            if not header_row:
                # Fallback lấy dòng 1
                for row in sheet.iter_rows(max_rows=1, values_only=True):
                    header_row = row
                    break
                    
            if not header_row:
                if log_func:
                    log_func(f"  ✗ Thất bại: Không tìm thấy tiêu đề cột trong file {filename}", "error")
                wb.close()
                continue
                
            # Bản đồ chỉ số cột dựa trên tiêu đề
            header_map = {}
            for c_idx, h in enumerate(header_row):
                if h is not None:
                    header_map[str(h).strip()] = c_idx
            
            # Lấy vị trí cột
            idx_rsm = header_map.get('RSM', 0)
            idx_asm = header_map.get('ASM', 1)
            idx_store_code = header_map.get('Store Code', 2)
            idx_store_name = header_map.get('Store Name', 3)
            idx_mch2_desc = header_map.get('MCH2 Desc', 5)
            idx_qty = header_map.get('Tồn cuối kỳ (D)', 14)
            idx_val = header_map.get('Giá trị tồn', 15)
            idx_class = header_map.get('Phân loại kiểm tra', 26)
            
            row_count = 0
            start_reading = False
            
            for row in sheet.iter_rows(values_only=True):
                if not start_reading:
                    if row == header_row:
                        start_reading = True
                    continue
                
                if not row or all(v is None for v in row):
                    continue
                
                # Trích xuất giá trị
                rsm = str(row[idx_rsm]).strip() if idx_rsm < len(row) and row[idx_rsm] is not None else ''
                asm = str(row[idx_asm]).strip() if idx_asm < len(row) and row[idx_asm] is not None else ''
                store_code = str(row[idx_store_code]).strip() if idx_store_code < len(row) and row[idx_store_code] is not None else ''
                store_name = str(row[idx_store_name]).strip() if idx_store_name < len(row) and row[idx_store_name] is not None else ''
                mch2 = str(row[idx_mch2_desc]).strip() if idx_mch2_desc < len(row) and row[idx_mch2_desc] is not None else ''
                check_class = str(row[idx_class]).strip() if idx_class < len(row) and row[idx_class] is not None else ''
                
                # Qty và Value
                try:
                    qty = float(row[idx_qty]) if idx_qty < len(row) and row[idx_qty] is not None else 0.0
                except (ValueError, TypeError):
                    qty = 0.0
                    
                try:
                    val = float(row[idx_val]) if idx_val < len(row) and row[idx_val] is not None else 0.0
                except (ValueError, TypeError):
                    val = 0.0
                
                # Bỏ qua dòng trống, N/A hoặc không có phân loại kiểm tra
                if not check_class or check_class.upper() in ('N/A', 'NONE', ''):
                    continue
                
                # Check siêu thị WM (mã cửa hàng chỉ gồm số hoặc là WM)
                if not has_wm and (store_code.isdigit() or 'WM' in store_code.upper()):
                    has_wm = True
                
                all_data.append({
                    'rsm': rsm,
                    'asm': asm,
                    'store_code': store_code,
                    'store_name': store_name,
                    'mch2': mch2,
                    'qty': qty,
                    'val': val,
                    'check_class': check_class
                })
                row_count += 1
                
            if log_func:
                log_func(f"  ✓ Đã nạp {row_count:,} dòng hợp lệ.")
            wb.close()
            
        except Exception as e:
            if log_func:
                log_func(f"  ✗ Lỗi khi đọc file {filename}: {e}", "error")
                
    if not all_data:
        return [], None, False
        
    # Tiến hành Groupby + Sum
    if log_func:
        log_func(f"\n💡 Đang gộp dữ liệu...")
        
    aggregated = {}
    for item in all_data:
        key = (
            item['rsm'],
            item['asm'],
            item['store_code'],
            item['store_name'],
            item['mch2'],
            item['check_class']
        )
        if key not in aggregated:
            aggregated[key] = {'qty': 0.0, 'val': 0.0, 'count': 0}
        aggregated[key]['qty'] += item['qty']
        aggregated[key]['val'] += item['val']
        aggregated[key]['count'] += 1
        
    # Chuyển đổi về list dict
    result_rows = []
    for key, metrics in aggregated.items():
        result_rows.append({
            'rsm': key[0],
            'asm': key[1],
            'store_code': key[2],
            'store_name': key[3],
            'mch2': key[4],
            'check_class': key[5],
            'qty': metrics['qty'],
            'val': metrics['val'],
            'count': metrics['count']
        })
        
    if log_func:
        log_func(f"  ✓ Gộp dữ liệu hoàn tất. Giảm từ {len(all_data):,} xuống {len(result_rows):,} dòng tổng hợp.")
        
    return result_rows, milestone_date, has_wm


def generate_overview_html(out_html_path: str, rows_data: list, has_wm: bool, 
                           log_func=None, target_date=None) -> bool:
    """
    Sinh file Dashboard HTML Tổng quan từ dữ liệu cộng gộp.
    """
    try:
        t0 = time.time()
        # Tìm đường dẫn template
        script_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(script_dir, DEFAULT_TEMPLATE_NAME)
        
        if not os.path.exists(template_path):
            template_path = r"D:\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu\dashboard_tong_quan_template.html"
            
        if not os.path.exists(template_path):
            # Thử tìm trong thư mục portable
            template_path = os.path.join(script_dir, "Tool_PhanTichTonKho_Portable", DEFAULT_TEMPLATE_NAME)
            
        if not os.path.exists(template_path):
            if log_func:
                log_func(f"  ✗ Lỗi: Không tìm thấy file template '{DEFAULT_TEMPLATE_NAME}'", "error")
            return False
            
        # Thu thập các danh sách duy nhất để mapping nén dữ liệu
        rsm_set = set()
        asm_set = set()
        store_map = {} # store_code: store_name
        mch2_set = set()
        class_set = set()
        
        for r in rows_data:
            rsm_set.add(r['rsm'])
            if r['asm']:
                asm_set.add(r['asm'])
            store_map[r['store_code']] = r['store_name']
            rsm_set.add(r['rsm'])
            mch2_set.add(r['mch2'])
            class_set.add(r['check_class'])
            
        rsm_list = sorted(list(rsm_set))
        asm_list = sorted(list(asm_set))
        store_list = [[code, name] for code, name in sorted(store_map.items())]
        mch2_list = sorted(list(mch2_set))
        class_list = sorted(list(class_set))
        
        # Tạo map index để nén
        rsm_idx_map = {name: idx for idx, name in enumerate(rsm_list)}
        asm_idx_map = {name: idx for idx, name in enumerate(asm_list)}
        store_idx_map = {code: idx for idx, [code, name] in enumerate(store_list)}
        mch2_idx_map = {name: idx for idx, name in enumerate(mch2_list)}
        class_idx_map = {name: idx for idx, name in enumerate(class_list)}
        
        compressed_rows = []
        for r in rows_data:
            store_idx = store_idx_map.get(r['store_code'], -1)
            rsm_idx = rsm_idx_map.get(r['rsm'], -1)
            asm_idx = asm_idx_map.get(r['asm'], -1)
            mch2_idx = mch2_idx_map.get(r['mch2'], -1)
            class_idx = class_idx_map.get(r['check_class'], -1)
            
            comp_row = [
                store_idx,
                rsm_idx,
                asm_idx,
                mch2_idx,
                class_idx,
                round(r['qty']),
                int(r['val']),
                r['count']
            ]
            compressed_rows.append(comp_row)
            
        out_dict = {
            "rsm_list": rsm_list,
            "asm_list": asm_list,
            "store_list": store_list,
            "mch2_list": mch2_list,
            "class_list": class_list,
            "has_wm": has_wm,
            "rows": compressed_rows
        }
        
        # Đọc nội dung template
        with open(template_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
            
        # Thay thế tiêu đề
        title = "Tổng Quan Chuỗi"
        html_content = html_content.replace(
            "Cập nhật: 01/07/2026",
            f"Cập nhật: {target_date or datetime.now().strftime('%d/%m/%Y')}"
        )
        
        # Tạo JS injection
        js_injection = f"const dbData = {json.dumps(out_dict, ensure_ascii=False, separators=(',', ':'))};\n"
        js_injection += "const rawData = dbData.rows.map(r => ({\n"
        js_injection += "    store_idx: r[0],\n"
        js_injection += "    store_code: dbData.store_list[r[0]][0],\n"
        js_injection += "    store_name: dbData.store_list[r[0]][1],\n"
        js_injection += "    rsm: dbData.rsm_list[r[1]],\n"
        js_injection += "    rsm_idx: r[1],\n"
        js_injection += "    asm: r[2] !== -1 ? dbData.asm_list[r[2]] : '',\n"
        js_injection += "    asm_idx: r[2],\n"
        js_injection += "    mch2: dbData.mch2_list[r[3]],\n"
        js_injection += "    mch2_idx: r[3],\n"
        js_injection += "    check_class: dbData.class_list[r[4]],\n"
        js_injection += "    class_idx: r[4],\n"
        js_injection += "    qty: r[5],\n"
        js_injection += "    value: r[6],\n"
        js_injection += "    count: r[7]\n"
        js_injection += "}));"
        
        html_content = html_content.replace('// DATA_PLACEHOLDER', js_injection)
        
        with open(out_html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
            
        elapsed = time.time() - t0
        if log_func:
            log_func(f"  ✓ Đã sinh file HTML: {os.path.basename(out_html_path)} | Thời gian: {elapsed:.1f}s")
        return True
        
    except Exception as e:
        if log_func:
            log_func(f"  ✗ Thất bại khi sinh file HTML: {e}", "error")
        return False


class App(tk.Tk):
    BG_DARK = '#1a1d2e'
    BG_CARD = '#252840'
    BG_INPUT = '#1e2235'
    ACCENT = '#6366f1'  # Indigo cho tông quan
    ACCENT_HOVER = '#818cf8'
    SUCCESS = '#10b981'
    WARNING = '#f59e0b'
    ERROR = '#ef4444'
    TEXT = '#e8eaed'
    TEXT_DIM = '#9aa0b8'
    BORDER = '#353a5e'

    def __init__(self):
        super().__init__()
        self.title(f"{APP_TITLE} v{APP_VERSION}")
        self.geometry("900x650")
        self.minsize(750, 550)
        self.configure(bg=self.BG_DARK)

        self.selected_files = []
        self.target_date_str = tk.StringVar(value=datetime.now().strftime('%d/%m/%Y'))
        self.is_running = False

        self._build_ui()
        self._center_window()

    def _center_window(self):
        self.update_idletasks()
        w = self.winfo_width()
        h = self.winfo_height()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _build_ui(self):
        # === Header ===
        header = tk.Frame(self, bg='#111322', pady=12)
        header.pack(fill='x')

        tk.Label(
            header, text="📈  Tool Tổng Quan Chuỗi - Báo Cáo Đối Soát",
            font=('Segoe UI', 15, 'bold'), bg='#111322', fg=self.TEXT
        ).pack(side='left', padx=20)

        tk.Label(
            header, text=f"v{APP_VERSION}", font=('Segoe UI', 9),
            bg='#111322', fg=self.TEXT_DIM
        ).pack(side='right', padx=20)

        # === Footer ===
        footer = tk.Frame(self, bg='#111322', pady=10)
        footer.pack(fill='x', side='bottom')

        self.run_btn = tk.Button(
            footer,
            text="📊  BẮT ĐẦU TỔNG HỢP",
            font=('Segoe UI', 11, 'bold'),
            bg=self.ACCENT, fg='white',
            activebackground=self.ACCENT_HOVER,
            activeforeground='white',
            cursor='hand2',
            relief='flat',
            padx=25, pady=8,
            command=self._on_run
        )
        self.run_btn.pack(side='left', padx=20)

        self.status_lbl = tk.Label(
            footer, text="Sẵn sàng.",
            font=('Segoe UI', 10), bg='#111322', fg=self.TEXT_DIM
        )
        self.status_lbl.pack(side='left', padx=10)

        # === Main Panel ===
        content = tk.Frame(self, bg=self.BG_DARK)
        content.pack(fill='both', expand=True, padx=15, pady=10)

        # Trái: Cấu hình
        left = tk.Frame(content, bg=self.BG_DARK)
        left.pack(side='left', fill='y', padx=(0, 10))

        # Phải: Log
        right = tk.Frame(content, bg=self.BG_DARK)
        right.pack(side='right', fill='both', expand=True)

        self._build_config_panel(left)
        self._build_log_panel(right)

    def _card(self, parent, title: str) -> tk.Frame:
        outer = tk.Frame(parent, bg=self.BG_CARD)
        outer.pack(fill='x', pady=5)
        tk.Label(
            outer, text=title, font=('Segoe UI', 9, 'bold'),
            bg=self.BG_CARD, fg=self.ACCENT, anchor='w', padx=12, pady=6
        ).pack(fill='x')
        tk.Frame(outer, bg=self.BORDER, height=1).pack(fill='x', padx=12)
        inner = tk.Frame(outer, bg=self.BG_CARD, padx=12, pady=10)
        inner.pack(fill='x')
        return inner

    def _build_config_panel(self, parent):
        parent.configure(width=320)

        # Card 1: Chọn files đầu vào
        c1 = self._card(parent, "1. File dữ liệu chi tiết *(done).xlsx")
        
        self.file_listbox = tk.Listbox(
            c1, font=('Segoe UI', 9), bg=self.BG_INPUT, fg=self.TEXT,
            selectbackground=self.ACCENT, selectforeground='white',
            height=8, relief='flat', bd=0,
            highlightthickness=1, highlightcolor=self.BORDER,
            highlightbackground=self.BORDER
        )
        self.file_listbox.pack(fill='x', pady=(0, 6))

        btn_frame = tk.Frame(c1, bg=self.BG_CARD)
        btn_frame.pack(fill='x')
        
        self._btn(btn_frame, "📂 Thêm file", self._add_files).pack(side='left')
        self._btn(btn_frame, "✖ Xóa", self._remove_selected, color='#d32f2f').pack(side='left', padx=4)
        self._btn(btn_frame, "🗑 Xóa hết", self._clear_files, color='#4b5563').pack(side='left')

        self.file_count_lbl = tk.Label(
            c1, text="Chưa chọn file nào.",
            font=('Segoe UI', 8), bg=self.BG_CARD, fg=self.TEXT_DIM, anchor='w'
        )
        self.file_count_lbl.pack(fill='x', pady=(4, 0))

        # Card 2: Cấu hình tham số
        c2 = self._card(parent, "2. Tham số báo cáo")
        
        frame = tk.Frame(c2, bg=self.BG_CARD)
        frame.pack(fill='x', pady=3)
        tk.Label(
            frame, text="Ngày cập nhật (hiển thị):", font=('Segoe UI', 9),
            bg=self.BG_CARD, fg=self.TEXT, anchor='w', width=22
        ).pack(side='left')
        
        self.date_entry = tk.Entry(
            frame, textvariable=self.target_date_str, font=('Segoe UI', 9),
            bg=self.BG_INPUT, fg=self.TEXT, relief='flat',
            insertbackground=self.TEXT,
            highlightthickness=1, highlightcolor=self.ACCENT,
            highlightbackground=self.BORDER
        )
        self.date_entry.pack(side='left', fill='x', expand=True)
        
        tk.Label(
            c2, text="Định dạng: DD/MM/YYYY (Ví dụ: 01/07/2026)", font=('Segoe UI', 7),
            bg=self.BG_CARD, fg=self.TEXT_DIM, anchor='e'
        ).pack(fill='x', pady=(2, 0))

        # Card 3: Ghi chú
        c3 = self._card(parent, "💡 Lưu ý")
        tk.Label(
            c3, text="• Nguồn dữ liệu đầu vào là các file kết quả\n  đã xử lý có đuôi *(done).xlsx\n• Tool sẽ tự động gộp dữ liệu theo RSM/ASM/\n  Cửa hàng/Ngành hàng và loại bỏ bảng\n  chi tiết sản phẩm để tối ưu hóa dung lượng.",
            font=('Segoe UI', 8), bg=self.BG_CARD, fg=self.TEXT_DIM, justify='left', anchor='w'
        ).pack(fill='x')

    def _build_log_panel(self, parent):
        tk.Label(
            parent, text="Nhật ký xử lý",
            font=('Segoe UI', 10, 'bold'), bg=self.BG_DARK, fg=self.TEXT, anchor='w'
        ).pack(fill='x', pady=(0, 4))

        log_frame = tk.Frame(parent, bg=self.BG_CARD)
        log_frame.pack(fill='both', expand=True)

        self.log_text = tk.Text(
            log_frame, font=('Consolas', 9), bg=self.BG_INPUT, fg=self.TEXT,
            relief='flat', bd=0, insertbackground=self.TEXT,
            highlightthickness=0, state='disabled'
        )
        self.log_text.pack(side='left', fill='both', expand=True, padx=(1, 0), pady=1)

        scrollbar = ttk.Scrollbar(log_frame, orient='vertical', command=self.log_text.yview)
        scrollbar.pack(side='right', fill='y')
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self.log_text.tag_config('info', foreground=self.TEXT)
        self.log_text.tag_config('success', foreground=self.SUCCESS)
        self.log_text.tag_config('warning', foreground=self.WARNING)
        self.log_text.tag_config('error', foreground=self.ERROR)
        self.log_text.tag_config('highlight', foreground=self.ACCENT, font=('Consolas', 9, 'bold'))

    def _btn(self, parent, text: str, cmd, color=None) -> tk.Button:
        bg_color = color if color else self.ACCENT
        active_bg = self.ACCENT_HOVER if not color else color
        btn = tk.Button(
            parent, text=text, font=('Segoe UI', 8, 'bold'),
            bg=bg_color, fg='white', activebackground=active_bg,
            activeforeground='white', relief='flat', cursor='hand2',
            padx=8, pady=4, command=cmd
        )
        return btn

    def log(self, message: str, tag: str = 'info'):
        self.log_text.config(state='normal')
        self.log_text.insert(tk.END, message + "\n", tag)
        self.log_text.see(tk.END)
        self.log_text.config(state='disabled')
        self.update_idletasks()

    def _add_files(self):
        files = filedialog.askopenfilenames(
            title="Chọn file dữ liệu chi tiết *(done).xlsx",
            filetypes=[("Excel Done Files", "*(done).xlsx"), ("Excel Files", "*.xlsx")]
        )
        for f in files:
            norm_f = os.path.abspath(f)
            if norm_f not in self.selected_files:
                self.selected_files.append(norm_f)
                self.file_listbox.insert(tk.END, os.path.basename(norm_f))
        self._update_file_count()

    def _remove_selected(self):
        selected_indices = list(self.file_listbox.curselection())
        for idx in reversed(selected_indices):
            self.selected_files.pop(idx)
            self.file_listbox.delete(idx)
        self._update_file_count()

    def _clear_files(self):
        self.selected_files.clear()
        self.file_listbox.delete(0, tk.END)
        self._update_file_count()

    def _update_file_count(self):
        count = len(self.selected_files)
        if count == 0:
            self.file_count_lbl.config(text="Chưa chọn file nào.", fg=self.TEXT_DIM)
        else:
            self.file_count_lbl.config(text=f"Đã chọn {count} file.", fg=self.SUCCESS)

    def _on_run(self):
        if self.is_running:
            return
            
        if not self.selected_files:
            messagebox.showerror("Lỗi", "Vui lòng chọn ít nhất 1 file dữ liệu chi tiết *(done).xlsx!")
            return

        date_str = self.target_date_str.get().strip()
        if date_str:
            try:
                # Kiểm tra định dạng ngày
                datetime.strptime(date_str, '%d/%m/%Y')
            except ValueError:
                messagebox.showerror("Lỗi", "Ngày cập nhật phải đúng định dạng DD/MM/YYYY!\nVí dụ: 01/07/2026")
                return

        self.is_running = True
        self.run_btn.config(state='disabled', text="⏳  ĐANG XỬ LÝ...")
        self.status_lbl.config(text="Đang tổng hợp dữ liệu...", fg=self.WARNING)
        
        self.log_text.config(state='normal')
        self.log_text.delete('1.0', tk.END)
        self.log_text.config(state='disabled')

        def worker():
            t_start = time.time()
            self.log(f"=== BẮT ĐẦU TỔNG HỢP TỒN KHO TOÀN CHUỖI ===", 'highlight')
            self.log(f"Thời gian chạy: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
            self.log(f"Ngày hiển thị trên Dashboard: {date_str if date_str else 'Hôm nay'}\n")
            
            # 1. Đọc và gộp dữ liệu
            aggregated_rows, milestone_date, has_wm = read_and_aggregate_files(
                self.selected_files,
                log_func=lambda msg, t='info': self.log(msg, t)
            )
            
            # 2. Sinh file html
            if aggregated_rows:
                # File đầu ra nằm cùng thư mục file đầu tiên
                out_dir = os.path.dirname(self.selected_files[0])
                out_html_path = os.path.join(out_dir, "Bao_cao_tong_quan_chuoi_dashboard.html")
                
                success = generate_overview_html(
                    out_html_path,
                    aggregated_rows,
                    has_wm,
                    log_func=lambda msg, t='info': self.log(msg, t),
                    target_date=date_str
                )
                
                elapsed = time.time() - t_start
                self.log(f"\n=== TIẾN TRÌNH HOÀN TẤT ===", 'highlight')
                if success:
                    self.log(f"Tổng hợp thành công từ {len(self.selected_files)} file.", 'success')
                    self.log(f"Đường dẫn file đầu ra: {out_html_path}", 'success')
                    self.log(f"Tổng thời gian thực hiện: {elapsed:.1f} giây.\n", 'success')
                    
                    try:
                        os.startfile(out_dir)
                    except Exception:
                        pass
                else:
                    self.log("✗ Lỗi sinh Dashboard HTML.", 'error')
            else:
                self.log("\n✗ Không có dữ liệu hợp lệ để tổng hợp. Vui lòng kiểm tra lại định dạng file đầu vào!", 'error')

            self.is_running = False
            self.run_btn.config(state='normal', text="📊  BẮT ĐẦU TỔNG HỢP")
            self.status_lbl.config(text="Đã hoàn tất.", fg=self.SUCCESS)

        Thread(target=worker, daemon=True).start()


if __name__ == '__main__':
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass

    app = App()
    app.mainloop()
