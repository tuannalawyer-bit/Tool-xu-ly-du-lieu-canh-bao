import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

folder = r"c:\Users\Tuan\OneDrive - WIN\Desktop\công việc thực hiện\Thang 7\Đổ dữ liệu"
files = [f for f in os.listdir(folder) if f.endswith(".xlsx")]

for f in files[:2]:
    path = os.path.join(folder, f)
    print(f"File: {f}")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        sheet = wb['Chi tiết data']
        # Let's read the first few rows
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= 5:
                break
            # Print index and Column AF value
            val = row[31] if len(row) > 31 else "N/A"
            print(f"Row {i} Col AF: '{val}' (Type: {type(val)})")
    except Exception as e:
        print(f"Error: {e}")
